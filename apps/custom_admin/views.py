"""
Views del Panel Administrativo Personalizado
Solo sirven HTML, la lógica está en JavaScript con JWT
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import ensure_csrf_cookie
from django.http import JsonResponse
from django.utils import timezone
from django.contrib import messages
from functools import wraps
from apps.authentication.models import Usuario
from apps.inventory_management.models import Marca,Proveedor
from functools import wraps
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import ensure_csrf_cookie
import logging 
import json
from django.http import JsonResponse
from django.db.models import Sum, Q
from apps.sales_management.models import Venta, DetalleVenta, Devolucion,Cliente
from django.db import IntegrityError
from django.core.paginator import Paginator
from apps.financial_management.models import (
    CuentaPorCobrar, 
    PagoCuentaPorCobrar,
    CuentaPorPagar,
    PagoCuentaPorPagar
)
from apps.system_configuration.models import (
    ConfiguracionSistema,
    ParametroSistema,
    LogConfiguracion,
    HealthCheck,
    get_parametro,
    set_parametro
)

logger = logging.getLogger(__name__)

# ========================================
# DECORATOR DE AUTENTICACIÓN
# ========================================


def get_authenticated_user(request):
    """Helper temporal para obtener usuario autenticado"""
    if request.user.is_authenticated:
        return request.user
    # Temporal: usar usuario edison
    from apps.authentication.models import Usuario
    return Usuario.objects.filter(username='edison').first()

def auth_required(view_func):
    """Decorator para vistas que requieren autenticación."""
    @wraps(view_func)
    def wrapper_function(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/login/')
        return view_func(request, *args, **kwargs)
    return wrapper_function

from django.views.decorators.cache import never_cache
from django.shortcuts import redirect

@ensure_csrf_cookie
@never_cache
def login_page_view(request):
    """Renderiza la página HTML de login - SIN autenticación requerida"""
    
    # ✅ SIEMPRE limpiar sesión al llegar a login
    request.session.flush()
    
    response = render(request, 'authentication/login.html')
    
    # ✅ Limpiar cookies JWT para evitar bucle
    response.delete_cookie('access_token')
    response.delete_cookie('refresh_token')
    
    # ✅ Headers anti-cache
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response


@ensure_csrf_cookie
@auth_required
def dashboard_view(request):
    """Dashboard principal del panel admin"""
    return render(request, 'custom_admin/dashboard.html')


# ========================================
# GESTIÓN DE USUARIOS
# ========================================

@ensure_csrf_cookie
@auth_required
def usuarios_view(request):
    """Lista de usuarios"""
    return render(request, 'custom_admin/usuarios/usuarios.html')


# ========================================
# ROLES Y PERMISOS
# ========================================

@ensure_csrf_cookie
@auth_required
def roles_view(request):
    """Lista de roles disponibles"""
    return render(request, 'custom_admin/roles/list.html')


# ========================================
# INVENTARIO
# ========================================

@ensure_csrf_cookie
@auth_required
def inventario_dashboard_view(request):
    """Dashboard de inventario"""
    return render(request, 'custom_admin/inventario/dashboard.html')


@ensure_csrf_cookie
@auth_required
def productos_view(request):
    """Lista de productos con datos reales - VERSIÓN ACTUALIZADA"""
    from apps.inventory_management.models import Producto, Categoria, Marca, Quintal, UnidadMedida
    from apps.system_configuration.models import ConfiguracionSistema
    from django.db.models import Q, Prefetch
    from django.core.paginator import Paginator
    
    productos = Producto.objects.select_related(
        'categoria',
        'unidad_medida_base', 
        'marca',
        'usuario_registro'
    ).prefetch_related(
        Prefetch(
            'quintales',
            queryset=Quintal.objects.select_related('proveedor', 'unidad_medida').filter(
                estado='DISPONIBLE'
            ).order_by('-fecha_ingreso')
        ),
        'inventario_normal'
    ).filter(activo=True).order_by('nombre')
    
    # Filtros
    search = request.GET.get('search', '')
    categoria_id = request.GET.get('categoria', '')
    tipo = request.GET.get('tipo', '')
    
    if search:
        productos = productos.filter(
            Q(nombre__icontains=search) | 
            Q(codigo_barras__icontains=search) |
            Q(descripcion__icontains=search)
        )
    
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    
    if tipo:
        productos = productos.filter(tipo_inventario=tipo)
    
    categorias = Categoria.objects.filter(activa=True).order_by('nombre')
    marcas = Marca.objects.filter(activa=True).order_by('nombre')
    unidades_medida = UnidadMedida.objects.filter(activa=True).order_by('orden_display')
    
    # ✅ OBTENER IVA DEFAULT DEL SISTEMA
    config = ConfiguracionSistema.get_config()
    iva_activo = config.iva_activo if config else False
    porcentaje_iva = config.porcentaje_iva if config else 0
    
    # Paginación
    paginator = Paginator(productos, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'productos': page_obj,
        'page_obj': page_obj,
        'marcas': marcas,
        'categorias': categorias,
        'unidades_medida': unidades_medida,
        'search': search,
        'categoria_selected': categoria_id,
        'tipo_selected': tipo,
        'iva_activo': iva_activo,
        'porcentaje_iva': porcentaje_iva  # ✅ Pasamos el porcentaje al template
    }
    
    return render(request, 'custom_admin/inventario/productos_list.html', context)

@ensure_csrf_cookie
@auth_required
def producto_detail_view(request, pk):
    """Detalle de producto"""
    return render(request, 'custom_admin/inventario/producto_detail.html', {'producto_id': pk})

@ensure_csrf_cookie
@auth_required
def producto_crear(request):
    """Crear nuevo producto - SISTEMA IVA SIMPLIFICADO"""
    from apps.inventory_management.models import Producto, ProductoNormal
    from apps.authentication.models import Usuario
    from apps.system_configuration.models import ConfiguracionSistema
    from decimal import Decimal
    
    if request.method == 'POST':
        try:
            usuario = Usuario.objects.filter(rol__codigo='ADMIN').first()
            if not usuario:
                usuario = Usuario.objects.first()
            if not usuario:
                messages.error(request, 'No hay usuarios en el sistema.')
                return redirect('custom_admin:productos')
            
            # Datos básicos
            nombre = request.POST.get('nombre', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            categoria_id = request.POST.get('categoria', '').strip()
            marca_id = request.POST.get('marca', '').strip()
            codigo_barras = request.POST.get('codigo_barras', '').strip()
            tipo_inventario = request.POST.get('tipo_inventario', 'NORMAL')
            activo = request.POST.get('activo') == 'on'
            
            # ✅ NUEVO: IVA simplificado - solo checkbox
            aplica_iva = request.POST.get('aplica_iva') == 'on'
            
            if not nombre:
                messages.error(request, 'El nombre del producto es obligatorio')
                return redirect('custom_admin:productos')
            
            producto_data = {
                'nombre': nombre,
                'descripcion': descripcion,
                'tipo_inventario': tipo_inventario,
                'activo': activo,
                'usuario_registro': usuario,
            }
            
            # ✅ Obtener IVA de la configuración si aplica
            if aplica_iva:
                config = ConfiguracionSistema.get_config()
                producto_data['iva'] = config.porcentaje_iva
            else:
                producto_data['iva'] = Decimal('0.00')
            
            if categoria_id:
                producto_data['categoria_id'] = categoria_id
            if marca_id:
                producto_data['marca_id'] = marca_id
            if codigo_barras:
                producto_data['codigo_barras'] = codigo_barras
            if 'imagen' in request.FILES:
                producto_data['imagen'] = request.FILES['imagen']
            
            # Precios según tipo
            if tipo_inventario == 'QUINTAL':
                precio_peso = request.POST.get('precio_por_unidad_peso', '0').strip()
                unidad_medida_id = request.POST.get('unidad_medida_base', '').strip()
                
                producto_data['precio_por_unidad_peso'] = Decimal(precio_peso) if precio_peso else Decimal('0.00')
                producto_data['precio_unitario'] = None
                
                if unidad_medida_id:
                    producto_data['unidad_medida_base_id'] = unidad_medida_id
                else:
                    precio_unit = request.POST.get('precio_unitario', '0').strip()  # El input sigue siendo precio_unitario
                    producto_data['precio_venta'] = Decimal(precio_unit) if precio_unit else Decimal('0.00')  # Pero se guarda como precio_venta
                    producto_data['precio_por_unidad_peso'] = None
            
            # Crear producto
            producto = Producto.objects.create(**producto_data)
            
            # Stock inicial para productos normales
            if tipo_inventario == 'NORMAL':
                stock_inicial = request.POST.get('stock_actual', '0').strip()
                ProductoNormal.objects.create(
                    producto=producto,
                    stock_actual=int(stock_inicial) if stock_inicial else 0,
                    stock_minimo=10,
                    stock_maximo=1000,
                    costo_unitario=Decimal('0.00')
                )
            
            messages.success(request, f'✅ Producto "{producto.nombre}" creado exitosamente')
            
        except ValueError as e:
            messages.error(request, f'Error en los datos numéricos: {str(e)}')
        except Exception as e:
            import traceback
            print(f"❌ Error: {traceback.format_exc()}")
            messages.error(request, f'Error al crear producto: {str(e)}')
    
    return redirect('custom_admin:productos')


@ensure_csrf_cookie
@auth_required
def producto_editar(request, producto_id):
    """Editar producto - SISTEMA IVA SIMPLIFICADO"""
    from apps.inventory_management.models import Producto, ProductoNormal, Quintal
    from apps.authentication.models import Usuario
    from apps.system_configuration.models import ConfiguracionSistema
    from decimal import Decimal
    
    try:
        producto = Producto.objects.get(id=producto_id)
        
        if request.method == 'POST':
            usuario = Usuario.objects.filter(rol__codigo='ADMIN').first()
            if not usuario:
                usuario = Usuario.objects.first()
            if not usuario:
                messages.error(request, 'No hay usuarios en el sistema.')
                return redirect('custom_admin:productos')
            
            # Datos básicos
            nombre = request.POST.get('nombre', '').strip()
            if not nombre:
                messages.error(request, 'El nombre del producto es obligatorio')
                return redirect('custom_admin:productos')
            
            producto.nombre = nombre
            producto.descripcion = request.POST.get('descripcion', '').strip()
            
            # Categoría y marca
            categoria_id = request.POST.get('categoria', '').strip()
            producto.categoria_id = categoria_id if categoria_id else None
            
            marca_id = request.POST.get('marca', '').strip()
            producto.marca_id = marca_id if marca_id else None
            
            codigo_barras = request.POST.get('codigo_barras', '').strip()
            if codigo_barras:
                producto.codigo_barras = codigo_barras
            
            # ✅ IVA simplificado
            aplica_iva = request.POST.get('aplica_iva') == 'on'
            if aplica_iva:
                config = ConfiguracionSistema.get_config()
                producto.iva = config.porcentaje_iva
            else:
                producto.iva = Decimal('0.00')
            
            print(f"🔧 IVA actualizado: {producto.iva}")
            
            # Imagen
            if 'imagen' in request.FILES:
                producto.imagen = request.FILES['imagen']
            
            # Tipo de inventario y precios
            tipo_inventario = request.POST.get('tipo_inventario', 'NORMAL')
            producto.tipo_inventario = tipo_inventario
            
            if tipo_inventario == 'QUINTAL':
                precio_peso = request.POST.get('precio_por_unidad_peso', '0').strip()
                producto.precio_por_unidad_peso = Decimal(precio_peso) if precio_peso else Decimal('0.00')
                producto.precio_unitario = None
                
                unidad_medida_id = request.POST.get('unidad_medida_base', '').strip()
                if unidad_medida_id:
                    producto.unidad_medida_base_id = unidad_medida_id
            else:
                precio_unit = request.POST.get('precio_unitario', '0').strip()
                producto.precio_unitario = Decimal(precio_unit) if precio_unit else Decimal('0.00')
                producto.precio_por_unidad_peso = None
            
            producto.activo = request.POST.get('activo') == 'on'
            
            if hasattr(producto, 'usuario_modificacion'):
                producto.usuario_modificacion = usuario
            
            # Guardar producto
            producto.save()
            
            # ✅ Actualizar stock
            if tipo_inventario == 'NORMAL':
                stock_actual_str = request.POST.get('stock_actual', '0').strip()
                
                try:
                    stock_actual = int(stock_actual_str) if stock_actual_str else 0
                    
                    try:
                        producto_normal = producto.inventario_normal
                        producto_normal.stock_actual = stock_actual
                        producto_normal.save()
                        print(f"✅ Stock actualizado: {stock_actual}")
                    except ProductoNormal.DoesNotExist:
                        ProductoNormal.objects.create(
                            producto=producto,
                            stock_actual=stock_actual,
                            stock_minimo=10,
                            stock_maximo=1000,
                            costo_unitario=Decimal('0.00')
                        )
                        print(f"✅ ProductoNormal creado con stock: {stock_actual}")
                
                except ValueError as e:
                    print(f"❌ Error al convertir stock: {e}")
                    messages.warning(request, 'El stock debe ser un número entero')
            
            messages.success(request, f'✅ Producto "{producto.nombre}" actualizado exitosamente')
            return redirect('custom_admin:productos')
            
    except Producto.DoesNotExist:
        messages.error(request, '❌ Producto no encontrado')
        return redirect('custom_admin:productos')
    except ValueError as e:
        messages.error(request, f'Error en los datos numéricos: {str(e)}')
        return redirect('custom_admin:productos')
    except Exception as e:
        import traceback
        print(f"❌ ERROR: {traceback.format_exc()}")
        messages.error(request, f'Error al editar producto: {str(e)}')
    
    return redirect('custom_admin:productos')


@ensure_csrf_cookie
@auth_required
def producto_eliminar(request, producto_id):
    """Eliminar o desactivar producto"""
    from apps.inventory_management.models import Producto
    from django.http import HttpResponseRedirect
    from django.db.models import ProtectedError
    
    try:
        producto = Producto.objects.get(id=producto_id)
        nombre = producto.nombre
        
        # ✅ Verificar todas las relaciones de forma segura
        tiene_relaciones = False
        relaciones_encontradas = []
        
        # Verificar detalles de venta (múltiples nombres posibles)
        for attr in ['detalleventa_set', 'detalles_venta', 'ventas']:
            if hasattr(producto, attr):
                try:
                    if getattr(producto, attr).exists():
                        tiene_relaciones = True
                        relaciones_encontradas.append('ventas')
                        break
                except:
                    pass
        
        # Verificar detalles de compra (múltiples nombres posibles)
        for attr in ['detallecompra_set', 'detalles_compra', 'compras']:
            if hasattr(producto, attr):
                try:
                    if getattr(producto, attr).exists():
                        tiene_relaciones = True
                        relaciones_encontradas.append('compras')
                        break
                except:
                    pass
        
        # Verificar quintales
        if hasattr(producto, 'quintales'):
            try:
                if producto.quintales.exists():
                    tiene_relaciones = True
                    relaciones_encontradas.append('quintales')
            except:
                pass
        
        # Verificar inventario normal
        if hasattr(producto, 'inventario_normal'):
            try:
                inventario = producto.inventario_normal
                # Si tiene stock o movimientos, considerar que tiene relaciones
                if inventario.stock_actual > 0:
                    tiene_relaciones = True
                    relaciones_encontradas.append('inventario')
                    
                # Verificar movimientos
                if hasattr(inventario, 'movimientos') and inventario.movimientos.exists():
                    tiene_relaciones = True
                    if 'inventario' not in relaciones_encontradas:
                        relaciones_encontradas.append('movimientos')
            except:
                pass
        
        if tiene_relaciones:
            # Tiene relaciones, solo desactivar
            producto.activo = False
            producto.save()
            
            relaciones_texto = ', '.join(relaciones_encontradas)
            messages.warning(
                request, 
                f'El producto "{nombre}" tiene registros asociados ({relaciones_texto}) y no puede eliminarse. Se ha desactivado.'
            )
        else:
            # No tiene relaciones, intentar eliminar
            try:
                producto.delete()
                messages.success(request, f'✅ Producto "{nombre}" eliminado correctamente.')
            except ProtectedError as e:
                # Si aún así hay error de protección, desactivar
                producto.activo = False
                producto.save()
                messages.warning(
                    request, 
                    f'El producto "{nombre}" tiene registros protegidos y no puede eliminarse. Se ha desactivado.'
                )
            except Exception as e:
                # Cualquier otro error al eliminar
                producto.activo = False
                producto.save()
                messages.warning(
                    request, 
                    f'Error al eliminar el producto "{nombre}": {str(e)}. Se ha desactivado.'
                )
        
        return HttpResponseRedirect('/panel/inventario/productos/')
        
    except Producto.DoesNotExist:
        messages.error(request, '❌ Producto no encontrado.')
        return HttpResponseRedirect('/panel/inventario/productos/')
    except Exception as e:
        import traceback
        error_completo = traceback.format_exc()
        print("=" * 80)
        print("❌ ERROR AL ELIMINAR PRODUCTO:")
        print(error_completo)
        print("=" * 80)
        messages.error(request, f'❌ Error al procesar producto: {str(e)}')
        return HttpResponseRedirect('/panel/inventario/productos/')
@ensure_csrf_cookie
@auth_required
def quintales_view(request):
    """Lista de quintales"""
    return render(request, 'custom_admin/inventario/quintales_list.html')


@ensure_csrf_cookie
@auth_required
def quintal_detail_view(request, pk):
    """Detalle de quintal"""
    return render(request, 'custom_admin/inventario/quintal_detail.html', {'quintal_id': pk})


@ensure_csrf_cookie
@auth_required
def categorias_view(request):
    """Lista de categorías con datos reales"""
    from apps.inventory_management.models import Categoria
    from django.db.models import Q
    from django.core.paginator import Paginator
    
    categorias = Categoria.objects.all().order_by('orden', 'nombre')
    
    search = request.GET.get('search', '')
    if search:
        categorias = categorias.filter(
            Q(nombre__icontains=search) | 
            Q(descripcion__icontains=search)
        )
    
    paginator = Paginator(categorias, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'categorias': page_obj,
        'page_obj': page_obj,
        'search': search,
    }
    
    return render(request, 'custom_admin/inventario/categorias_list.html', context)


@ensure_csrf_cookie
@auth_required
def categoria_crear_view(request):
    """Crear categoría"""
    from apps.inventory_management.models import Categoria
    
    if request.method == 'POST':
        try:
            categoria = Categoria.objects.create(
                nombre=request.POST.get('nombre'),
                descripcion=request.POST.get('descripcion', ''),
                margen_ganancia_sugerido=request.POST.get('margen_ganancia_sugerido', 30.00),
                descuento_maximo_permitido=request.POST.get('descuento_maximo_permitido', 10.00),
                orden=request.POST.get('orden', 0),
                activa=request.POST.get('activa') == 'on'
            )
            messages.success(request, f'Categoría "{categoria.nombre}" creada exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al crear la categoría: {str(e)}')
    
    return redirect('custom_admin:categorias')


@ensure_csrf_cookie
@auth_required
def categoria_editar_view(request, pk):
    """Editar categoría"""
    from apps.inventory_management.models import Categoria
    
    categoria = get_object_or_404(Categoria, pk=pk)
    
    if request.method == 'POST':
        try:
            categoria.nombre = request.POST.get('nombre')
            categoria.descripcion = request.POST.get('descripcion', '')
            categoria.margen_ganancia_sugerido = request.POST.get('margen_ganancia_sugerido', 30.00)
            categoria.descuento_maximo_permitido = request.POST.get('descuento_maximo_permitido', 10.00)
            categoria.orden = request.POST.get('orden', 0)
            categoria.activa = request.POST.get('activa') == 'on'
            categoria.save()
            messages.success(request, f'Categoría "{categoria.nombre}" actualizada exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al actualizar la categoría: {str(e)}')
    
    return redirect('custom_admin:categorias')


@ensure_csrf_cookie
@auth_required
def categoria_eliminar_view(request, pk):
    """Eliminar categoría"""
    from apps.inventory_management.models import Categoria
    
    categoria = get_object_or_404(Categoria, pk=pk)
    
    if request.method == 'POST':
        nombre = categoria.nombre
        try:
            if categoria.productos.exists():
                messages.error(request, f'No se puede eliminar la categoría "{nombre}" porque tiene productos asociados.')
            else:
                categoria.delete()
                messages.success(request, f'Categoría "{nombre}" eliminada exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar la categoría: {str(e)}')
    
    return redirect('custom_admin:categorias')


# ============================================================================
# VISTAS DE MARCAS - ✅ VERSIÓN CORREGIDA
# ============================================================================

@ensure_csrf_cookie
@auth_required
def marcas_list(request):
    """Lista todas las marcas de productos con filtros y paginación"""
    from apps.inventory_management.models import Marca
    from django.core.paginator import Paginator
    from django.db.models import Q, Count
    
    marcas = Marca.objects.annotate(
        total_productos=Count('productos', distinct=True),
        productos_con_stock=Count(
            'productos',
            filter=Q(
                productos__activo=True,
                productos__tipo_inventario='NORMAL',
                productos__inventario_normal__stock_actual__gt=0
            ),
            distinct=True
        )
    ).order_by('orden', 'nombre')
    
    # Filtros
    search = request.GET.get('search', '')
    destacada_selected = request.GET.get('destacada', '')
    pais_selected = request.GET.get('pais', '')
    
    if search:
        marcas = marcas.filter(
            Q(nombre__icontains=search) |
            Q(descripcion__icontains=search) |
            Q(fabricante__icontains=search)
        )
    
    if destacada_selected:
        marcas = marcas.filter(destacada=True)
    
    if pais_selected:
        marcas = marcas.filter(pais_origen=pais_selected)
    
    # Obtener lista de países únicos para el filtro
    paises = Marca.objects.exclude(
        pais_origen__isnull=True
    ).exclude(
        pais_origen=''
    ).values_list('pais_origen', flat=True).distinct().order_by('pais_origen')
    
    # Paginación
    paginator = Paginator(marcas, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'marcas': page_obj,
        'page_obj': page_obj,
        'paises': paises,
        'search': search,
        'destacada_selected': destacada_selected,
        'pais_selected': pais_selected,
    }
    
    return render(request, 'custom_admin/inventario/marcas_list.html', context)


@ensure_csrf_cookie
@auth_required
def marca_crear(request):
    """Crear una nueva marca - VERSIÓN COMPLETA"""
    from apps.inventory_management.models import Marca
    
    if request.method == 'POST':
        try:
            # ✅ Obtener todos los campos del formulario
            nombre = request.POST.get('nombre', '').strip()
            descripcion = request.POST.get('descripcion', '').strip()
            pais_origen = request.POST.get('pais_origen', '').strip()
            fabricante = request.POST.get('fabricante', '').strip()
            sitio_web = request.POST.get('sitio_web', '').strip()
            orden = request.POST.get('orden', '0')
            
            # ✅ Campos booleanos (checkbox)
            activa = request.POST.get('activa') == 'on'
            destacada = request.POST.get('destacada') == 'on'
            
            # ✅ Archivo de logo (si existe)
            logo = request.FILES.get('logo')
            
            # Validación básica
            if not nombre:
                messages.error(request, '❌ El nombre de la marca es obligatorio.')
                return redirect('custom_admin:marcas')
            
            # Verificar si ya existe
            if Marca.objects.filter(nombre__iexact=nombre).exists():
                messages.error(request, f'❌ Ya existe una marca con el nombre "{nombre}".')
                return redirect('custom_admin:marcas')
            
            # ✅ Crear la marca con TODOS los campos
            marca_data = {
                'nombre': nombre,
                'descripcion': descripcion,
                'pais_origen': pais_origen,
                'fabricante': fabricante,
                'sitio_web': sitio_web,
                'activa': activa,
                'destacada': destacada,
                'orden': int(orden) if orden else 0,
            }
            
            # ✅ Agregar logo si existe
            if logo:
                marca_data['logo'] = logo
            
            marca = Marca.objects.create(**marca_data)
            
            messages.success(
                request, 
                f'✅ Marca "{marca.nombre}" creada exitosamente.'
            )
            
        except Exception as e:
            messages.error(request, f'❌ Error al crear la marca: {str(e)}')
    
    return redirect('custom_admin:marcas')


@ensure_csrf_cookie
@auth_required
def marca_editar(request, pk):
    """Editar una marca existente - VERSIÓN COMPLETA"""
    from apps.inventory_management.models import Marca
    
    marca = get_object_or_404(Marca, pk=pk)
    
    if request.method == 'POST':
        try:
            # ✅ Actualizar todos los campos
            nombre = request.POST.get('nombre', '').strip()
            
            if not nombre:
                messages.error(request, '❌ El nombre de la marca es obligatorio.')
                return redirect('custom_admin:marcas')
            
            # Verificar duplicados (excluyendo la marca actual)
            if Marca.objects.filter(nombre__iexact=nombre).exclude(pk=pk).exists():
                messages.error(request, f'❌ Ya existe otra marca con el nombre "{nombre}".')
                return redirect('custom_admin:marcas')
            
            # ✅ Actualizar campos
            marca.nombre = nombre
            marca.descripcion = request.POST.get('descripcion', '').strip()
            marca.pais_origen = request.POST.get('pais_origen', '').strip()
            marca.fabricante = request.POST.get('fabricante', '').strip()
            marca.sitio_web = request.POST.get('sitio_web', '').strip()
            marca.orden = int(request.POST.get('orden', '0'))
            marca.activa = request.POST.get('activa') == 'on'
            marca.destacada = request.POST.get('destacada') == 'on'
            
            # ✅ Actualizar logo si se subió uno nuevo
            logo = request.FILES.get('logo')
            if logo:
                # Eliminar logo anterior si existe
                if marca.logo:
                    marca.logo.delete(save=False)
                marca.logo = logo
            
            marca.save()
            
            messages.success(
                request, 
                f'✅ Marca "{marca.nombre}" actualizada exitosamente.'
            )
            
        except Exception as e:
            messages.error(request, f'❌ Error al actualizar la marca: {str(e)}')
    
    return redirect('custom_admin:marcas')


@ensure_csrf_cookie
@auth_required
def marca_eliminar(request, pk):
    """Eliminar una marca (o desactivar si tiene productos)"""
    from apps.inventory_management.models import Marca
    
    marca = get_object_or_404(Marca, pk=pk)
    nombre = marca.nombre
    
    if request.method == 'POST':
        try:
            # Verificar si tiene productos asociados
            if marca.productos.exists():
                # Tiene productos, solo desactivar
                marca.activa = False
                marca.save()
                messages.warning(
                    request,
                    f'⚠️ La marca "{nombre}" tiene productos asociados y no puede eliminarse. Se ha desactivado.'
                )
            else:
                # No tiene productos, eliminar completamente
                marca.delete()
                messages.success(
                    request,
                    f'✅ Marca "{nombre}" eliminada exitosamente.'
                )
        except Exception as e:
            messages.error(request, f'❌ Error al eliminar la marca: {str(e)}')
    
    return redirect('custom_admin:marcas')

# ============================================================================
# PROVEEDORES - MÓDULO COMPLETO CORREGIDO
# ============================================================================

@ensure_csrf_cookie
@auth_required
def proveedores_view(request):
    """Lista de proveedores con datos reales"""
    from apps.inventory_management.models import Proveedor
    from django.db.models import Q, Count
    from django.core.paginator import Paginator
    
    proveedores = Proveedor.objects.annotate(
        total_productos=Count('productos', distinct=True)
    ).order_by('nombre_comercial')
    
    # Filtros
    search = request.GET.get('search', '')
    activo_selected = request.GET.get('activo', '')
    
    if search:
        proveedores = proveedores.filter(
            Q(nombre_comercial__icontains=search) |
            Q(razon_social__icontains=search) |
            Q(ruc_nit__icontains=search) |
            Q(email__icontains=search)
        )
    
    if activo_selected:
        proveedores = proveedores.filter(activo=(activo_selected == 'true'))
    
    # Paginación
    paginator = Paginator(proveedores, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'proveedores': page_obj,
        'page_obj': page_obj,
        'search': search,
        'activo_selected': activo_selected,
    }
    
    return render(request, 'custom_admin/inventario/proveedores_list.html', context)


@ensure_csrf_cookie
@auth_required
def proveedor_crear(request):
    """Crear un nuevo proveedor"""
    from apps.inventory_management.models import Proveedor
    from decimal import Decimal
    
    if request.method == 'POST':
        try:
            nombre_comercial = request.POST.get('nombre_comercial', '').strip()
            ruc_nit = request.POST.get('ruc_nit', '').strip()
            
            if not nombre_comercial:
                messages.error(request, '❌ El nombre comercial es obligatorio.')
                return redirect('custom_admin:proveedores')
            
            if not ruc_nit:
                messages.error(request, '❌ El RUC/NIT es obligatorio.')
                return redirect('custom_admin:proveedores')
            
            # Verificar si ya existe por nombre
            if Proveedor.objects.filter(nombre_comercial__iexact=nombre_comercial).exists():
                messages.error(request, f'❌ Ya existe un proveedor con el nombre "{nombre_comercial}".')
                return redirect('custom_admin:proveedores')
            
            # Verificar si ya existe por RUC
            if Proveedor.objects.filter(ruc_nit=ruc_nit).exists():
                messages.error(request, f'❌ Ya existe un proveedor con el RUC/NIT "{ruc_nit}".')
                return redirect('custom_admin:proveedores')
            
            # Preparar datos
            dias_credito = request.POST.get('dias_credito', '0').strip()
            limite_credito = request.POST.get('limite_credito', '0').strip()
            
            proveedor = Proveedor.objects.create(
                nombre_comercial=nombre_comercial,
                razon_social=request.POST.get('razon_social', '').strip(),
                ruc_nit=ruc_nit,
                direccion=request.POST.get('direccion', '').strip(),
                telefono=request.POST.get('telefono', '').strip(),
                email=request.POST.get('email', '').strip(),
                dias_credito=int(dias_credito) if dias_credito else 0,
                limite_credito=Decimal(limite_credito) if limite_credito else Decimal('0.00'),
                activo=request.POST.get('activo') == 'on'
            )
            
            messages.success(request, f'✅ Proveedor "{proveedor.nombre_comercial}" creado exitosamente.')
            
        except ValueError as e:
            messages.error(request, f'❌ Error en los datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'❌ Error al crear el proveedor: {str(e)}')
            import traceback
            traceback.print_exc()
    
    return redirect('custom_admin:proveedores')


@ensure_csrf_cookie
@auth_required
def proveedor_editar(request, pk):
    """Editar un proveedor existente"""
    from apps.inventory_management.models import Proveedor
    from decimal import Decimal
    
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    if request.method == 'POST':
        try:
            nombre_comercial = request.POST.get('nombre_comercial', '').strip()
            ruc_nit = request.POST.get('ruc_nit', '').strip()
            
            if not nombre_comercial:
                messages.error(request, '❌ El nombre comercial es obligatorio.')
                return redirect('custom_admin:proveedores')
            
            if not ruc_nit:
                messages.error(request, '❌ El RUC/NIT es obligatorio.')
                return redirect('custom_admin:proveedores')
            
            # Verificar duplicados por nombre (excluyendo el proveedor actual)
            if Proveedor.objects.filter(nombre_comercial__iexact=nombre_comercial).exclude(pk=pk).exists():
                messages.error(request, f'❌ Ya existe otro proveedor con el nombre "{nombre_comercial}".')
                return redirect('custom_admin:proveedores')
            
            # Verificar duplicados por RUC (excluyendo el proveedor actual)
            if Proveedor.objects.filter(ruc_nit=ruc_nit).exclude(pk=pk).exists():
                messages.error(request, f'❌ Ya existe otro proveedor con el RUC/NIT "{ruc_nit}".')
                return redirect('custom_admin:proveedores')
            
            # Actualizar campos
            proveedor.nombre_comercial = nombre_comercial
            proveedor.razon_social = request.POST.get('razon_social', '').strip()
            proveedor.ruc_nit = ruc_nit
            proveedor.direccion = request.POST.get('direccion', '').strip()
            proveedor.telefono = request.POST.get('telefono', '').strip()
            proveedor.email = request.POST.get('email', '').strip()
            
            dias_credito = request.POST.get('dias_credito', '0').strip()
            limite_credito = request.POST.get('limite_credito', '0').strip()
            
            proveedor.dias_credito = int(dias_credito) if dias_credito else 0
            proveedor.limite_credito = Decimal(limite_credito) if limite_credito else Decimal('0.00')
            proveedor.activo = request.POST.get('activo') == 'on'
            
            proveedor.save()
            
            messages.success(request, f'✅ Proveedor "{proveedor.nombre_comercial}" actualizado exitosamente.')
            
        except ValueError as e:
            messages.error(request, f'❌ Error en los datos: {str(e)}')
        except Exception as e:
            messages.error(request, f'❌ Error al actualizar el proveedor: {str(e)}')
            import traceback
            traceback.print_exc()
    
    return redirect('custom_admin:proveedores')


@ensure_csrf_cookie
@auth_required
def proveedor_eliminar(request, pk):
    """Eliminar un proveedor (o desactivar si tiene productos)"""
    from apps.inventory_management.models import Proveedor
    
    proveedor = get_object_or_404(Proveedor, pk=pk)
    nombre = proveedor.nombre_comercial
    
    if request.method == 'POST':
        try:
            # Verificar si tiene productos asociados
            if proveedor.productos.exists():
                # Tiene productos, solo desactivar
                proveedor.activo = False
                proveedor.save()
                messages.warning(
                    request,
                    f'⚠️ El proveedor "{nombre}" tiene {proveedor.productos.count()} productos asociados y no puede eliminarse. Se ha desactivado.'
                )
            else:
                # No tiene productos, eliminar completamente
                proveedor.delete()
                messages.success(request, f'✅ Proveedor "{nombre}" eliminado exitosamente.')
        except Exception as e:
            messages.error(request, f'❌ Error al eliminar el proveedor: {str(e)}')
            import traceback
            traceback.print_exc()
    
    return redirect('custom_admin:proveedores')


@ensure_csrf_cookie
@auth_required
def proveedor_detalle_api(request, pk):
    """API para obtener detalles de un proveedor (para edición)"""
    from apps.inventory_management.models import Proveedor
    
    try:
        proveedor = get_object_or_404(Proveedor, pk=pk)
        
        data = {
            'success': True,
            'proveedor': {
                'id': str(proveedor.id),
                'nombre_comercial': proveedor.nombre_comercial,
                'razon_social': proveedor.razon_social or '',
                'ruc_nit': proveedor.ruc_nit,
                'direccion': proveedor.direccion or '',
                'telefono': proveedor.telefono or '',
                'email': proveedor.email or '',
                'dias_credito': proveedor.dias_credito,
                'limite_credito': str(proveedor.limite_credito),
                'activo': proveedor.activo,
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

# ========================================
# VENTAS
# ========================================

@ensure_csrf_cookie
@auth_required
def ventas_dashboard_view(request):
    """Dashboard de ventas"""
    return render(request, 'custom_admin/ventas/dashboard.html')


@ensure_csrf_cookie
@auth_required
def ventas_view(request):
    """Historial de ventas con datos reales"""
    from apps.sales_management.models import Venta, Cliente
    from apps.authentication.models import Usuario
    from django.db.models import Q, Sum, Count
    from django.core.paginator import Paginator
    from decimal import Decimal
    
    ventas = Venta.objects.select_related('cliente', 'vendedor').all().order_by('-fecha_venta')
    
    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    estado = request.GET.get('estado', '')
    tipo_venta = request.GET.get('tipo_venta', '')
    vendedor_id = request.GET.get('vendedor', '')
    cliente_id = request.GET.get('cliente', '')
    
    if fecha_inicio:
        ventas = ventas.filter(fecha_venta__date__gte=fecha_inicio)
    
    if fecha_fin:
        ventas = ventas.filter(fecha_venta__date__lte=fecha_fin)
    
    if estado:
        ventas = ventas.filter(estado=estado)
    
    if tipo_venta:
        ventas = ventas.filter(tipo_venta=tipo_venta)
    
    if vendedor_id:
        ventas = ventas.filter(vendedor_id=vendedor_id)
    
    if cliente_id:
        ventas = ventas.filter(cliente_id=cliente_id)
    
    # Estadísticas
    stats = ventas.filter(estado='COMPLETADA').aggregate(
        total=Sum('total'),
        cantidad=Count('id')
    )
    
    total_ventas = stats['total'] or Decimal('0')
    count_ventas = stats['cantidad'] or 0
    
    # Ventas de HOY
    hoy = timezone.now().date()
    ventas_hoy = Venta.objects.filter(
        fecha_venta__date=hoy,
        estado='COMPLETADA'
    ).count()
    
    # Ventas del MES actual
    inicio_mes = timezone.now().replace(day=1).date()
    ventas_mes = Venta.objects.filter(
        fecha_venta__date__gte=inicio_mes,
        estado='COMPLETADA'
    ).count()
    
    # Paginación
    paginator = Paginator(ventas, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Datos para formulario
    vendedores = Usuario.objects.filter(
        rol__codigo__in=['VENDEDOR', 'ADMIN', 'SUPERVISOR']
    ).order_by('nombres')
    
    clientes = Cliente.objects.filter(activo=True).order_by('apellidos', 'nombres')[:100]
    
    context = {
        'ventas': page_obj,
        'page_obj': page_obj,
        'total_ventas': total_ventas,
        'count_ventas': count_ventas,
        'ventas_hoy': ventas_hoy,
        'ventas_mes': ventas_mes,
        'vendedores': vendedores,
        'clientes': clientes,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'estado_selected': estado,
        'tipo_venta_selected': tipo_venta,
        'vendedor_selected': vendedor_id,
        'cliente_selected': cliente_id,
        'estados': [
            ('', 'Todos'),
            ('PENDIENTE', 'Pendiente'),
            ('COMPLETADA', 'Completada'),
            ('ANULADA', 'Anulada'),
        ],
        'tipos_venta': [
            ('', 'Todos'),
            ('CONTADO', 'Contado'),
            ('CREDITO', 'Crédito'),
        ],
    }
    
    return render(request, 'custom_admin/ventas/list.html', context)


@ensure_csrf_cookie
@auth_required
def venta_detail_view(request, pk):
    """Detalle de venta con datos reales"""
    from apps.sales_management.models import Venta
    
    venta = get_object_or_404(Venta.objects.select_related('cliente', 'vendedor', 'caja'), pk=pk)
    detalles = venta.detalles.select_related('producto', 'quintal', 'unidad_medida').all()
    pagos = venta.pagos.select_related('usuario').all()
    devoluciones = venta.devoluciones.select_related('usuario_solicita', 'usuario_aprueba').all()
    
    context = {
        'venta': venta,
        'detalles': detalles,
        'pagos': pagos,
        'devoluciones': devoluciones,
    }
    
    return render(request, 'custom_admin/ventas/detail.html', context)


@ensure_csrf_cookie
@auth_required
def venta_anular_view(request, pk):
    """Anular una venta"""
    from apps.sales_management.models import Venta
    from apps.authentication.models import Usuario
    
    if request.method == 'POST':
        try:
            venta = get_object_or_404(Venta, pk=pk)
            
            if venta.estado == 'ANULADA':
                messages.error(request, '❌ La venta ya está anulada.')
                return redirect('custom_admin:venta_detail', pk=pk)
            
            if venta.monto_pagado > 0:
                messages.error(request, '❌ No se puede anular una venta con pagos registrados.')
                return redirect('custom_admin:venta_detail', pk=pk)
            
            from apps.sales_management.services.pos_service import POSService
            usuario = Usuario.objects.filter(rol__codigo='ADMIN').first()
            
            if not usuario:
                usuario = Usuario.objects.first()
            
            POSService.anular_venta(venta, usuario)
            
            messages.success(request, f'✅ Venta {venta.numero_venta} anulada exitosamente.')
            
        except Exception as e:
            messages.error(request, f'❌ Error al anular venta: {str(e)}')
    
    return redirect('custom_admin:ventas_list')


@ensure_csrf_cookie
@auth_required
def venta_ticket_view(request, pk):
    """Imprimir ticket de venta"""
    from apps.sales_management.models import Venta
    from django.http import HttpResponse
    
    venta = get_object_or_404(Venta.objects.select_related('cliente', 'vendedor'), pk=pk)
    detalles = venta.detalles.select_related('producto', 'unidad_medida').all()
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Ticket - {venta.numero_venta}</title>
        <style>
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{
                font-family: 'Courier New', monospace;
                width: 80mm;
                margin: 0 auto;
                padding: 10px;
                font-size: 12px;
            }}
            .header {{
                text-align: center;
                margin-bottom: 15px;
                border-bottom: 2px dashed #000;
                padding-bottom: 10px;
            }}
            .header h1 {{ font-size: 18px; margin-bottom: 5px; }}
            .header p {{ font-size: 10px; margin: 2px 0; }}
            .info {{
                margin-bottom: 15px;
                border-bottom: 1px dashed #000;
                padding-bottom: 10px;
            }}
            .info-row {{
                display: flex;
                justify-content: space-between;
                margin: 3px 0;
            }}
            .items {{ margin-bottom: 15px; }}
            .item {{
                margin: 5px 0;
                padding: 5px 0;
                border-bottom: 1px dotted #ccc;
            }}
            .item-name {{
                font-weight: bold;
                margin-bottom: 3px;
            }}
            .item-details {{
                display: flex;
                justify-content: space-between;
                font-size: 11px;
            }}
            .totals {{
                border-top: 2px solid #000;
                padding-top: 10px;
                margin-top: 10px;
            }}
            .total-row {{
                display: flex;
                justify-content: space-between;
                margin: 3px 0;
            }}
            .total-row.final {{
                font-size: 16px;
                font-weight: bold;
                margin-top: 10px;
                padding-top: 10px;
                border-top: 2px dashed #000;
            }}
            .footer {{
                text-align: center;
                margin-top: 20px;
                padding-top: 10px;
                border-top: 2px dashed #000;
                font-size: 10px;
            }}
            @media print {{ body {{ width: 80mm; }} }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🛒 CommerceBox</h1>
            <p>Sistema de Gestión</p>
            <p>──────────────────</p>
        </div>
        
        <div class="info">
            <div class="info-row">
                <span><strong>N° Venta:</strong></span>
                <span>{venta.numero_venta}</span>
            </div>
            <div class="info-row">
                <span><strong>Fecha:</strong></span>
                <span>{venta.fecha_venta.strftime('%d/%m/%Y %H:%M')}</span>
            </div>
            <div class="info-row">
                <span><strong>Vendedor:</strong></span>
                <span>{venta.vendedor.get_full_name()}</span>
            </div>
            {"" if not venta.cliente else f'''
            <div class="info-row">
                <span><strong>Cliente:</strong></span>
                <span>{venta.cliente.nombres} {venta.cliente.apellidos}</span>
            </div>
            <div class="info-row">
                <span><strong>Documento:</strong></span>
                <span>{venta.cliente.numero_documento}</span>
            </div>
            '''}
            <div class="info-row">
                <span><strong>Tipo:</strong></span>
                <span>{venta.get_tipo_venta_display()}</span>
            </div>
        </div>
        
        <div class="items">
            <h3 style="margin-bottom: 10px;">PRODUCTOS</h3>
    """
    
    for detalle in detalles:
        if detalle.cantidad_unidades:
            cantidad = f"{detalle.cantidad_unidades} und"
            precio = f"${float(detalle.precio_unitario):,.2f}"
        else:
            cantidad = f"{float(detalle.peso_vendido):,.3f} {detalle.unidad_medida.abreviatura if detalle.unidad_medida else 'kg'}"
            precio = f"${float(detalle.precio_por_unidad_peso):,.2f}/{detalle.unidad_medida.abreviatura if detalle.unidad_medida else 'kg'}"
        
        html += f"""
            <div class="item">
                <div class="item-name">{detalle.producto.nombre}</div>
                <div class="item-details">
                    <span>{cantidad} × {precio}</span>
                    <span>${float(detalle.total):,.2f}</span>
                </div>
            </div>
        """
    
    html += f"""
        </div>
        
        <div class="totals">
            <div class="total-row">
                <span>Subtotal:</span>
                <span>${float(venta.subtotal):,.2f}</span>
            </div>
            {"" if venta.descuento == 0 else f'''
            <div class="total-row">
                <span>Descuento:</span>
                <span>-${float(venta.descuento):,.2f}</span>
            </div>
            '''}
            {"" if venta.impuestos == 0 else f'''
            <div class="total-row">
                <span>Impuestos:</span>
                <span>${float(venta.impuestos):,.2f}</span>
            </div>
            '''}
            <div class="total-row final">
                <span>TOTAL:</span>
                <span>${float(venta.total):,.2f}</span>
            </div>
            {"" if venta.tipo_venta != 'CONTADO' else f'''
            <div class="total-row">
                <span>Pagado:</span>
                <span>${float(venta.monto_pagado):,.2f}</span>
            </div>
            <div class="total-row">
                <span>Cambio:</span>
                <span>${float(venta.cambio):,.2f}</span>
            </div>
            '''}
        </div>
        
        <div class="footer">
            <p>¡Gracias por su compra!</p>
            <p>──────────────────</p>
            <p>Este documento no tiene validez fiscal</p>
        </div>
        
        <script>
            window.onload = function() {{
                window.print();
            }};
        </script>
    </body>
    </html>
    """
    
    return HttpResponse(html, content_type='text/html')


@ensure_csrf_cookie
@auth_required
def venta_factura_view(request, pk):
    """Generar factura de venta"""
    from apps.sales_management.models import Venta
    from django.http import HttpResponse
    
    venta = get_object_or_404(Venta.objects.select_related('cliente', 'vendedor'), pk=pk)
    
    if not venta.cliente:
        messages.error(request, 'No se puede generar factura para ventas sin cliente.')
        return redirect('custom_admin:venta_detail', pk=pk)
    
    detalles = venta.detalles.select_related('producto', 'unidad_medida').all()
    
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Factura - {venta.numero_factura or venta.numero_venta}</title>
        <style>
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{
                font-family: 'Arial', sans-serif;
                padding: 40px;
                color: #333;
            }}
            .factura {{
                max-width: 800px;
                margin: 0 auto;
                border: 2px solid #3b82f6;
                padding: 30px;
            }}
            .header {{
                display: flex;
                justify-content: space-between;
                align-items: start;
                border-bottom: 3px solid #3b82f6;
                padding-bottom: 20px;
                margin-bottom: 30px;
            }}
            .logo {{
                font-size: 32px;
                font-weight: bold;
                color: #3b82f6;
            }}
            .factura-info {{
                text-align: right;
            }}
            .factura-numero {{
                font-size: 24px;
                font-weight: bold;
                color: #3b82f6;
                margin-bottom: 5px;
            }}
            .info-grid {{
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 30px;
                margin-bottom: 30px;
            }}
            .info-section {{
                background: #f8fafc;
                padding: 15px;
                border-radius: 8px;
            }}
            .info-section h3 {{
                color: #3b82f6;
                margin-bottom: 10px;
                font-size: 14px;
                text-transform: uppercase;
            }}
            .info-section p {{
                margin: 5px 0;
                font-size: 13px;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
            }}
            thead {{
                background: #3b82f6;
                color: white;
            }}
            th {{
                padding: 12px;
                text-align: left;
                font-size: 13px;
                text-transform: uppercase;
            }}
            td {{
                padding: 10px 12px;
                border-bottom: 1px solid #e2e8f0;
                font-size: 13px;
            }}
            tbody tr:hover {{
                background: #f8fafc;
            }}
            .text-right {{
                text-align: right;
            }}
            .totals {{
                margin-top: 20px;
                display: flex;
                justify-content: flex-end;
            }}
            .totals-table {{
                width: 300px;
            }}
            .totals-table tr td {{
                padding: 8px 12px;
                border: none;
            }}
            .totals-table tr.total {{
                background: #3b82f6;
                color: white;
                font-size: 18px;
                font-weight: bold;
            }}
            .footer {{
                margin-top: 40px;
                padding-top: 20px;
                border-top: 2px solid #e2e8f0;
                text-align: center;
                color: #64748b;
                font-size: 12px;
            }}
            @media print {{
                body {{ padding: 0; }}
                .factura {{ border: none; }}
            }}
        </style>
    </head>
    <body>
        <div class="factura">
            <div class="header">
                <div>
                    <div class="logo">🛒 CommerceBox</div>
                    <p style="color: #64748b; margin-top: 5px;">Sistema de Gestión Comercial</p>
                </div>
                <div class="factura-info">
                    <div class="factura-numero">FACTURA</div>
                    <p><strong>N°:</strong> {venta.numero_factura or venta.numero_venta}</p>
                    <p><strong>Fecha:</strong> {venta.fecha_venta.strftime('%d/%m/%Y')}</p>
                </div>
            </div>
            
            <div class="info-grid">
                <div class="info-section">
                    <h3>📦 Información del Cliente</h3>
                    <p><strong>Cliente:</strong> {venta.cliente.nombres} {venta.cliente.apellidos}</p>
                    <p><strong>Documento:</strong> {venta.cliente.get_tipo_documento_display()} - {venta.cliente.numero_documento}</p>
                    {"" if not venta.cliente.telefono else f"<p><strong>Teléfono:</strong> {venta.cliente.telefono}</p>"}
                    {"" if not venta.cliente.email else f"<p><strong>Email:</strong> {venta.cliente.email}</p>"}
                    {"" if not venta.cliente.direccion else f"<p><strong>Dirección:</strong> {venta.cliente.direccion}</p>"}
                </div>
                
                <div class="info-section">
                    <h3>💼 Información de la Venta</h3>
                    <p><strong>Vendedor:</strong> {venta.vendedor.get_full_name()}</p>
                    <p><strong>Tipo de Venta:</strong> {venta.get_tipo_venta_display()}</p>
                    <p><strong>Estado:</strong> {venta.get_estado_display()}</p>
                    {"" if venta.tipo_venta != 'CREDITO' else f'''
                    <p><strong>Fecha Vencimiento:</strong> {venta.fecha_vencimiento.strftime('%d/%m/%Y') if venta.fecha_vencimiento else 'N/A'}</p>
                    '''}
                </div>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th style="width: 50%;">Producto</th>
                        <th class="text-right">Cantidad</th>
                        <th class="text-right">Precio Unit.</th>
                        <th class="text-right">Descuento</th>
                        <th class="text-right">Total</th>
                    </tr>
                </thead>
                <tbody>
    """
    
    for detalle in detalles:
        if detalle.cantidad_unidades:
            cantidad = f"{detalle.cantidad_unidades} und"
            precio = f"${float(detalle.precio_unitario):,.2f}"
        else:
            cantidad = f"{float(detalle.peso_vendido):,.3f} {detalle.unidad_medida.abreviatura if detalle.unidad_medida else 'kg'}"
            precio = f"${float(detalle.precio_por_unidad_peso):,.2f}"
        
        html += f"""
                    <tr>
                        <td><strong>{detalle.producto.nombre}</strong></td>
                        <td class="text-right">{cantidad}</td>
                        <td class="text-right">{precio}</td>
                        <td class="text-right">${float(detalle.descuento_monto):,.2f}</td>
                        <td class="text-right"><strong>${float(detalle.total):,.2f}</strong></td>
                    </tr>
        """
    
    html += f"""
                </tbody>
            </table>
            
            <div class="totals">
                <table class="totals-table">
                    <tr>
                        <td>Subtotal:</td>
                        <td class="text-right"><strong>${float(venta.subtotal):,.2f}</strong></td>
                    </tr>
                    {"" if venta.descuento == 0 else f'''
                    <tr>
                        <td>Descuento:</td>
                        <td class="text-right" style="color: #dc2626;"><strong>-${float(venta.descuento):,.2f}</strong></td>
                    </tr>
                    '''}
                    {"" if venta.impuestos == 0 else f'''
                    <tr>
                        <td>Impuestos:</td>
                        <td class="text-right"><strong>${float(venta.impuestos):,.2f}</strong></td>
                    </tr>
                    '''}
                    <tr class="total">
                        <td>TOTAL:</td>
                        <td class="text-right">${float(venta.total):,.2f}</td>
                    </tr>
                </table>
            </div>
            
            <div class="footer">
                <p>Gracias por su preferencia</p>
                <p style="margin-top: 10px;">Este documento es una representación impresa de una factura electrónica</p>
            </div>
        </div>
        
        <script>
            window.onload = function() {{
                window.print();
            }};
        </script>
    </body>
    </html>
    """
    
    return HttpResponse(html, content_type='text/html')


# ========================================
# NUEVAS VISTAS PARA EXPORTACIONES
# ========================================

@ensure_csrf_cookie
def venta_detalle_api(request, pk):
    """API endpoint para obtener detalles de una venta"""
    from apps.sales_management.models import Venta
    
    try:
        venta = get_object_or_404(Venta.objects.select_related('cliente', 'vendedor'), pk=pk)
        
        # Obtener items de la venta
        items = []
        for item in venta.detalles.select_related('producto', 'unidad_medida').all():
            items.append({
                'producto_nombre': item.producto.nombre,
                'codigo_barras': item.producto.codigo_barras or 'Sin código',
                'cantidad': str(item.cantidad_unidades) if item.cantidad_unidades else str(float(item.peso_vendido)),
                'precio_unitario': str(item.precio_unitario) if item.precio_unitario else str(float(item.precio_por_unidad_peso)),
                'descuento': str(item.descuento_monto),
                'subtotal': str(item.subtotal),
                'total': str(item.total),
            })
        
        # Construir respuesta
        data = {
            'success': True,
            'venta': {
                'numero_venta': venta.numero_venta,
                'fecha_venta': venta.fecha_venta.strftime('%d/%m/%Y %H:%M'),
                'estado': venta.get_estado_display(),
                'tipo_venta': venta.get_tipo_venta_display(),
                'vendedor': venta.vendedor.get_full_name(),
                'cliente': f"{venta.cliente.nombres} {venta.cliente.apellidos}" if venta.cliente else None,
                'metodo_pago': venta.get_metodo_pago_display() if hasattr(venta, 'metodo_pago') else 'N/A',
                'subtotal': str(venta.subtotal),
                'descuento': str(venta.descuento),
                'total': str(venta.total),
                'monto_pagado': str(venta.monto_pagado),
                'cambio': str(venta.cambio) if venta.cambio else '0.00',
                'saldo_pendiente': str(venta.saldo_pendiente),
                'items': items,
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@ensure_csrf_cookie
@auth_required
def exportar_venta_excel_individual(request, pk):
    """Exportar una venta individual a Excel"""
    from apps.sales_management.models import Venta
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    venta = get_object_or_404(Venta.objects.select_related('cliente', 'vendedor'), pk=pk)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Venta {venta.numero_venta}"
    
    # Estilos
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # TÍTULO
    ws.merge_cells('A1:F1')
    ws['A1'] = f"DETALLE DE VENTA - {venta.numero_venta}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # INFORMACIÓN GENERAL
    row = 3
    ws[f'A{row}'] = "Fecha:"
    ws[f'B{row}'] = venta.fecha_venta.strftime('%d/%m/%Y %H:%M')
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Cliente:"
    ws[f'B{row}'] = f"{venta.cliente.nombres} {venta.cliente.apellidos}" if venta.cliente else "Público General"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Vendedor:"
    ws[f'B{row}'] = venta.vendedor.get_full_name()
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Tipo de Venta:"
    ws[f'B{row}'] = venta.get_tipo_venta_display()
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Estado:"
    ws[f'B{row}'] = venta.get_estado_display()
    ws[f'A{row}'].font = Font(bold=True)
    
    # PRODUCTOS
    row += 2
    headers = ['Producto', 'Cantidad', 'Precio Unit.', 'Descuento', 'Subtotal', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Items
    for item in venta.detalles.select_related('producto', 'unidad_medida').all():
        row += 1
        ws.cell(row=row, column=1, value=item.producto.nombre).border = border
        
        if item.cantidad_unidades:
            ws.cell(row=row, column=2, value=float(item.cantidad_unidades)).border = border
            ws.cell(row=row, column=3, value=float(item.precio_unitario)).border = border
        else:
            ws.cell(row=row, column=2, value=float(item.peso_vendido)).border = border
            ws.cell(row=row, column=3, value=float(item.precio_por_unidad_peso)).border = border
        
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).number_format = '$#,##0.00'
        ws.cell(row=row, column=4, value=float(item.descuento_monto)).border = border
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=5, value=float(item.subtotal)).border = border
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        ws.cell(row=row, column=6, value=float(item.total)).border = border
        ws.cell(row=row, column=6).number_format = '$#,##0.00'
    
    # TOTALES
    row += 2
    ws.cell(row=row, column=5, value="Subtotal:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(venta.subtotal)).number_format = '$#,##0.00'
    
    row += 1
    ws.cell(row=row, column=5, value="Descuento:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(venta.descuento)).number_format = '$#,##0.00'
    
    row += 1
    ws.cell(row=row, column=5, value="TOTAL:").font = Font(bold=True, size=14)
    ws.cell(row=row, column=6, value=float(venta.total)).font = Font(bold=True, size=14)
    ws.cell(row=row, column=6).number_format = '$#,##0.00'
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Preparar respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=venta_{venta.numero_venta}.xlsx'
    
    return response


@ensure_csrf_cookie
@auth_required
def exportar_venta_pdf_individual(request, pk):
    """Exportar una venta individual a PDF"""
    from apps.sales_management.models import Venta
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO
    
    venta = get_object_or_404(Venta.objects.select_related('cliente', 'vendedor'), pk=pk)
    
    # Crear buffer
    buffer = BytesIO()
    
    # Crear documento
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo personalizado para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=30,
        alignment=1  # Centro
    )
    
    # TÍTULO
    title = Paragraph(f"DETALLE DE VENTA - {venta.numero_venta}", title_style)
    elements.append(title)
    
    # INFORMACIÓN GENERAL
    info_data = [
        ['Fecha:', venta.fecha_venta.strftime('%d/%m/%Y %H:%M')],
        ['Cliente:', f"{venta.cliente.nombres} {venta.cliente.apellidos}" if venta.cliente else "Público General"],
        ['Vendedor:', venta.vendedor.get_full_name()],
        ['Tipo de Venta:', venta.get_tipo_venta_display()],
        ['Estado:', venta.get_estado_display()],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e293b')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # PRODUCTOS
    productos_header = Paragraph('<b>PRODUCTOS</b>', styles['Heading2'])
    elements.append(productos_header)
    elements.append(Spacer(1, 10))
    
    # Tabla de productos
    productos_data = [['Producto', 'Cant.', 'Precio Unit.', 'Desc.', 'Subtotal', 'Total']]
    
    for item in venta.detalles.select_related('producto', 'unidad_medida').all():
        if item.cantidad_unidades:
            cantidad = str(item.cantidad_unidades)
            precio = f"${item.precio_unitario:.2f}"
        else:
            cantidad = f"{float(item.peso_vendido):.3f}"
            precio = f"${item.precio_por_unidad_peso:.2f}"
        
        productos_data.append([
            item.producto.nombre,
            cantidad,
            precio,
            f"${item.descuento_monto:.2f}",
            f"${item.subtotal:.2f}",
            f"${item.total:.2f}",
        ])
    
    productos_table = Table(productos_data, colWidths=[2.5*inch, 0.7*inch, 1*inch, 0.9*inch, 1*inch, 1*inch])
    productos_table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Body
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 9),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    elements.append(productos_table)
    elements.append(Spacer(1, 20))
    
    # TOTALES
    totales_data = [
        ['Subtotal:', f"${venta.subtotal:.2f}"],
        ['Descuento:', f"${venta.descuento:.2f}"],
        ['', ''],
        ['TOTAL:', f"${venta.total:.2f}"],
    ]
    
    totales_table = Table(totales_data, colWidths=[4.5*inch, 1.5*inch])
    totales_table.setStyle(TableStyle([
        ('FONT', (0, 0), (0, 1), 'Helvetica-Bold', 10),
        ('FONT', (1, 0), (1, 1), 'Helvetica', 10),
        ('FONT', (0, 3), (-1, 3), 'Helvetica-Bold', 14),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e293b')),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('LINEABOVE', (0, 3), (-1, 3), 2, colors.HexColor('#3B82F6')),
        ('TOPPADDING', (0, 3), (-1, 3), 10),
    ]))
    elements.append(totales_table)
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=venta_{venta.numero_venta}.pdf'
    
    return response


@ensure_csrf_cookie
@auth_required
def exportar_ventas_excel_general(request):
    """Exportar todas las ventas filtradas a Excel"""
    from apps.sales_management.models import Venta
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime
    from io import BytesIO
    
    # Obtener filtros de la URL
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    estado = request.GET.get('estado')
    tipo_venta = request.GET.get('tipo_venta')
    vendedor_id = request.GET.get('vendedor')
    cliente_id = request.GET.get('cliente')
    
    # Filtrar ventas
    ventas = Venta.objects.select_related('cliente', 'vendedor').all().order_by('-fecha_venta')
    
    if fecha_inicio:
        ventas = ventas.filter(fecha_venta__date__gte=fecha_inicio)
    if fecha_fin:
        ventas = ventas.filter(fecha_venta__date__lte=fecha_fin)
    if estado:
        ventas = ventas.filter(estado=estado)
    if tipo_venta:
        ventas = ventas.filter(tipo_venta=tipo_venta)
    if vendedor_id:
        ventas = ventas.filter(vendedor_id=vendedor_id)
    if cliente_id:
        ventas = ventas.filter(cliente_id=cliente_id)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    # Estilos
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # TÍTULO
    ws.merge_cells('A1:J1')
    ws['A1'] = f"REPORTE DE VENTAS - {datetime.now().strftime('%d/%m/%Y')}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # HEADERS
    headers = ['N° Venta', 'Fecha', 'Cliente', 'Vendedor', 'Tipo', 'Estado', 
               'Subtotal', 'Descuento', 'Total', 'Pagado']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # DATOS
    row = 4
    for venta in ventas:
        ws.cell(row=row, column=1, value=venta.numero_venta).border = border
        ws.cell(row=row, column=2, value=venta.fecha_venta.strftime('%d/%m/%Y %H:%M')).border = border
        ws.cell(row=row, column=3, value=f"{venta.cliente.nombres} {venta.cliente.apellidos}" if venta.cliente else "Público General").border = border
        ws.cell(row=row, column=4, value=venta.vendedor.get_full_name()).border = border
        ws.cell(row=row, column=5, value=venta.get_tipo_venta_display()).border = border
        ws.cell(row=row, column=6, value=venta.get_estado_display()).border = border
        
        ws.cell(row=row, column=7, value=float(venta.subtotal)).border = border
        ws.cell(row=row, column=7).number_format = '$#,##0.00'
        
        ws.cell(row=row, column=8, value=float(venta.descuento)).border = border
        ws.cell(row=row, column=8).number_format = '$#,##0.00'
        
        ws.cell(row=row, column=9, value=float(venta.total)).border = border
        ws.cell(row=row, column=9).number_format = '$#,##0.00'
        
        ws.cell(row=row, column=10, value=float(venta.monto_pagado)).border = border
        ws.cell(row=row, column=10).number_format = '$#,##0.00'
        
        row += 1
    
    # TOTALES
    row += 1
    ws.cell(row=row, column=8, value="TOTALES:").font = Font(bold=True)
    
    total_ventas = sum(float(v.total) for v in ventas)
    ws.cell(row=row, column=9, value=total_ventas).font = Font(bold=True)
    ws.cell(row=row, column=9).number_format = '$#,##0.00'
    
    total_pagado = sum(float(v.monto_pagado) for v in ventas)
    ws.cell(row=row, column=10, value=total_pagado).font = Font(bold=True)
    ws.cell(row=row, column=10).number_format = '$#,##0.00'
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    
    # Preparar respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


@ensure_csrf_cookie
@auth_required
def exportar_ventas_pdf_general(request):
    """Exportar todas las ventas filtradas a PDF"""
    from apps.sales_management.models import Venta
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    from datetime import datetime
    
    # Obtener filtros (mismo código que Excel)
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    estado = request.GET.get('estado')
    tipo_venta = request.GET.get('tipo_venta')
    vendedor_id = request.GET.get('vendedor')
    cliente_id = request.GET.get('cliente')
    
    # Filtrar ventas
    ventas = Venta.objects.select_related('cliente', 'vendedor').all().order_by('-fecha_venta')
    
    if fecha_inicio:
        ventas = ventas.filter(fecha_venta__date__gte=fecha_inicio)
    if fecha_fin:
        ventas = ventas.filter(fecha_venta__date__lte=fecha_fin)
    if estado:
        ventas = ventas.filter(estado=estado)
    if tipo_venta:
        ventas = ventas.filter(tipo_venta=tipo_venta)
    if vendedor_id:
        ventas = ventas.filter(vendedor_id=vendedor_id)
    if cliente_id:
        ventas = ventas.filter(cliente_id=cliente_id)
    
    # Crear buffer
    buffer = BytesIO()
    
    # Crear documento (landscape para más espacio)
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    title = Paragraph(f"REPORTE DE VENTAS - {datetime.now().strftime('%d/%m/%Y')}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 10))
    
    # Tabla de ventas
    data = [['N° Venta', 'Fecha', 'Cliente', 'Tipo', 'Total', 'Estado']]
    
    for venta in ventas:
        data.append([
            venta.numero_venta,
            venta.fecha_venta.strftime('%d/%m/%Y'),
            (f"{venta.cliente.nombres} {venta.cliente.apellidos}" if venta.cliente else "Público")[:20],
            venta.get_tipo_venta_display(),
            f"${venta.total:.2f}",
            venta.get_estado_display(),
        ])
    
    # Totales
    total = sum(float(v.total) for v in ventas)
    data.append(['', '', '', 'TOTAL:', f"${total:.2f}", ''])
    
    table = Table(data, colWidths=[1.2*inch, 1.1*inch, 1.8*inch, 1*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Body
        ('FONT', (0, 1), (-1, -2), 'Helvetica', 8),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
        
        # Total row
        ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 10),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#3B82F6')),
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    filename = f"ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


@ensure_csrf_cookie
@auth_required
def clientes_view(request):
    """Lista y gestión de clientes"""
    from apps.sales_management.models import Cliente
    from django.db.models import Q
    from django.core.paginator import Paginator
    
    # Obtener filtros
    search = request.GET.get('search', '')
    tipo_cliente = request.GET.get('tipo', '')
    estado = request.GET.get('estado', '')
    
    # Consulta base
    clientes_qs = Cliente.objects.all()
    
    # Aplicar filtros
    if search:
        clientes_qs = clientes_qs.filter(
            Q(nombres__icontains=search) | 
            Q(apellidos__icontains=search) |
            Q(numero_documento__icontains=search) |
            Q(nombre_comercial__icontains=search)
        )
    
    if tipo_cliente:
        clientes_qs = clientes_qs.filter(tipo_cliente=tipo_cliente)
    
    if estado:
        if estado == 'activo':
            clientes_qs = clientes_qs.filter(activo=True)
        elif estado == 'inactivo':
            clientes_qs = clientes_qs.filter(activo=False)
    
    clientes_qs = clientes_qs.order_by('apellidos', 'nombres')
    
    # Convertir a lista para el template
    clientes = list(clientes_qs)
    
    context = {
        'clientes': clientes,
        'total_clientes': Cliente.objects.count(),
        'clientes_activos': Cliente.objects.filter(activo=True).count(),
        'clientes_mayoristas': Cliente.objects.filter(tipo_cliente='MAYORISTA').count(),
        'clientes_frecuentes': Cliente.objects.filter(tipo_cliente='FRECUENTE').count(),
        'search': search,
        'tipo_selected': tipo_cliente,
        'estado_selected': estado,
    }
    
    return render(request, 'custom_admin/ventas/clientes_list.html', context)



@ensure_csrf_cookie
@auth_required
def cliente_crear(request):
    """Crear nuevo cliente"""
    from apps.sales_management.models import Cliente
    from django.contrib import messages
    
    if request.method == 'POST':
        try:
            cliente = Cliente.objects.create(
                tipo_documento=request.POST.get('tipo_documento'),
                numero_documento=request.POST.get('numero_documento'),
                nombres=request.POST.get('nombres'),
                apellidos=request.POST.get('apellidos'),
                nombre_comercial=request.POST.get('nombre_comercial', ''),
                telefono=request.POST.get('telefono', ''),
                email=request.POST.get('email', ''),
                direccion=request.POST.get('direccion', ''),
                tipo_cliente=request.POST.get('tipo_cliente', 'OCASIONAL'),
                limite_credito=request.POST.get('limite_credito', 0),
                dias_credito=request.POST.get('dias_credito', 0),
                descuento_general=request.POST.get('descuento_general', 0),
                activo=request.POST.get('activo') == 'on'
            )
            messages.success(request, f'Cliente {cliente.nombres} {cliente.apellidos} creado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al crear cliente: {str(e)}')
    
    return redirect('custom_admin:clientes')

@ensure_csrf_cookie
@auth_required
def cliente_editar(request, pk):
    """Editar cliente existente"""
    from apps.sales_management.models import Cliente
    from django.contrib import messages
    
    cliente = get_object_or_404(Cliente, pk=pk)
    
    if request.method == 'POST':
        try:
            cliente.tipo_documento = request.POST.get('tipo_documento')
            cliente.numero_documento = request.POST.get('numero_documento')
            cliente.nombres = request.POST.get('nombres')
            cliente.apellidos = request.POST.get('apellidos')
            cliente.nombre_comercial = request.POST.get('nombre_comercial', '')
            cliente.telefono = request.POST.get('telefono', '')
            cliente.email = request.POST.get('email', '')
            cliente.direccion = request.POST.get('direccion', '')
            cliente.tipo_cliente = request.POST.get('tipo_cliente', 'OCASIONAL')
            cliente.limite_credito = request.POST.get('limite_credito', 0)
            cliente.dias_credito = request.POST.get('dias_credito', 0)
            cliente.descuento_general = request.POST.get('descuento_general', 0)
            cliente.activo = request.POST.get('activo') == 'on'
            cliente.save()
            
            messages.success(request, f'Cliente {cliente.nombres} {cliente.apellidos} actualizado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al actualizar cliente: {str(e)}')
    
    return redirect('custom_admin:clientes')

@ensure_csrf_cookie
@auth_required
def cliente_eliminar(request, pk):
    """Eliminar cliente"""
    from apps.sales_management.models import Cliente
    from django.contrib import messages
    
    if request.method == 'POST':
        cliente = get_object_or_404(Cliente, pk=pk)
        nombre_completo = f"{cliente.nombres} {cliente.apellidos}"
        try:
            cliente.delete()
            messages.success(request, f'Cliente {nombre_completo} eliminado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al eliminar cliente: {str(e)}')
    
    return redirect('custom_admin:clientes')

def api_clientes_list(request):
    """API para listar clientes"""
    from apps.sales_management.models import Cliente
    
    clientes = Cliente.objects.all().order_by('apellidos', 'nombres')
    data = []
    
    for cliente in clientes:
        data.append({
            'id': str(cliente.id),
            'tipo_documento': cliente.tipo_documento,
            'numero_documento': cliente.numero_documento,
            'nombres': cliente.nombres,
            'apellidos': cliente.apellidos,
            'nombre_comercial': cliente.nombre_comercial,
            'telefono': cliente.telefono,
            'email': cliente.email,
            'direccion': cliente.direccion,
            'tipo_cliente': cliente.tipo_cliente,
            'limite_credito': float(cliente.limite_credito),
            'credito_disponible': float(cliente.credito_disponible),
            'dias_credito': cliente.dias_credito,
            'descuento_general': float(cliente.descuento_general),
            'total_compras': float(cliente.total_compras),
            'activo': cliente.activo,
        })
    
    return JsonResponse({'clientes': data})

def api_cliente_detail(request, pk):
    """API para obtener detalle de un cliente"""
    from apps.sales_management.models import Cliente
    
    cliente = get_object_or_404(Cliente, pk=pk)
    
    data = {
        'id': str(cliente.id),
        'tipo_documento': cliente.tipo_documento,
        'numero_documento': cliente.numero_documento,
        'nombres': cliente.nombres,
        'apellidos': cliente.apellidos,
        'nombre_comercial': cliente.nombre_comercial,
        'telefono': cliente.telefono,
        'email': cliente.email,
        'direccion': cliente.direccion,
        'tipo_cliente': cliente.tipo_cliente,
        'limite_credito': float(cliente.limite_credito),
        'credito_disponible': float(cliente.credito_disponible),
        'dias_credito': cliente.dias_credito,
        'descuento_general': float(cliente.descuento_general),
        'total_compras': float(cliente.total_compras),
        'activo': cliente.activo,
    }
    
    return JsonResponse(data)

@ensure_csrf_cookie
@auth_required
def devoluciones_view(request):
    """Lista de devoluciones"""
    from apps.sales_management.models import Devolucion, Venta
    from django.db.models import Sum, Count
    
    devoluciones = Devolucion.objects.all().select_related(
        'venta_original', 
        'venta_original__cliente',
        'detalle_venta',
        'detalle_venta__producto'
    ).order_by('-fecha_devolucion')
    
    # Estadísticas
    total_devoluciones = devoluciones.count()
    pendientes = devoluciones.filter(estado='PENDIENTE').count()
    aprobadas = devoluciones.filter(estado='APROBADA').count()
    monto_total = devoluciones.filter(estado='APROBADA').aggregate(
        total=Sum('monto_devolucion')
    )['total'] or 0
    
    context = {
        'devoluciones': devoluciones,
        'total_devoluciones': total_devoluciones,
        'pendientes': pendientes,
        'aprobadas': aprobadas,
        'monto_total': monto_total,
    }
    
    return render(request, 'custom_admin/ventas/devoluciones_list.html', context)



@ensure_csrf_cookie
@auth_required
def devolucion_crear(request):
    """Crear nueva devolución - versión simplificada"""
    from apps.sales_management.models import Devolucion, Venta
    from django.contrib import messages
    from django.shortcuts import redirect
    from datetime import datetime
    from decimal import Decimal

    if request.method == 'POST':
        print("=" * 50)
        print("DEVOLUCION_CREAR - POST recibido")
        print(f"Usuario: {request.user}")
        print(f"Is authenticated: {request.user.is_authenticated}")
        print("Datos POST:", dict(request.POST))
        print("=" * 50)
        
        try:
#             # Verificar autenticación PRIMERO
#             if not request.user.is_authenticated:
#                 messages.error(request, "Debe iniciar sesión para crear devoluciones")
#                 return redirect("custom_admin:devoluciones")
            
            # Obtener datos del formulario
            numero_venta = request.POST.get('numero_venta')
            motivo = request.POST.get('motivo', 'DEFECTUOSO')
            descripcion = request.POST.get('observaciones', '')

            # Buscar la venta
            try:
                venta = Venta.objects.get(numero_venta=numero_venta)
            except Venta.DoesNotExist:
                messages.error(request, f'Venta {numero_venta} no encontrada')
                return redirect('custom_admin:devoluciones')

            # Generar número de devolución
            year = datetime.now().year
            last_dev = Devolucion.objects.filter(
                numero_devolucion__startswith=f'DEV-{year}-'
            ).order_by('numero_devolucion').last()

            if last_dev:
                try:
                    last_num = int(last_dev.numero_devolucion.split('-')[-1])
                    new_num = f"DEV-{year}-{str(last_num + 1).zfill(5)}"
                except:
                    new_num = f"DEV-{year}-00001"
            else:
                new_num = f"DEV-{year}-00001"

            # Obtener primer detalle de venta
            detalle = venta.detalles.first()
            if not detalle:
                messages.error(request, 'La venta no tiene productos')
                return redirect('custom_admin:devoluciones')

            # Calcular cantidad
            cantidad = Decimal('1.0')
            if detalle.peso_vendido:
                cantidad = detalle.peso_vendido
            elif detalle.cantidad_unidades:
                cantidad = Decimal(str(detalle.cantidad_unidades))

            # Obtener el usuario correcto ANTES de crear el objeto
            usuario_para_devolucion = request.user if request.user.is_authenticated else None
            if not usuario_para_devolucion:
                from django.contrib.auth import get_user_model
                Usuario = get_user_model()
                usuario_para_devolucion = Usuario.objects.filter(is_superuser=True).first()

            # Crear la devolución con el usuario ya definido
            devolucion = Devolucion(
                numero_devolucion=new_num,
                venta_original=venta,
                detalle_venta=detalle,
                cantidad_devuelta=cantidad,
                monto_devolucion=venta.total,
                motivo=motivo,
                descripcion=descripcion,
                estado='PENDIENTE',
                usuario_solicita=usuario_para_devolucion
            )

            # Guardar
            devolucion.save()

            messages.success(request, f'Devolución {new_num} creada exitosamente')
            return redirect('custom_admin:devoluciones')

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            import traceback
            print(f"ERROR EN DEVOLUCION: {traceback.format_exc()}")
            return redirect('custom_admin:devoluciones')
    return redirect('custom_admin:devoluciones')
            
    
def api_buscar_venta(request, numero):
    """Buscar venta por número"""
    from apps.sales_management.models import Venta
    
    try:
        venta = Venta.objects.select_related('cliente').prefetch_related('detalles__producto').get(numero_venta=numero)
        
        detalles = []
        for detalle in venta.detalles.all():
            # Usar cantidad_unidades para productos normales o peso_vendido para quintales
            if detalle.peso_vendido:  # Es un quintal
                cantidad = float(detalle.peso_vendido)
                unidad = detalle.unidad_medida.abreviatura if detalle.unidad_medida else 'kg'
                cantidad_str = f"{cantidad} {unidad}"
            else:  # Es un producto normal
                cantidad = detalle.cantidad_unidades if detalle.cantidad_unidades else 1
                cantidad_str = f"{cantidad} unidades"
            
            detalles.append({
                'producto': detalle.producto.nombre if detalle.producto else 'Producto',
                'cantidad': cantidad_str,
                'precio': float(detalle.precio_unitario) if detalle.precio_unitario else 0.0,
                'subtotal': float(detalle.subtotal) if detalle.subtotal else 0.0
            })
        
        data = {
            'success': True,
            'venta': {
                'numero': venta.numero_venta,
                'cliente': f"{venta.cliente.nombres} {venta.cliente.apellidos}" if venta.cliente else "Cliente General",
                'fecha': venta.fecha_venta.strftime('%Y-%m-%d %H:%M'),
                'total': float(venta.total) if venta.total else 0.0,
                'detalles': detalles
            }
        }
    except Venta.DoesNotExist:
        data = {'success': False, 'error': 'Venta no encontrada'}
    except Exception as e:
        data = {'success': False, 'error': str(e)}
    
    return JsonResponse(data)

def api_devoluciones_list(request):
    """API para listar devoluciones"""
    from apps.sales_management.models import Devolucion
    
    devoluciones = Devolucion.objects.select_related(
        'venta_original__cliente',
        'detalle_venta__producto'
    ).all()
    
    data = []
    for dev in devoluciones:
        data.append({
            'numero': dev.numero_devolucion,
            'venta': dev.venta_original.numero_venta,
            'cliente': f"{dev.venta_original.cliente.nombres} {dev.venta_original.cliente.apellidos}" if dev.venta_original.cliente else "N/A",
            'fecha': dev.fecha_devolucion.strftime('%Y-%m-%d %H:%M'),
            'motivo': dev.get_motivo_display(),
            'monto': float(dev.monto_devolucion),
            'estado': dev.estado
        })
    
    return JsonResponse({'devoluciones': data})


@ensure_csrf_cookie
@auth_required
def devolucion_aprobar(request, devolucion_id):
    """Aprobar una devolución"""
    from apps.sales_management.models import Devolucion
    from django.contrib import messages
    
    try:
        devolucion = Devolucion.objects.get(id=devolucion_id)
        
        if devolucion.estado != 'PENDIENTE':
            messages.warning(request, 'Esta devolución ya fue procesada.')
            return redirect('custom_admin:devoluciones')
        
        devolucion.estado = 'APROBADA'
        devolucion.usuario_aprueba = request.user
        devolucion.fecha_procesado = timezone.now()
        
        # Si debe devolver al inventario
        if hasattr(devolucion, 'devolver_inventario') and devolucion.devolver_inventario:
            detalle = devolucion.detalle_venta
            if detalle.producto:
                if detalle.producto.es_quintal() and detalle.quintal:
                    # Devolver peso al quintal
                    quintal = detalle.quintal
                    quintal.peso_actual += devolucion.cantidad_devuelta
                    quintal.save()
                elif detalle.cantidad_unidades:
                    # Devolver unidades al producto
                    producto = detalle.producto
                    producto.stock_actual += int(devolucion.cantidad_devuelta)
                    producto.save()
        
        devolucion.save()
        messages.success(request, f'Devolución {devolucion.numero_devolucion} aprobada.')
        
    except Devolucion.DoesNotExist:
        messages.error(request, 'Devolución no encontrada.')
    except Exception as e:
        messages.error(request, f'Error al aprobar: {str(e)}')
    
    return redirect('custom_admin:devoluciones')

@ensure_csrf_cookie
@auth_required
def devolucion_detalle(request, devolucion_id):
    """Ver detalle de una devolución"""
    from apps.sales_management.models import Devolucion
    import json
    
    try:
        devolucion = Devolucion.objects.select_related(
            'venta_original',
            'venta_original__cliente',
            'detalle_venta__producto',
            'usuario_solicita',
            'usuario_aprueba'
        ).get(id=devolucion_id)
        
        # Construir el diccionario con todos los valores como strings/números básicos
        data = {
            'success': True,
            'devolucion': {
                'id': str(devolucion.id),
                'numero_devolucion': str(devolucion.numero_devolucion),
                'venta_original': str(devolucion.venta_original.numero_venta),
                'cliente': str(devolucion.venta_original.cliente.nombre_completo if devolucion.venta_original.cliente else 'Cliente General'),
                'producto': str(devolucion.detalle_venta.producto.nombre),
                'cantidad_devuelta': float(devolucion.cantidad_devuelta),
                'monto': float(devolucion.monto_devolucion),
                'motivo': str(devolucion.motivo),
                'motivo_display': str(devolucion.get_motivo_display()),
                'descripcion': str(devolucion.descripcion),
                'estado': str(devolucion.estado),
                'estado_display': str(devolucion.get_estado_display()),
                'fecha_devolucion': devolucion.fecha_devolucion.strftime('%d/%m/%Y %H:%M'),
                'usuario_solicita': str(devolucion.usuario_solicita.username) if devolucion.usuario_solicita else 'N/A',
                'usuario_aprueba': str(devolucion.usuario_aprueba.username) if devolucion.usuario_aprueba else None,
                'fecha_procesado': devolucion.fecha_procesado.strftime('%d/%m/%Y %H:%M') if devolucion.fecha_procesado else None
            }
        }
        
        # Serializar manualmente y retornar como HttpResponse
        from django.http import HttpResponse
        json_data = json.dumps(data)
        return HttpResponse(json_data, content_type='application/json')
        
    except Devolucion.DoesNotExist:
        return HttpResponse(
            json.dumps({'success': False, 'error': 'Devolución no encontrada'}),
            content_type='application/json'
        )
    except Exception as e:
        import traceback
        print(f"❌ Error en devolucion_detalle: {traceback.format_exc()}")
        return HttpResponse(
            json.dumps({'success': False, 'error': str(e)}),
            content_type='application/json'
        )
@ensure_csrf_cookie
@auth_required
@require_http_methods(["POST"])
def aprobar_devolucion_api(request, id):
    """API para aprobar o rechazar una devolución"""
    import json
    from apps.sales_management.models import Devolucion
    
    try:
        data = json.loads(request.body)
        devolucion = Devolucion.objects.get(id=id)
        
        if devolucion.estado != 'PENDIENTE':
            return JsonResponse({
                'success': False,
                'error': 'Solo se pueden procesar devoluciones pendientes'
            })
        
        decision = data.get('decision')
        
        if decision == 'APROBADA':
            devolucion.estado = 'APROBADA'
            devolucion.usuario_aprueba = request.user
            devolucion.fecha_procesado = timezone.now()
            devolucion.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Devolución aprobada exitosamente'
            })
            
        elif decision == 'RECHAZADA':
            devolucion.estado = 'RECHAZADA'
            devolucion.usuario_aprueba = request.user
            devolucion.fecha_procesado = timezone.now()
            
            if 'observaciones' in data:
                devolucion.descripcion += f"\n\n--- MOTIVO DE RECHAZO ---\n{data['observaciones']}"
            
            devolucion.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Devolución rechazada'
            })
        
        else:
            return JsonResponse({
                'success': False,
                'error': 'Decisión no válida'
            })
        
    except Devolucion.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Devolución no encontrada'})
    except Exception as e:
        import traceback
        print(f"❌ Error en aprobar_devolucion_api: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'error': str(e)})
@ensure_csrf_cookie
@auth_required
def ventas_export_excel(request):
    """Exportar ventas a Excel - FUNCIÓN ANTIGUA (mantener por compatibilidad)"""
    return exportar_ventas_excel_general(request)


@ensure_csrf_cookie
@auth_required
def ventas_export_pdf(request):
    """Exportar ventas a PDF - FUNCIÓN ANTIGUA (mantener por compatibilidad)"""
    return exportar_ventas_pdf_general(request)


# ========================================
# PUNTO DE VENTA (POS)
# ========================================

@ensure_csrf_cookie
@auth_required
def pos_view(request):
    """Punto de Venta principal"""
    from apps.sales_management.models import Cliente
    from apps.inventory_management.models import Categoria
    from apps.system_configuration.models import ConfiguracionSistema  # ✅ AGREGAR IMPORT
    
    clientes = Cliente.objects.filter(activo=True).order_by('apellidos', 'nombres')
    categorias = Categoria.objects.filter(activa=True).order_by('nombre')
    
    # ✅ OBTENER CONFIGURACIÓN DE IVA
    config = ConfiguracionSistema.get_config()
    iva_porcentaje = float(config.porcentaje_iva) if config else 0
    iva_activo = config.iva_activo if config else False
    
    context = {
        'clientes': clientes,
        'categorias': categorias,
        'iva_porcentaje': iva_porcentaje,  # ✅ AGREGAR AL CONTEXTO
        'iva_activo': iva_activo,  # ✅ OPCIONAL: para mostrar/ocultar IVA en UI
    }
    
    return render(request, 'custom_admin/pos/punto_venta.html', context)


def api_buscar_productos(request):
    """API para buscar productos"""
    from apps.inventory_management.models import Producto, ProductoNormal, Quintal
    from django.db.models import Q, Sum
    from django.http import JsonResponse
    
    query = request.GET.get('q', '').strip()
    categoria_id = request.GET.get('categoria', '')
    cargar_todos = request.GET.get('all', '') == 'true'
    
    if not query and not categoria_id and not cargar_todos:
        return JsonResponse({'productos': []})
    
    productos = Producto.objects.select_related('categoria', 'unidad_medida_base', 'marca').filter(activo=True)
    
    if query:
        productos = productos.filter(
            Q(nombre__icontains=query) |
            Q(codigo_barras__icontains=query) |
            Q(descripcion__icontains=query)
        )
    
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    
    if not cargar_todos:
        productos = productos[:20]
    else:
        productos = productos[:100]
    
    data = []
    for p in productos:
        # Obtener stock según tipo de inventario
        stock_actual = 0
        
        if p.tipo_inventario == 'NORMAL':
            try:
                inventario = p.inventario_normal
                stock_actual = float(inventario.stock_actual) if inventario else 0
            except ProductoNormal.DoesNotExist:
                stock_actual = 0
        elif p.tipo_inventario == 'QUINTAL':
            stock_actual = float(
                Quintal.objects.filter(
                    producto=p,
                    estado='DISPONIBLE'
                ).aggregate(
                    total=Sum('peso_actual')
                )['total'] or 0
            )
        
        # ✅ UNIFICAR PRECIO - siempre devolver 'precio' como campo principal
        if p.tipo_inventario == 'NORMAL':
            precio = float(p.precio_venta) if p.precio_venta else 0
        else:  # QUINTAL
            precio = float(p.precio_por_unidad_peso) if p.precio_por_unidad_peso else 0
        
        data.append({
            'id': str(p.id),
            'nombre': p.nombre,
            'codigo_barras': p.codigo_barras or '',
            'categoria': p.categoria.nombre if p.categoria else '',
            'marca': p.marca.nombre if p.marca else '',
            'tipo_inventario': p.tipo_inventario,
            'precio': precio,  # ✅ CAMPO UNIFICADO
            'precio_unitario': precio,  # ✅ MANTENER POR COMPATIBILIDAD
            'precio_por_unidad_peso': precio if p.tipo_inventario == 'QUINTAL' else 0,
            'stock_actual': stock_actual,
            'unidad_medida': p.unidad_medida_base.abreviatura if p.unidad_medida_base else 'und',
            'aplica_impuestos': p.aplica_impuestos,
        })
    
    return JsonResponse({'productos': data})


@ensure_csrf_cookie
def api_obtener_producto(request, producto_id):
    """API para obtener info completa de un producto"""
    from apps.inventory_management.models import Producto
    from django.http import JsonResponse
    
    producto = get_object_or_404(Producto.objects.select_related('categoria', 'unidad_medida_base'), pk=producto_id)
    
    data = {
        'id': str(producto.id),
        'nombre': producto.nombre,
        'descripcion': producto.descripcion or '',
        'codigo_barras': producto.codigo_barras or '',
        'categoria': producto.categoria.nombre if producto.categoria else '',
        'tipo_inventario': producto.tipo_inventario,
        'precio_unitario': float(producto.precio_unitario) if producto.precio_unitario else 0,
        'precio_por_unidad_peso': float(producto.precio_por_unidad_peso) if producto.precio_por_unidad_peso else 0,
        'stock_actual': float(producto.stock_actual) if hasattr(producto, 'stock_actual') else 0,
        'unidad_medida': producto.unidad_medida_base.abreviatura if producto.unidad_medida_base else 'und',
        'permite_descuento': True,
        'descuento_maximo': float(producto.categoria.descuento_maximo_permitido) if producto.categoria else 10.0,
    }
    
    return JsonResponse(data)


@ensure_csrf_cookie
def api_procesar_venta(request):
    """
    API para procesar y guardar la venta con impresión térmica
    VERSIÓN CON IVA CORREGIDO
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    import json
    from django.http import JsonResponse
    from apps.sales_management.models import Venta, DetalleVenta, Cliente, Pago
    from apps.inventory_management.models import Producto
    from apps.authentication.models import Usuario
    from apps.hardware_integration.models import TrabajoImpresion, Impresora
    from apps.system_configuration.models import ConfiguracionSistema
    from decimal import Decimal
    from django.utils import timezone
    from django.db import transaction
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        data = json.loads(request.body)
        
        items = data.get('items', [])
        if not items:
            return JsonResponse({'success': False, 'error': 'No hay productos en el carrito'})
        
        cliente_id = data.get('cliente_id')
        tipo_venta = data.get('tipo_venta', 'CONTADO')
        metodo_pago = data.get('metodo_pago', 'EFECTIVO')
        
        try:
            monto_recibido = Decimal(str(data.get('monto_recibido', '0')))
        except:
            monto_recibido = Decimal('0')
        
        # Obtener usuario
        usuario = request.user if request.user.is_authenticated else None
        if not usuario:
            usuario = Usuario.objects.filter(rol__codigo__in=['VENDEDOR', 'ADMIN', 'CAJERO']).first()
        if not usuario:
            usuario = Usuario.objects.first()
        
        if not usuario:
            return JsonResponse({'success': False, 'error': 'No hay usuarios en el sistema'})
        
        cliente = None
        if cliente_id:
            try:
                cliente = Cliente.objects.get(id=cliente_id)
            except Cliente.DoesNotExist:
                pass
        
        # ============================================================================
        # ✅ OBTENER CONFIGURACIÓN DE IVA
        # ============================================================================
        config = ConfiguracionSistema.get_config()
        iva_activo = config.iva_activo if config else False
        porcentaje_iva = config.porcentaje_iva if config else Decimal('0')
        
        logger.info(f"💰 IVA Activo: {iva_activo}, Porcentaje: {porcentaje_iva}%")
        
        # ============================================================================
        # ✅ CALCULAR TOTALES CON IVA
        # ============================================================================
        subtotal = Decimal('0')
        descuento_total = Decimal('0')
        impuestos_total = Decimal('0')
        
        for item in items:
            try:
                item_subtotal = Decimal(str(item.get('subtotal', '0')))
                item_descuento = Decimal(str(item.get('descuento', '0')))
                
                subtotal += item_subtotal
                descuento_total += item_descuento
                
                # ✅ CALCULAR IVA DEL ITEM SI APLICA
                producto_id = item.get('producto_id')
                if producto_id and iva_activo:
                    try:
                        producto = Producto.objects.get(id=producto_id)
                        if producto.aplica_impuestos:
                            base_imponible = item_subtotal - item_descuento
                            iva_item = base_imponible * (porcentaje_iva / Decimal('100'))
                            impuestos_total += iva_item
                    except Producto.DoesNotExist:
                        pass
            except:
                continue
        
        # ✅ TOTAL CON IVA
        total = subtotal - descuento_total + impuestos_total
        cambio = monto_recibido - total if monto_recibido >= total else Decimal('0')
        
        logger.info(f"📊 Subtotal: ${subtotal}, Descuento: ${descuento_total}, IVA: ${impuestos_total}, Total: ${total}")
        
        with transaction.atomic():
            # Generar número de venta
            año = timezone.now().year
            ventas_año = Venta.objects.filter(
                numero_venta__startswith=f'VNT-{año}-'
            ).count()
            siguiente = ventas_año + 1
            numero_venta_generado = 'VNT-{}-{:05d}'.format(año, siguiente)
            
            # ✅ CREAR VENTA CON IMPUESTOS
            venta = Venta.objects.create(
                numero_venta=numero_venta_generado,
                cliente=cliente,
                vendedor=usuario,
                tipo_venta=tipo_venta,
                subtotal=subtotal,
                descuento=descuento_total,
                impuestos=impuestos_total,
                total=total,
                estado='COMPLETADA',
                monto_pagado=monto_recibido,
                cambio=cambio,
            )
            
            logger.info(f"✅ Venta creada: {venta.numero_venta} - Total: ${total} (IVA: ${impuestos_total})")
            
            # Crear detalles
            orden = 1
            for item in items:
                try:
                    producto = Producto.objects.get(id=item['producto_id'])
                    
                    cantidad = Decimal(str(item.get('cantidad', '0')))
                    precio = Decimal(str(item.get('precio', '0')))
                    descuento = Decimal(str(item.get('descuento', '0')))
                    descuento_porcentaje = Decimal(str(item.get('descuento_porcentaje', '0')))
                    
                    costo_unitario = precio * Decimal('0.7')
                    costo_total = cantidad * costo_unitario
                    item_subtotal = cantidad * precio
                    
                    # ✅ CALCULAR IVA DEL DETALLE
                    base_imponible = item_subtotal - descuento
                    iva_detalle = Decimal('0')
                    
                    if producto.aplica_impuestos and iva_activo:
                        iva_detalle = base_imponible * (porcentaje_iva / Decimal('100'))
                    
                    item_total = base_imponible + iva_detalle
                    
                    # ✅ USAR CAMPOS CORRECTOS DEL MODELO
                    detalle_data = {
                        'venta': venta,
                        'producto': producto,
                        'orden': orden,
                        'costo_unitario': costo_unitario,
                        'costo_total': costo_total,
                        'descuento_porcentaje': descuento_porcentaje,
                        'descuento_monto': descuento,
                        'aplica_iva': producto.aplica_impuestos if iva_activo else False,  # ✅ CORRECTO
                        'monto_iva': iva_detalle,  # ✅ CORRECTO
                        'subtotal': item_subtotal,
                        'total': item_total,
                    }
                    
                    # Manejar quintales o productos normales
                    if item.get('es_quintal') and item.get('quintal_id'):
                        from apps.inventory_management.models import Quintal
                        try:
                            quintal = Quintal.objects.get(id=item['quintal_id'])
                            detalle_data['quintal'] = quintal
                            detalle_data['peso_vendido'] = Decimal(str(item.get('peso_vendido', 0)))
                            detalle_data['precio_por_unidad_peso'] = producto.precio_por_unidad_peso
                            detalle_data['unidad_medida'] = quintal.unidad_medida
                        except Quintal.DoesNotExist:
                            logger.error(f"Quintal no encontrado: {item['quintal_id']}")
                    elif producto.tipo_inventario == 'QUINTAL':
                        detalle_data['peso_vendido'] = cantidad
                        detalle_data['precio_por_unidad_peso'] = precio
                    else:
                        detalle_data['cantidad_unidades'] = int(cantidad)
                        detalle_data['precio_unitario'] = precio
                    
                    # ✅ CREAR DETALLE (el signal se encargará de descontar inventario)
                    detalle = DetalleVenta.objects.create(**detalle_data)
                    orden += 1
                    
                except Producto.DoesNotExist:
                    logger.error(f"Producto no encontrado: {item.get('producto_id')}")
                    continue
                except Exception as e:
                    logger.error(f"Error en detalle: {e}", exc_info=True)
                    continue
            
            # ============================================================================
            # REGISTRAR PAGO
            # ============================================================================
            if monto_recibido > 0:
                Pago.objects.create(
                    venta=venta,
                    forma_pago=metodo_pago,
                    monto=monto_recibido,
                    numero_referencia=f"Pago {metodo_pago} - {venta.numero_venta}",
                    banco='',
                    usuario=usuario
                )
                logger.info(f"✅ Pago registrado: {metodo_pago} - ${monto_recibido}")
            
            # ============================================================================
            # CREAR TRABAJO DE IMPRESIÓN
            # ============================================================================
            try:
                impresora = Impresora.objects.filter(
                    es_principal_tickets=True,
                    estado='ACTIVA'
                ).first()
                
                if not impresora:
                    impresora = Impresora.objects.filter(
                        tipo_impresora__in=['TERMICA_TICKET', 'TERMICA_FACTURA'],
                        estado='ACTIVA'
                    ).first()
                
                if impresora:
                    # Generar comandos ESC/POS
                    comandos_bytes = generar_comandos_ticket_bytes(venta)
                    comandos_hex = comandos_bytes.hex()
                    
                    logger.info(f"📄 Comandos generados: {len(comandos_bytes)} bytes → {len(comandos_hex)} chars hex")
                    
                    # Crear trabajo de impresión
                    trabajo = TrabajoImpresion.objects.create(
                        tipo='TICKET',
                        prioridad=1,
                        estado='PENDIENTE',
                        impresora=impresora,
                        venta=venta,
                        datos_impresion=comandos_hex,
                        formato='ESC_POS',
                        abrir_gaveta=True if metodo_pago == 'EFECTIVO' else False,
                        copias=1,
                        creado_por=usuario,
                        max_intentos=3
                    )
                    
                    logger.info(f"✅ Trabajo de impresión creado: {trabajo.id}")
                else:
                    logger.warning("⚠️ No hay impresora configurada para tickets")
                    
            except Exception as e:
                logger.error(f"❌ Error creando trabajo de impresión: {e}", exc_info=True)
        
        return JsonResponse({
            'success': True,
            'venta_id': str(venta.id),
            'numero_venta': venta.numero_venta,
            'subtotal': float(subtotal),
            'descuento': float(descuento_total),
            'impuestos': float(impuestos_total),
            'total': float(total),
            'cambio': float(cambio),
            'monto_recibido': float(monto_recibido),
        })
        
    except Exception as e:
        logger.error("❌ Error procesando venta:", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': f'Error: {str(e)}'
        }, status=500)

@ensure_csrf_cookie
def api_calcular_precio_iva(request):
    """API para calcular el precio con IVA de un producto"""
    from apps.inventory_management.models import Producto
    from apps.system_configuration.models import ConfiguracionSistema
    from decimal import Decimal
    from django.http import JsonResponse
    
    producto_id = request.GET.get('producto_id')
    cantidad = Decimal(str(request.GET.get('cantidad', '1')))
    
    if not producto_id:
        return JsonResponse({'success': False, 'error': 'ID de producto requerido'})
    
    try:
        producto = Producto.objects.get(id=producto_id)
        
        # Obtener precio base según tipo
        if producto.tipo_inventario == 'NORMAL':
            precio_base = producto.precio_venta or Decimal('0')
        else:  # QUINTAL
            precio_base = producto.precio_por_unidad_peso or Decimal('0')
        
        # Calcular subtotal
        subtotal = precio_base * cantidad
        
        # Calcular IVA si aplica
        monto_iva = Decimal('0')
        porcentaje_iva = Decimal('0')
        
        if producto.aplica_impuestos:
            config = ConfiguracionSistema.get_config()
            if config and config.iva_activo:
                porcentaje_iva = config.porcentaje_iva
                monto_iva = subtotal * (porcentaje_iva / Decimal('100'))
        
        total_con_iva = subtotal + monto_iva
        
        return JsonResponse({
            'success': True,
            'precio_base': float(precio_base),
            'subtotal': float(subtotal),
            'aplica_iva': producto.aplica_impuestos,
            'porcentaje_iva': float(porcentaje_iva),
            'monto_iva': float(monto_iva),
            'total': float(total_con_iva)
        })
        
    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Producto no encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
@ensure_csrf_cookie
def api_reimprimir_ticket(request, venta_id):
    """
    API para reimprimir un ticket de venta
    Crea un nuevo trabajo de impresión para el agente
    
    POST /panel/api/ventas/<venta_id>/reimprimir-ticket/
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    import logging
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from apps.sales_management.models import Venta
    from apps.hardware_integration.models import Impresora, TrabajoImpresion
    from apps.authentication.models import Usuario
    
    logger = logging.getLogger(__name__)
    
    try:
        # Obtener la venta
        venta = get_object_or_404(Venta, pk=venta_id)
        
        # Obtener usuario
        usuario = request.user if request.user.is_authenticated else None
        if not usuario:
            usuario = Usuario.objects.filter(rol__codigo__in=['VENDEDOR', 'ADMIN', 'CAJERO']).first()
        if not usuario:
            usuario = Usuario.objects.first()
        
        # Obtener impresora principal de tickets
        impresora = Impresora.objects.filter(
            es_principal_tickets=True,
            estado='ACTIVA'
        ).first()
        
        if not impresora:
            impresora = Impresora.objects.filter(
                tipo_impresora__in=['TERMICA_TICKET', 'TERMICA_FACTURA'],
                estado='ACTIVA'
            ).first()
        
        if not impresora:
            logger.warning("⚠️ No hay impresora configurada")
            return JsonResponse({
                'success': False,
                'error': 'No hay impresora configurada. Configure una impresora en Hardware Integration.'
            })
        
        # ✅ GENERAR COMANDOS ESC/POS
        comandos_bytes = generar_comandos_ticket_bytes(venta)
        comandos_hex = comandos_bytes.hex()
        
        logger.info(f"📄 Reimpresión - Comandos: {len(comandos_bytes)} bytes")
        
        # Verificar si quiere abrir gaveta (opcional en reimpresión)
        abrir_gaveta = request.POST.get('abrir_gaveta', 'false') == 'true'
        
        # Crear trabajo de impresión
        trabajo = TrabajoImpresion.objects.create(
            tipo='TICKET',
            prioridad=1,  # Alta prioridad
            estado='PENDIENTE',
            impresora=impresora,
            venta=venta,
            datos_impresion=comandos_hex,
            formato='ESC_POS',
            abrir_gaveta=abrir_gaveta,
            copias=1,
            creado_por=usuario,
            max_intentos=3
        )
        
        logger.info(f"✅ Trabajo de reimpresión creado: {trabajo.id}")
        logger.info(f"   Venta: {venta.numero_venta}")
        logger.info(f"   Impresora: {impresora.nombre}")
        
        return JsonResponse({
            'success': True,
            'trabajo_id': str(trabajo.id),
            'mensaje': f'Ticket de venta {venta.numero_venta} enviado a la impresora {impresora.nombre}'
        })
        
    except Exception as e:
        logger.error(f"❌ Error en reimpresión: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': f'Error al reimprimir: {str(e)}'
        }, status=500)

# ============================================================================
# ✅ FUNCIÓN NUEVA: GENERAR COMANDOS TICKET EN BYTES
# ============================================================================

def generar_comandos_ticket_bytes(venta):
    """
    Genera comandos ESC/POS en formato BYTES para impresora térmica
    VERSIÓN CORREGIDA - Error de get_metodo_pago_display solucionado
    
    Args:
        venta: Instancia del modelo Venta
    
    Returns:
        bytes: Comandos ESC/POS listos para enviar a la impresora
    """
    from django.conf import settings
    
    # Comandos ESC/POS (BYTES, no strings)
    ESC = b'\x1b'
    GS = b'\x1d'
    
    INIT = ESC + b'@'  # Inicializar
    CENTER = ESC + b'a\x01'  # Centrar
    LEFT = ESC + b'a\x00'  # Izquierda
    BOLD_ON = ESC + b'E\x01'  # Negrita ON
    BOLD_OFF = ESC + b'E\x00'  # Negrita OFF
    DOUBLE_HEIGHT = ESC + b'!\x10'  # Doble altura
    NORMAL = ESC + b'!\x00'  # Normal
    CUT = GS + b'V\x00'  # Cortar papel
    
    # Construir ticket
    ticket = b''
    
    # Inicializar
    ticket += INIT
    
    # ========================================
    # ENCABEZADO
    # ========================================
    ticket += CENTER
    ticket += BOLD_ON + DOUBLE_HEIGHT + b"COMMERCEBOX\n" + NORMAL + BOLD_OFF
    ticket += b"RUC: 1234567890001\n"
    ticket += b"Av. Principal #123\n"
    ticket += b"Quito, Ecuador\n"
    ticket += b"Tel: (02) 123-4567\n"
    ticket += b"\n"
    
    # ========================================
    # INFORMACIÓN DE VENTA
    # ========================================
    ticket += LEFT
    ticket += b"=" * 42 + b"\n"
    ticket += BOLD_ON + f"TICKET: {venta.numero_venta}\n".encode('utf-8') + BOLD_OFF
    ticket += f"Fecha: {venta.fecha_creacion.strftime('%d/%m/%Y %H:%M')}\n".encode('utf-8')
    ticket += f"Cajero: {venta.vendedor.get_full_name()}\n".encode('utf-8')
    
    if venta.cliente:
        ticket += f"Cliente: {venta.cliente.nombres} {venta.cliente.apellidos}\n".encode('utf-8')
        if hasattr(venta.cliente, 'numero_documento'):
            ticket += f"CI/RUC: {venta.cliente.numero_documento}\n".encode('utf-8')
    
    ticket += b"=" * 42 + b"\n"
    ticket += b"\n"
    
    # ========================================
    # PRODUCTOS
    # ========================================
    ticket += b"CANT  PRODUCTO           PRECIO    TOTAL\n"
    ticket += b"-" * 42 + b"\n"
    
    for detalle in venta.detalles.all():
        producto = detalle.producto
        nombre = producto.nombre[:20]  # Máximo 20 caracteres
        
        if hasattr(detalle, 'quintal') and detalle.quintal:
            cant_str = f"{detalle.peso_vendido:.2f}kg"
            precio_str = f"${detalle.precio_por_unidad_peso:.2f}"
        else:
            cant_str = f"{detalle.cantidad_unidades}"
            precio_str = f"${detalle.precio_unitario:.2f}"
        
        total_str = f"${detalle.total:.2f}"
        
        linea = f"{cant_str:<5} {nombre:<20} {precio_str:>7} {total_str:>7}\n"
        ticket += linea.encode('utf-8')
        
        if detalle.descuento_monto and detalle.descuento_monto > 0:
            desc_linea = f"      Descuento: -${detalle.descuento_monto:.2f}\n"
            ticket += desc_linea.encode('utf-8')
    
    ticket += b"\n"
    ticket += b"=" * 42 + b"\n"
    
    # ========================================
    # TOTALES
    # ========================================
    ticket += LEFT
    ticket += f"{'SUBTOTAL:':<32}${venta.subtotal:>9.2f}\n".encode('utf-8')
    
    if venta.descuento and venta.descuento > 0:
        ticket += f"{'DESCUENTO:':<32}-${venta.descuento:>8.2f}\n".encode('utf-8')
    
    if venta.impuestos and venta.impuestos > 0:
        ticket += f"{'IVA (15%):':<32}${venta.impuestos:>9.2f}\n".encode('utf-8')
    
    ticket += b"\n"
    ticket += BOLD_ON + DOUBLE_HEIGHT
    ticket += f"{'TOTAL:':<16}${venta.total:>9.2f}\n".encode('utf-8')
    ticket += NORMAL + BOLD_OFF
    
    # ========================================
    # INFORMACIÓN DE PAGO - ✅ CORREGIDO
    # ========================================
    if venta.monto_pagado and venta.monto_pagado > 0:
        ticket += b"\n"
        ticket += LEFT
        
        pago = venta.pagos.first()
        
        # ✅ CORRECCIÓN: usar get_forma_pago_display() en lugar de get_metodo_pago_display()
        if pago:
            metodo = pago.get_forma_pago_display()  # ✅ CORRECTO
        else:
            metodo = "EFECTIVO"
        
        ticket += f"Forma de pago: {metodo}\n".encode('utf-8')
        ticket += f"{'Recibido:':<32}${venta.monto_pagado:>9.2f}\n".encode('utf-8')
        
        if venta.cambio and venta.cambio > 0:
            ticket += BOLD_ON
            ticket += f"{'Cambio:':<32}${venta.cambio:>9.2f}\n".encode('utf-8')
            ticket += BOLD_OFF
    
    # ========================================
    # PIE DE PÁGINA
    # ========================================
    ticket += b"\n"
    ticket += CENTER
    ticket += b"=" * 42 + b"\n"
    ticket += b"GRACIAS POR SU COMPRA!\n"
    ticket += b"\n"
    ticket += b"Este documento no tiene\n"
    ticket += b"validez tributaria\n"
    ticket += b"\n"
    ticket += b"Conserve este ticket para\n"
    ticket += b"cambios y devoluciones\n"
    ticket += b"\n"
    ticket += b"\n"
    ticket += b"\n"
    
    # Cortar papel
    ticket += CUT
    
    return ticket


@ensure_csrf_cookie
def api_venta_detalle(request, venta_id):
    """API para obtener detalle de una venta"""
    from apps.sales_management.models import Venta
    from django.http import JsonResponse
    
    venta = get_object_or_404(
        Venta.objects.select_related('cliente', 'vendedor'), 
        pk=venta_id
    )
    
    detalles = venta.detalles.select_related('producto', 'unidad_medida').all()
    pagos = venta.pagos.select_related('usuario').all()
    
    venta_data = {
        'numero_venta': venta.numero_venta,
        'fecha_venta': venta.fecha_venta.strftime('%d/%m/%Y %H:%M'),
        'estado': venta.estado,
        'tipo_venta': venta.get_tipo_venta_display(),
        'subtotal': float(venta.subtotal),
        'descuento': float(venta.descuento),
        'impuestos': float(venta.impuestos),
        'total': float(venta.total),
        'monto_pagado': float(venta.monto_pagado),
        'cliente_nombre': f"{venta.cliente.nombres} {venta.cliente.apellidos}" if venta.cliente else None,
        'cliente_documento': venta.cliente.numero_documento if venta.cliente else None,
        'cliente_telefono': venta.cliente.telefono if venta.cliente else None,
        'vendedor_nombre': venta.vendedor.get_full_name(),
    }
    
    detalles_data = []
    for detalle in detalles:
        detalles_data.append({
            'producto_nombre': detalle.producto.nombre,
            'cantidad_unidades': detalle.cantidad_unidades,
            'peso_vendido': float(detalle.peso_vendido) if detalle.peso_vendido else None,
            'unidad_medida': detalle.unidad_medida.abreviatura if detalle.unidad_medida else 'und',
            'precio_unitario': float(detalle.precio_unitario) if detalle.precio_unitario else None,
            'precio_por_unidad_peso': float(detalle.precio_por_unidad_peso) if detalle.precio_por_unidad_peso else None,
            'descuento_monto': float(detalle.descuento_monto),
            'total': float(detalle.total),
        })
    
    pagos_data = []
    for pago in pagos:
        pagos_data.append({
            'forma_pago': pago.get_forma_pago_display(),
            'monto': float(pago.monto),
            'fecha_pago': pago.fecha_pago.strftime('%d/%m/%Y %H:%M'),
            'usuario': pago.usuario.get_full_name(),
        })
    
    return JsonResponse({
        'venta': venta_data,
        'detalles': detalles_data,
        'pagos': pagos_data,
    })

# ========================================
# ENTRADA DE INVENTARIO UNIFICADA
# ========================================

from apps.system_configuration.models import ConfiguracionSistema

@ensure_csrf_cookie
@auth_required
def entrada_inventario_view(request):
    from apps.inventory_management.models import Proveedor, UnidadMedida, Categoria, Marca
    
    categorias = Categoria.objects.filter(activa=True).order_by('orden', 'nombre')
    proveedores = Proveedor.objects.filter(activo=True).order_by('nombre_comercial')
    marcas = Marca.objects.filter(activa=True).order_by('nombre')
    unidades_medida = UnidadMedida.objects.filter(activa=True).order_by('orden_display')
    
    # ✅ OBTENER IVA DESDE CONFIGURACIÓN
    config = ConfiguracionSistema.get_config()
    iva_porcentaje = config.porcentaje_iva
    
    context = {
        'categorias': categorias,
        'proveedores': proveedores,
        'marcas': marcas,
        'unidades_medida': unidades_medida,
        'iva_porcentaje': iva_porcentaje,       # ✅ Para usar en la vista si es necesario
        'iva_default': iva_porcentaje,          # ✅ AGREGAR: Para el HTML del select
        'porcentaje_iva': iva_porcentaje,       # ✅ AGREGAR: Para el JavaScript
    }
    
    return render(request, 'custom_admin/inventario/entrada_inventario.html', context)
def api_buscar_producto_codigo(request):
    """API para buscar producto por código de barras"""
    from apps.inventory_management.models import Producto
    from django.http import JsonResponse
    
    codigo = request.GET.get('codigo', '').strip()
    
    if not codigo:
        return JsonResponse({
            'success': False,
            'mensaje': 'Código vacío'
        })
    
    try:
        producto = Producto.objects.get(codigo_barras=codigo, activo=True)
        
        data = {
            'success': True,
            'producto': {
                'id': str(producto.id),
                'nombre': producto.nombre,
                'codigo_barras': producto.codigo_barras,
                'tipo_inventario': producto.tipo_inventario,
                'unidad_medida_id': str(producto.unidad_medida_base.id) if producto.unidad_medida_base else None,
            }
        }
        
        return JsonResponse(data)
        
    except Producto.DoesNotExist:
        return JsonResponse({
            'success': False,
            'mensaje': f'Producto no encontrado: {codigo}'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'mensaje': f'Error: {str(e)}'
        }, status=500)


@ensure_csrf_cookie
def api_procesar_entrada_masiva(request):
    """API para procesar entrada masiva de productos"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    import json
    from django.http import JsonResponse
    from django.db import transaction
    from apps.inventory_management.models import (
        Producto, Quintal, ProductoNormal, MovimientoInventario,
        Proveedor, UnidadMedida
    )
    from apps.authentication.models import Usuario
    from decimal import Decimal
    from django.utils import timezone
    
    try:
        data = json.loads(request.body)
        entradas = data.get('entradas', [])
        numero_factura = data.get('numero_factura', '')
        observaciones = data.get('observaciones', '')
        
        if not entradas:
            return JsonResponse({'success': False, 'error': 'No hay productos para procesar'})
        
        usuario = Usuario.objects.filter(rol__codigo__in=['ADMIN', 'SUPERVISOR']).first()
        if not usuario:
            usuario = Usuario.objects.first()
        
        codigos_generados = []
        quintales_creados = 0
        productos_actualizados = 0
        
        with transaction.atomic():
            for entrada in entradas:
                try:
                    producto = Producto.objects.get(id=entrada['producto_id'])
                    proveedor = Proveedor.objects.get(id=entrada['proveedor_id'])
                    
                    if producto.tipo_inventario == 'QUINTAL':
                        # Crear Quintal
                        from apps.inventory_management.utils.barcode_generator import BarcodeGenerator
                        
                        unidad_medida = UnidadMedida.objects.get(id=entrada['unidad_medida_id'])
                        peso_inicial = Decimal(str(entrada['peso_inicial']))
                        costo_total = Decimal(str(entrada['costo_total']))
                        
                        quintal = Quintal.objects.create(
                            codigo_unico=BarcodeGenerator.generar_codigo_quintal(producto),
                            producto=producto,
                            proveedor=proveedor,
                            peso_inicial=peso_inicial,
                            peso_actual=peso_inicial,
                            unidad_medida=unidad_medida,
                            costo_total=costo_total,
                            costo_por_unidad=costo_total / peso_inicial,
                            fecha_ingreso=timezone.now(),
                            fecha_vencimiento=entrada.get('fecha_vencimiento') or None,
                            lote_proveedor=entrada.get('lote_proveedor', ''),
                            numero_factura_compra=numero_factura,
                            usuario_registro=usuario,
                            estado='DISPONIBLE'
                        )
                        
                        quintales_creados += 1
                        
                        # Generar códigos de barras
                        cantidad_etiquetas = int(peso_inicial)  # 1 etiqueta por unidad de peso
                        
                        codigos_generados.append({
                            'producto_id': str(producto.id),
                            'tipo': 'QUINTAL',
                            'codigo': quintal.codigo_unico,
                            'producto_nombre': producto.nombre,
                            'cantidad_etiquetas': cantidad_etiquetas,
                            'peso_unitario': 1,  # 1 lb/kg por etiqueta
                            'unidad': unidad_medida.abreviatura,
                            'pdf_url': f'/panel/api/inventario/generar-pdf-codigos/?quintal_id={quintal.id}'
                        })
                        
                    else:
                        # Producto Normal
                        cantidad = int(entrada['cantidad_unidades'])
                        costo_unitario = Decimal(str(entrada['costo_unitario']))
                        
                        producto_normal, created = ProductoNormal.objects.get_or_create(
                            producto=producto,
                            defaults={
                                'stock_actual': 0,
                                'stock_minimo': 10,
                                'costo_unitario': costo_unitario
                            }
                        )
                        
                        stock_antes = producto_normal.stock_actual
                        producto_normal.stock_actual += cantidad
                        
                        # Actualizar costo promedio ponderado
                        if stock_antes > 0:
                            costo_total_anterior = stock_antes * producto_normal.costo_unitario
                            costo_total_nuevo = cantidad * costo_unitario
                            stock_total = stock_antes + cantidad
                            producto_normal.costo_unitario = (costo_total_anterior + costo_total_nuevo) / stock_total
                        else:
                            producto_normal.costo_unitario = costo_unitario
                        
                        producto_normal.fecha_ultima_entrada = timezone.now()
                        producto_normal.lote = entrada.get('lote', '')
                        producto_normal.fecha_vencimiento = entrada.get('fecha_vencimiento') or None
                        producto_normal.save()
                        
                        # Registrar movimiento
                        MovimientoInventario.objects.create(
                            producto_normal=producto_normal,
                            tipo_movimiento='ENTRADA_COMPRA',
                            cantidad=cantidad,
                            stock_antes=stock_antes,
                            stock_despues=producto_normal.stock_actual,
                            costo_unitario=costo_unitario,
                            costo_total=cantidad * costo_unitario,
                            usuario=usuario,
                            observaciones=observaciones or f"Entrada masiva - Factura: {numero_factura}"
                        )
                        
                        productos_actualizados += 1
                        
                        # Generar códigos de barras
                        codigos_generados.append({
                            'producto_id': str(producto.id),
                            'tipo': 'NORMAL',
                            'codigo': producto.codigo_barras,
                            'producto_nombre': producto.nombre,
                            'cantidad_etiquetas': cantidad,
                            'pdf_url': f'/panel/api/inventario/generar-pdf-codigos/?producto_id={producto.id}&cantidad={cantidad}'
                        })
                        
                except Exception as e:
                    print(f"Error procesando entrada: {e}")
                    continue
        
        return JsonResponse({
            'success': True,
            'quintales_creados': quintales_creados,
            'productos_actualizados': productos_actualizados,
            'codigos_generados': codigos_generados,
            'mensaje': f'Entrada procesada: {quintales_creados} quintales, {productos_actualizados} productos actualizados'
        })
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Error: {str(e)}'
        }, status=500)

@ensure_csrf_cookie
def api_generar_pdf_codigos(request):
    """Genera y devuelve PDF con códigos de barras - VERSIÓN ORGANIZADA"""
    from django.http import HttpResponse, JsonResponse
    from apps.inventory_management.models import Quintal, Producto
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.graphics.barcode import code128
    from io import BytesIO
    from datetime import datetime
    
    quintal_id = request.GET.get('quintal_id')
    producto_id = request.GET.get('producto_id')
    cantidad = int(request.GET.get('cantidad', 50))
    
    try:
        # ========================================
        # CONFIGURACIÓN OPTIMIZADA PARA ETIQUETAS
        # ========================================
        ETIQUETAS_POR_FILA = 3
        ETIQUETAS_POR_COLUMNA = 8
        ANCHO_ETIQUETA = 63 * mm
        ALTO_ETIQUETA = 29 * mm
        MARGEN_IZQUIERDO = 8 * mm
        MARGEN_SUPERIOR = 12 * mm
        ESPACIO_HORIZONTAL = 5 * mm
        ESPACIO_VERTICAL = 4 * mm
        
        # Crear PDF
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=letter)
        
        if quintal_id:
            # ========================================
            # GENERAR ETIQUETAS PARA QUINTAL
            # ========================================
            quintal = Quintal.objects.get(id=quintal_id)
            codigo = quintal.codigo_unico
            nombre_producto = quintal.producto.nombre
            peso_unitario = 1.0
            unidad = quintal.unidad_medida.abreviatura
            precio = float(quintal.producto.precio_por_unidad_peso or 0)
            fecha = quintal.fecha_ingreso.strftime('%d/%m/%Y')
            filename = f'etiquetas_{quintal.codigo_unico}.pdf'
            
            etiqueta_num = 0
            
            while etiqueta_num < cantidad:
                for fila in range(ETIQUETAS_POR_COLUMNA):
                    for columna in range(ETIQUETAS_POR_FILA):
                        if etiqueta_num >= cantidad:
                            break
                        
                        # Calcular posición
                        x = MARGEN_IZQUIERDO + columna * (ANCHO_ETIQUETA + ESPACIO_HORIZONTAL)
                        y = letter[1] - MARGEN_SUPERIOR - (fila + 1) * (ALTO_ETIQUETA + ESPACIO_VERTICAL)
                        
                        # === NOMBRE DEL PRODUCTO (ARRIBA) ===
                        pdf.setFont("Helvetica-Bold", 8)
                        nombre_truncado = nombre_producto[:22] + '...' if len(nombre_producto) > 22 else nombre_producto
                        pdf.drawCentredString(x + ANCHO_ETIQUETA / 2, y + 24*mm, nombre_truncado)
                        
                        # === CÓDIGO DE BARRAS (CENTRO) ===
                        barcode_width = 45 * mm
                        barcode_height = 10 * mm
                        barcode_x = x + (ANCHO_ETIQUETA - barcode_width) / 2
                        barcode_y = y + 12 * mm
                        
                        try:
                            barcode_obj = code128.Code128(codigo, barWidth=0.32*mm, barHeight=barcode_height)
                            barcode_obj.drawOn(pdf, barcode_x, barcode_y)
                        except:
                            pass
                        
                        # === CÓDIGO (DEBAJO DEL BARCODE) ===
                        pdf.setFont("Helvetica-Bold", 6)
                        pdf.drawCentredString(x + ANCHO_ETIQUETA / 2, y + 9*mm, codigo)
                        
                        # === PESO ===
                        pdf.setFont("Helvetica", 7)
                        pdf.drawCentredString(x + ANCHO_ETIQUETA / 2, y + 6*mm, f"{peso_unitario} {unidad}")
                        
                        # === PRECIO (DESTACADO) ===
                        pdf.setFont("Helvetica-Bold", 10)
                        pdf.drawCentredString(x + ANCHO_ETIQUETA / 2, y + 3*mm, f"${precio:.2f}")
                        
                        # === FECHA (ABAJO) ===
                        pdf.setFont("Helvetica", 5)
                        pdf.drawCentredString(x + ANCHO_ETIQUETA / 2, y + 0.5*mm, f"{fecha}")
                        
                        etiqueta_num += 1
                
                if etiqueta_num < cantidad:
                    pdf.showPage()
            
        elif producto_id:
            # ========================================
            # GENERAR ETIQUETAS PARA PRODUCTO NORMAL
            # ========================================
            producto = Producto.objects.get(id=producto_id)
            codigo = producto.codigo_barras
            nombre_producto = producto.nombre
            precio = float(producto.precio_unitario or 0)
            fecha = datetime.now().strftime('%d/%m/%Y')
            filename = f'etiquetas_{producto.codigo_barras}.pdf'
            
            etiqueta_num = 0
            
            while etiqueta_num < cantidad:
                for fila in range(ETIQUETAS_POR_COLUMNA):
                    for columna in range(ETIQUETAS_POR_FILA):
                        if etiqueta_num >= cantidad:
                            break
                        
                        # Calcular posición
                        x = MARGEN_IZQUIERDO + columna * (ANCHO_ETIQUETA + ESPACIO_HORIZONTAL)
                        y = letter[1] - MARGEN_SUPERIOR - (fila + 1) * (ALTO_ETIQUETA + ESPACIO_VERTICAL)
                        
                        # === NOMBRE DEL PRODUCTO (ARRIBA) ===
                        pdf.setFont("Helvetica-Bold", 8)
                        nombre_truncado = nombre_producto[:22] + '...' if len(nombre_producto) > 22 else nombre_producto
                        pdf.drawCentredString(x + ANCHO_ETIQUETA / 2, y + 24*mm, nombre_truncado)
                        
                        # === CÓDIGO DE BARRAS (CENTRO) ===
                        barcode_width = 45 * mm
                        barcode_height = 10 * mm
                        barcode_x = x + (ANCHO_ETIQUETA - barcode_width) / 2
                        barcode_y = y + 12 * mm
                        
                        try:
                            barcode_obj = code128.Code128(codigo, barWidth=0.32*mm, barHeight=barcode_height)
                            barcode_obj.drawOn(pdf, barcode_x, barcode_y)
                        except:
                            pass
                        
                        # === CÓDIGO (DEBAJO DEL BARCODE) ===
                        pdf.setFont("Helvetica-Bold", 6)
                        pdf.drawCentredString(x + ANCHO_ETIQUETA / 2, y + 9*mm, codigo)
                        
                        # === PRECIO (DESTACADO) ===
                        pdf.setFont("Helvetica-Bold", 11)
                        pdf.drawCentredString(x + ANCHO_ETIQUETA / 2, y + 4*mm, f"${precio:.2f}")
                        
                        # === FECHA (ABAJO) ===
                        pdf.setFont("Helvetica", 5)
                        pdf.drawCentredString(x + ANCHO_ETIQUETA / 2, y + 1*mm, f"{fecha}")
                        
                        etiqueta_num += 1
                
                if etiqueta_num < cantidad:
                    pdf.showPage()
            
        else:
            return JsonResponse({'error': 'Parámetros inválidos'}, status=400)
        
        # Finalizar PDF
        pdf.save()
        buffer.seek(0)
        
        # Retornar PDF
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Quintal.DoesNotExist:
        return JsonResponse({'error': 'Quintal no encontrado'}, status=404)
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)
    except Exception as e:
        import traceback
        print("=" * 80)
        print("ERROR GENERANDO PDF:")
        print(traceback.format_exc())
        print("=" * 80)
        return JsonResponse({
            'error': f'Error generando PDF: {str(e)}',
            'detalle': traceback.format_exc()
        }, status=500)
@ensure_csrf_cookie
def api_quintales_disponibles(request):
    """
    API para obtener quintales disponibles de un producto
    
    Parámetros GET:
        - producto_id (UUID): ID del producto del cual obtener quintales
    
    Retorna:
        JSON con lista de quintales disponibles y sus detalles
    
    Uso:
        GET /panel/api/inventario/quintales-disponibles/?producto_id=<uuid>
    """
    from apps.inventory_management.models import Quintal
    from django.http import JsonResponse
    
    # Obtener producto_id del request
    producto_id = request.GET.get('producto_id')
    
    # Validar que se envió el producto_id
    if not producto_id:
        return JsonResponse({
            'success': False,
            'error': 'Se requiere producto_id'
        }, status=400)
    
    try:
        # Obtener quintales disponibles del producto
        quintales = Quintal.objects.filter(
            producto_id=producto_id,
            estado='DISPONIBLE'
        ).select_related(
            'producto',
            'proveedor',
            'unidad_medida'
        ).order_by('fecha_ingreso')  # FIFO: primero los más antiguos
        
        # Construir lista de quintales para respuesta
        quintales_data = []
        for quintal in quintales:
            quintales_data.append({
                'id': str(quintal.id),
                'codigo_unico': quintal.codigo_unico,
                'peso_inicial': float(quintal.peso_inicial),
                'peso_actual': float(quintal.peso_actual),
                'unidad_medida': quintal.unidad_medida.abreviatura if quintal.unidad_medida else 'lb',
                'costo_por_unidad': float(quintal.costo_por_unidad),
                'fecha_ingreso': quintal.fecha_ingreso.strftime('%d/%m/%Y'),
                'proveedor': quintal.proveedor.nombre_comercial if quintal.proveedor else 'Sin proveedor',
                'lote': quintal.lote_proveedor or 'Sin lote',
            })
        
        # Retornar respuesta exitosa
        return JsonResponse({
            'success': True,
            'quintales': quintales_data,
            'total': len(quintales_data)
        })
        
    except Exception as e:
        # Manejar errores
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
@ensure_csrf_cookie
def api_procesar_entrada_unificada(request):
    """API para procesar entrada unificada de inventario - VERSIÓN CORREGIDA PARA QUINTALES"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    import json
    from django.db import transaction
    from django.db.models import Q
    from apps.inventory_management.models import (
        Producto, Categoria, Marca, Proveedor, ProductoNormal, MovimientoInventario,
        Quintal, UnidadMedida
    )
    from apps.authentication.models import Usuario
    from apps.system_configuration.models import ConfiguracionSistema
    from decimal import Decimal
    from django.utils import timezone
    
    # ✅ FUNCIÓN HELPER PARA CONVERTIR VALORES A DECIMAL DE FORMA SEGURA
    def safe_decimal(value, default='0.00'):
        """Convierte un valor a Decimal de forma segura, manejando None, '', 'None', etc."""
        if value is None or value == '' or value == 'None':
            return Decimal(default)
        try:
            return Decimal(str(value))
        except (ValueError, TypeError):
            return Decimal(default)
    
    try:
        # Leer datos del FormData
        productos_json = request.POST.get('productos')
        
        if not productos_json:
            return JsonResponse({
                'success': False,
                'error': 'No se recibieron datos de productos'
            })
        
        productos_data = json.loads(productos_json)
        
        print("=" * 80)
        print(f"📦 PRODUCTOS RECIBIDOS: {len(productos_data)}")
        print("=" * 80)
        
        if not productos_data:
            return JsonResponse({
                'success': False,
                'error': 'No hay productos para procesar'
            })
        
        # Obtener usuario del sistema
        usuario = Usuario.objects.filter(rol__codigo__in=['ADMIN', 'SUPERVISOR']).first()
        if not usuario:
            usuario = Usuario.objects.first()
        
        if not usuario:
            return JsonResponse({
                'success': False,
                'error': 'No hay usuarios en el sistema'
            })
        
        productos_creados = 0
        productos_reabastecidos = 0
        quintales_creados = 0
        quintales_reabastecidos = 0
        codigos_generados = []
        errores = []
        
        with transaction.atomic():
            for idx, prod_data in enumerate(productos_data):
                print(f"\n{'='*60}")
                print(f"🔄 PROCESANDO PRODUCTO {idx + 1}/{len(productos_data)}")
                print(f"📝 Datos: {prod_data}")
                
                try:
                    # Validar datos requeridos
                    if not prod_data.get('nombre'):
                        error_msg = f"Producto {idx + 1}: Falta el nombre"
                        print(f"❌ {error_msg}")
                        errores.append(error_msg)
                        continue
                    
                    # Obtener marca
                    marca_id = prod_data.get('marca_id')
                    if not marca_id:
                        error_msg = f"Producto {idx + 1}: Falta la marca"
                        print(f"❌ {error_msg}")
                        errores.append(error_msg)
                        continue
                    
                    try:
                        marca = Marca.objects.get(id=marca_id)
                        print(f"✅ Marca encontrada: {marca.nombre}")
                    except Marca.DoesNotExist:
                        error_msg = f"Producto {idx + 1}: Marca no encontrada"
                        print(f"❌ {error_msg}")
                        errores.append(error_msg)
                        continue
                    
                    # Obtener categoría
                    if not prod_data.get('categoria_id'):
                        error_msg = f"Producto {idx + 1}: Falta la categoría"
                        print(f"❌ {error_msg}")
                        errores.append(error_msg)
                        continue
                    
                    try:
                        categoria = Categoria.objects.get(id=prod_data['categoria_id'])
                        print(f"✅ Categoría encontrada: {categoria.nombre}")
                    except Categoria.DoesNotExist:
                        error_msg = f"Producto {idx + 1}: Categoría no encontrada"
                        print(f"❌ {error_msg}")
                        errores.append(error_msg)
                        continue
                    
                    # Obtener proveedor
                    if not prod_data.get('proveedor_id'):
                        error_msg = f"Producto {idx + 1}: Falta el proveedor"
                        print(f"❌ {error_msg}")
                        errores.append(error_msg)
                        continue
                    
                    try:
                        proveedor = Proveedor.objects.get(id=prod_data['proveedor_id'])
                        print(f"✅ Proveedor encontrado: {proveedor.nombre_comercial}")
                    except Proveedor.DoesNotExist:
                        error_msg = f"Producto {idx + 1}: Proveedor no encontrado"
                        print(f"❌ {error_msg}")
                        errores.append(error_msg)
                        continue
                    
                    tipo_inventario = prod_data.get('tipo_inventario', 'NORMAL')
                    
                    # ✅ OBTENER ESTADO DE IVA (solo booleano)
                    aplica_impuestos = prod_data.get('aplica_impuestos', False)
                    
                    # Obtener porcentaje solo para logging
                    if aplica_impuestos:
                        config = ConfiguracionSistema.get_config()
                        iva_porcentaje = config.porcentaje_iva
                        print(f"✅ IVA activado: {iva_porcentaje}%")
                    else:
                        iva_porcentaje = Decimal('0.00')
                        print(f"⚠️ IVA desactivado")
                    
                    # ========================================
                    # 🌾 PROCESAR QUINTAL
                    # ========================================
                    if tipo_inventario == 'QUINTAL':
                        print(f"🌾 PROCESANDO QUINTAL")
                        
                        # Obtener unidad de medida
                        unidad_codigo = prod_data.get('unidad_medida')
                        if not unidad_codigo:
                            error_msg = f"Producto {idx + 1}: Falta la unidad de medida"
                            print(f"❌ {error_msg}")
                            errores.append(error_msg)
                            continue
                        
                        try:
                            # Buscar por abreviatura o nombre
                            unidad_medida = UnidadMedida.objects.filter(
                                Q(abreviatura=unidad_codigo) | Q(nombre__icontains=unidad_codigo)
                            ).first()
                            
                            if not unidad_medida:
                                error_msg = f"Producto {idx + 1}: Unidad de medida no encontrada: {unidad_codigo}"
                                print(f"❌ {error_msg}")
                                errores.append(error_msg)
                                continue
                            
                            print(f"✅ Unidad de medida: {unidad_medida.nombre}")
                        except Exception as e:
                            error_msg = f"Producto {idx + 1}: Error buscando unidad: {str(e)}"
                            print(f"❌ {error_msg}")
                            errores.append(error_msg)
                            continue
                        
                        # ✅ USAR safe_decimal PARA CONVERSIONES
                        peso_inicial = safe_decimal(prod_data.get('cantidad', 0))
                        costo_unitario_base = safe_decimal(prod_data.get('costo_unitario', 0))
                        costo_total = costo_unitario_base * peso_inicial
                        precio_por_unidad_peso = safe_decimal(prod_data.get('precio_venta', 0))
                        
                        # ✅ BUSCAR SI EL PRODUCTO YA EXISTE
                        nombre_producto = prod_data['nombre'].strip()
                        producto_existente = Producto.objects.filter(
                            nombre__iexact=nombre_producto,
                            marca=marca,
                            categoria=categoria,
                            tipo_inventario='QUINTAL',
                            activo=True
                        ).first()
                        
                        # ✅ OBTENER IMAGEN SI EXISTE
                        imagen_file = None
                        if prod_data.get('tiene_imagen'):
                            imagen_key = f'imagen_{idx}'
                            if imagen_key in request.FILES:
                                imagen_file = request.FILES[imagen_key]
                                print(f"📷 Imagen recibida: {imagen_file.name}")
                        
                        if not producto_existente:
                            # ✅ CREAR PRODUCTO NUEVO
                            from apps.inventory_management.utils.barcode_generator import BarcodeGenerator
                            codigo_barras = BarcodeGenerator.generar_codigo_producto()
                            
                            producto_data = {
                                'codigo_barras': codigo_barras,
                                'nombre': nombre_producto,
                                'descripcion': prod_data.get('descripcion', ''),
                                'marca': marca,
                                'categoria': categoria,
                                'tipo_inventario': 'QUINTAL',
                                'unidad_medida_base': unidad_medida,
                                'precio_por_unidad_peso': precio_por_unidad_peso,
                                'aplica_impuestos': aplica_impuestos,  # ✅ Solo booleano
                                'activo': True,
                                'usuario_registro': usuario
                            }
                            
                            if imagen_file:
                                producto_data['imagen'] = imagen_file
                            
                            producto = Producto.objects.create(**producto_data)
                            print(f"✅ Producto QUINTAL creado: {producto.nombre}")
                            print(f"   Aplica IVA: {producto.aplica_impuestos} ({iva_porcentaje}% del sistema)")
                            productos_creados += 1
                        else:
                            producto = producto_existente
                            print(f"✅ Producto QUINTAL existente: {producto.nombre}")
                            
                            # ✅ Actualizar precio y aplica_impuestos si cambió
                            actualizado = False
                            if producto.precio_por_unidad_peso != precio_por_unidad_peso:
                                producto.precio_por_unidad_peso = precio_por_unidad_peso
                                actualizado = True
                            
                            if producto.aplica_impuestos != aplica_impuestos:
                                producto.aplica_impuestos = aplica_impuestos
                                actualizado = True
                                print(f"   Aplica IVA actualizado: {aplica_impuestos}")
                            
                            if actualizado:
                                producto.save()
                            
                            productos_reabastecidos += 1
                        
                        # ✅ CREAR QUINTAL
                        from apps.inventory_management.utils.barcode_generator import BarcodeGenerator
                        codigo_quintal = BarcodeGenerator.generar_codigo_quintal(producto)
                        
                        quintal = Quintal.objects.create(
                            codigo_quintal=codigo_quintal,
                            producto=producto,
                            proveedor=proveedor,
                            peso_inicial=peso_inicial,
                            peso_actual=peso_inicial,
                            unidad_medida=unidad_medida,
                            costo_total=costo_total,
                            costo_por_unidad=costo_total / peso_inicial if peso_inicial > 0 else Decimal('0'),
                            fecha_ingreso=timezone.now(),
                            fecha_vencimiento=prod_data.get('fecha_vencimiento') or None,
                            usuario_registro=usuario,
                            estado='DISPONIBLE'
                        )
                        
                        print(f"✅ QUINTAL CREADO: {quintal.codigo_quintal}")
                        print(f"   Peso: {quintal.peso_inicial} {quintal.unidad_medida.abreviatura}")
                        print(f"   Costo: ${quintal.costo_total}")
                        quintales_creados += 1
                        
                        # ✅ GENERAR CÓDIGOS DE BARRAS
                        cantidad_etiquetas = int(peso_inicial)  # 1 etiqueta por unidad
                        
                        codigos_generados.append({
                            'producto_id': str(producto.id),
                            'tipo': 'QUINTAL',
                            'codigo_base': quintal.codigo_quintal,
                            'producto_nombre': producto.nombre,
                            'cantidad_codigos': cantidad_etiquetas,
                            'cantidad_stock': float(peso_inicial),
                            'unidad_medida': unidad_medida.abreviatura,
                            'pdf_url': f'/panel/api/inventario/generar-pdf-codigos/?quintal_id={quintal.id}&cantidad={cantidad_etiquetas}',
                            'tipo_operacion': 'NUEVO_QUINTAL'
                        })
                    
                    # ========================================
                    # 📦 PROCESAR PRODUCTO NORMAL
                    # ========================================
                    else:
                        print(f"📦 PROCESANDO PRODUCTO NORMAL")
                        
                        nombre_producto = prod_data['nombre'].strip()
                        producto_existente = Producto.objects.filter(
                            nombre__iexact=nombre_producto,
                            marca=marca,
                            categoria=categoria,
                            tipo_inventario='NORMAL',
                            activo=True
                        ).first()
                        
                        # ✅ USAR safe_decimal PARA TODAS LAS CONVERSIONES
                        cantidad = int(prod_data.get('cantidad', 0))
                        costo_unitario = safe_decimal(prod_data.get('costo_unitario', '0'))
                        precio_venta = safe_decimal(prod_data.get('precio_venta', '0'))
                        
                        # Obtener imagen
                        imagen_file = None
                        if prod_data.get('tiene_imagen'):
                            imagen_key = f'imagen_{idx}'
                            if imagen_key in request.FILES:
                                imagen_file = request.FILES[imagen_key]
                        
                        if producto_existente:
                            # REABASTECER
                            producto = producto_existente
                            
                            # ✅ Actualizar precio y aplica_impuestos
                            actualizado = False
                            if producto.precio_venta != precio_venta:
                                producto.precio_venta = precio_venta
                                actualizado = True
                            
                            if producto.aplica_impuestos != aplica_impuestos:
                                producto.aplica_impuestos = aplica_impuestos
                                actualizado = True
                                print(f"   Aplica IVA actualizado: {aplica_impuestos}")
                            
                            if imagen_file:
                                if producto.imagen:
                                    producto.imagen.delete(save=False)
                                producto.imagen = imagen_file
                                actualizado = True
                            
                            if actualizado:
                                producto.save()
                            
                            # Reabastecer inventario
                            try:
                                producto_normal = producto.inventario_normal
                                stock_antes = producto_normal.stock_actual
                                
                                producto_normal.stock_actual += cantidad
                                
                                # Costo promedio ponderado
                                if stock_antes > 0:
                                    costo_total_anterior = stock_antes * producto_normal.costo_unitario
                                    costo_total_nuevo = cantidad * costo_unitario
                                    stock_total = stock_antes + cantidad
                                    producto_normal.costo_unitario = (costo_total_anterior + costo_total_nuevo) / stock_total
                                else:
                                    producto_normal.costo_unitario = costo_unitario
                                
                                producto_normal.fecha_ultima_entrada = timezone.now()
                                
                                if prod_data.get('lote'):
                                    producto_normal.lote = prod_data.get('lote')
                                if prod_data.get('fecha_vencimiento'):
                                    producto_normal.fecha_vencimiento = prod_data.get('fecha_vencimiento')
                                
                                producto_normal.save()
                                
                                # Registrar movimiento
                                MovimientoInventario.objects.create(
                                    producto_normal=producto_normal,
                                    tipo_movimiento='ENTRADA_COMPRA',
                                    cantidad=cantidad,
                                    stock_antes=stock_antes,
                                    stock_despues=producto_normal.stock_actual,
                                    costo_unitario=costo_unitario,
                                    costo_total=cantidad * costo_unitario,
                                    usuario=usuario,
                                    observaciones=f"Reabastecimiento - {prod_data.get('lote', 'Sin lote')}"
                                )
                                
                                productos_reabastecidos += 1
                                
                            except ProductoNormal.DoesNotExist:
                                producto_normal = ProductoNormal.objects.create(
                                    producto=producto,
                                    stock_actual=cantidad,
                                    stock_minimo=10,
                                    costo_unitario=costo_unitario,
                                    lote=prod_data.get('lote', ''),
                                    fecha_vencimiento=prod_data.get('fecha_vencimiento') or None,
                                    fecha_ultima_entrada=timezone.now()
                                )
                                
                                MovimientoInventario.objects.create(
                                    producto_normal=producto_normal,
                                    tipo_movimiento='ENTRADA_COMPRA',
                                    cantidad=cantidad,
                                    stock_antes=0,
                                    stock_despues=cantidad,
                                    costo_unitario=costo_unitario,
                                    costo_total=cantidad * costo_unitario,
                                    usuario=usuario,
                                    observaciones=f"Primer inventario - {prod_data.get('lote', 'Sin lote')}"
                                )
                                
                                productos_reabastecidos += 1
                        
                        else:
                            # CREAR NUEVO
                            from apps.inventory_management.utils.barcode_generator import BarcodeGenerator
                            codigo_barras = BarcodeGenerator.generar_codigo_producto()
                            
                            producto_data = {
                                'codigo_barras': codigo_barras,
                                'nombre': nombre_producto,
                                'descripcion': prod_data.get('descripcion', ''),
                                'marca': marca,
                                'categoria': categoria,
                                'tipo_inventario': 'NORMAL',
                                'precio_venta': precio_venta,
                                'aplica_impuestos': aplica_impuestos,  # ✅ Solo booleano
                                'activo': True,
                                'usuario_registro': usuario
                            }
                            
                            if imagen_file:
                                producto_data['imagen'] = imagen_file
                            
                            producto = Producto.objects.create(**producto_data)
                            print(f"✅ Producto NORMAL creado: {producto.nombre}")
                            print(f"   Aplica IVA: {producto.aplica_impuestos} ({iva_porcentaje}% del sistema)")
                            
                            producto_normal = ProductoNormal.objects.create(
                                producto=producto,
                                stock_actual=cantidad,
                                stock_minimo=10,
                                costo_unitario=costo_unitario,
                                lote=prod_data.get('lote', ''),
                                fecha_vencimiento=prod_data.get('fecha_vencimiento') or None,
                                fecha_ultima_entrada=timezone.now()
                            )
                            
                            MovimientoInventario.objects.create(
                                producto_normal=producto_normal,
                                tipo_movimiento='ENTRADA_COMPRA',
                                cantidad=cantidad,
                                stock_antes=0,
                                stock_despues=cantidad,
                                costo_unitario=costo_unitario,
                                costo_total=cantidad * costo_unitario,
                                usuario=usuario,
                                observaciones=f"Entrada inicial - {prod_data.get('lote', 'Sin lote')}"
                            )
                            
                            productos_creados += 1
                        
                        # Generar códigos
                        cantidad_codigos = int(prod_data.get('cantidad_codigos', cantidad))
                        
                        codigos_generados.append({
                            'producto_id': str(producto.id),
                            'tipo': 'NORMAL',
                            'codigo_base': producto.codigo_barras,
                            'producto_nombre': producto.nombre,
                            'cantidad_codigos': cantidad_codigos,
                            'cantidad_stock': cantidad,
                            'unidad_medida': 'UNIDAD',
                            'pdf_url': f'/panel/api/inventario/generar-pdf-codigos/?producto_id={producto.id}&cantidad={cantidad_codigos}',
                            'tipo_operacion': 'REABASTECIMIENTO' if producto_existente else 'NUEVO'
                        })
                    
                except Exception as e:
                    error_msg = f"Producto {idx + 1} ({prod_data.get('nombre', 'Sin nombre')}): {str(e)}"
                    print(f"❌ ERROR: {error_msg}")
                    import traceback
                    traceback.print_exc()
                    errores.append(error_msg)
                    continue
        
        print("\n" + "=" * 80)
        print(f"✅ RESUMEN FINAL:")
        print(f"   - Productos NUEVOS: {productos_creados}")
        print(f"   - Productos REABASTECIDOS: {productos_reabastecidos}")
        print(f"   - Quintales CREADOS: {quintales_creados}")
        print(f"   - Códigos generados: {len(codigos_generados)}")
        print(f"   - Errores: {len(errores)}")
        print("=" * 80)
        
        return JsonResponse({
            'success': True,
            'productos_creados': productos_creados,
            'productos_reabastecidos': productos_reabastecidos,
            'quintales_creados': quintales_creados,
            'codigos_generados': codigos_generados,
            'errores': errores,
            'mensaje': f'{productos_creados} nuevos, {productos_reabastecidos} reabastecidos, {quintales_creados} quintales'
        })
        
    except json.JSONDecodeError as e:
        print(f"❌ ERROR JSON: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Error al decodificar JSON: {str(e)}'
        }, status=400)
    except Exception as e:
        import traceback
        error_detalle = traceback.format_exc()
        print("=" * 80)
        print("❌ ERROR CRÍTICO:")
        print(error_detalle)
        print("=" * 80)
        return JsonResponse({
            'success': False,
            'error': f'Error al procesar inventario: {str(e)}'
        }, status=500)

# ============================================================================
# MOVIMIENTOS DE INVENTARIO
# ============================================================================

@ensure_csrf_cookie
@auth_required
def movimientos_inventario_view(request):
    """Vista principal de movimientos de inventario con filtros"""
    from apps.inventory_management.models import MovimientoInventario, Producto
    from apps.authentication.models import Usuario
    from django.db.models import Q, Sum, Count
    from django.core.paginator import Paginator
    from decimal import Decimal
    
    # Base queryset
    movimientos = MovimientoInventario.objects.select_related(
        'producto_normal__producto',
        'usuario'
    ).all().order_by('-fecha_movimiento')
    
    # ========================================
    # FILTROS
    # ========================================
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    tipo_movimiento = request.GET.get('tipo_movimiento', '')
    producto_id = request.GET.get('producto', '')
    usuario_id = request.GET.get('usuario', '')
    search = request.GET.get('search', '')
    
    if fecha_inicio:
        movimientos = movimientos.filter(fecha_movimiento__date__gte=fecha_inicio)
    
    if fecha_fin:
        movimientos = movimientos.filter(fecha_movimiento__date__lte=fecha_fin)
    
    if tipo_movimiento:
        movimientos = movimientos.filter(tipo_movimiento=tipo_movimiento)
    
    if producto_id:
        movimientos = movimientos.filter(producto_normal__producto_id=producto_id)
    
    if usuario_id:
        movimientos = movimientos.filter(usuario_id=usuario_id)
    
    if search:
        movimientos = movimientos.filter(
            Q(producto_normal__producto__nombre__icontains=search) |
            Q(observaciones__icontains=search)
        )
    
    # ========================================
    # ESTADÍSTICAS
    # ========================================
    total_movimientos = movimientos.count()
    
    entradas = movimientos.filter(
        tipo_movimiento__in=['ENTRADA_COMPRA', 'ENTRADA_AJUSTE', 'ENTRADA_DEVOLUCION']
    ).aggregate(
        total=Sum('cantidad'),
        count=Count('id')
    )
    
    salidas = movimientos.filter(
        tipo_movimiento__in=['SALIDA_VENTA', 'SALIDA_AJUSTE', 'SALIDA_MERMA', 'SALIDA_DEVOLUCION']
    ).aggregate(
        total=Sum('cantidad'),
        count=Count('id')
    )
    
    total_entradas = abs(entradas['total'] or 0)
    count_entradas = entradas['count'] or 0
    
    total_salidas = abs(salidas['total'] or 0)
    count_salidas = salidas['count'] or 0
    
    # Movimientos del día
    hoy = timezone.now().date()
    movimientos_hoy = MovimientoInventario.objects.filter(
        fecha_movimiento__date=hoy
    ).count()
    
    # ========================================
    # PAGINACIÓN
    # ========================================
    paginator = Paginator(movimientos, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # ========================================
    # DATOS PARA FORMULARIOS
    # ========================================
    productos = Producto.objects.filter(
        tipo_inventario='NORMAL',
        activo=True
    ).order_by('nombre')
    
    usuarios = Usuario.objects.filter(is_active=True).order_by('nombres')
    
    # Tipos de movimiento
    tipos_movimiento = [
        ('', 'Todos'),
        ('ENTRADA_COMPRA', 'Entrada por Compra'),
        ('ENTRADA_DEVOLUCION', 'Entrada por Devolución'),
        ('ENTRADA_AJUSTE', 'Entrada por Ajuste'),
        ('SALIDA_VENTA', 'Salida por Venta'),
        ('SALIDA_DEVOLUCION', 'Salida por Devolución'),
        ('SALIDA_MERMA', 'Salida por Merma'),
        ('SALIDA_AJUSTE', 'Salida por Ajuste'),
    ]
    
    context = {
        'page_obj': page_obj,
        'movimientos': page_obj,
        'total_movimientos': total_movimientos,
        'total_entradas': total_entradas,
        'count_entradas': count_entradas,
        'total_salidas': total_salidas,
        'count_salidas': count_salidas,
        'movimientos_hoy': movimientos_hoy,
        'productos': productos,
        'usuarios': usuarios,
        'tipos_movimiento': tipos_movimiento,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'tipo_movimiento_selected': tipo_movimiento,
        'producto_selected': producto_id,
        'usuario_selected': usuario_id,
        'search': search,
    }
    
    return render(request, 'custom_admin/inventario/movimientos_list.html', context)


@ensure_csrf_cookie
@auth_required
def movimiento_detalle_api(request, pk):
    """API para obtener detalles de un movimiento de inventario"""
    from apps.inventory_management.models import MovimientoInventario
    from django.http import JsonResponse
    
    try:
        movimiento = MovimientoInventario.objects.select_related(
            'producto_normal__producto',
            'usuario'
        ).get(pk=pk)
        
        data = {
            'success': True,
            'movimiento': {
                'fecha_movimiento': movimiento.fecha_movimiento.strftime('%d/%m/%Y %H:%M'),
                'tipo_movimiento': movimiento.get_tipo_movimiento_display(),
                'tipo_movimiento_code': movimiento.tipo_movimiento,
                'producto_nombre': movimiento.producto_normal.producto.nombre,
                'producto_codigo': movimiento.producto_normal.producto.codigo_barras or 'Sin código',
                'cantidad': str(movimiento.cantidad),
                'stock_antes': str(movimiento.stock_antes),
                'stock_despues': str(movimiento.stock_despues),
                'costo_unitario': str(movimiento.costo_unitario),
                'costo_total': str(movimiento.costo_total),
                'usuario': movimiento.usuario.get_full_name() if movimiento.usuario else 'Sistema',
                'referencia': movimiento.venta.numero_venta if movimiento.venta else 'N/A',
                'observaciones': movimiento.observaciones or 'Sin observaciones',
            }
        }
        return JsonResponse(data)
        
    except MovimientoInventario.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Movimiento no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@ensure_csrf_cookie
@auth_required
def exportar_movimiento_excel(request, pk):
    """Exporta un movimiento individual a Excel"""
    from apps.inventory_management.models import MovimientoInventario
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    try:
        movimiento = MovimientoInventario.objects.select_related(
            'producto_normal__producto',
            'usuario'
        ).get(pk=pk)
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimiento"
        
        # Encabezado
        ws['A1'] = 'DETALLE DE MOVIMIENTO DE INVENTARIO'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # Información
        row = 3
        ws[f'A{row}'] = 'Fecha:'
        ws[f'B{row}'] = movimiento.fecha_movimiento.strftime('%d/%m/%Y %H:%M')
        row += 1
        
        ws[f'A{row}'] = 'Tipo:'
        ws[f'B{row}'] = movimiento.get_tipo_movimiento_display()
        row += 1
        
        ws[f'A{row}'] = 'Producto:'
        ws[f'B{row}'] = movimiento.producto_normal.producto.nombre
        row += 1
        
        ws[f'A{row}'] = 'Cantidad:'
        ws[f'B{row}'] = movimiento.cantidad
        row += 1
        
        ws[f'A{row}'] = 'Stock Antes:'
        ws[f'B{row}'] = movimiento.stock_antes
        row += 1
        
        ws[f'A{row}'] = 'Stock Después:'
        ws[f'B{row}'] = movimiento.stock_despues
        row += 1
        
        ws[f'A{row}'] = 'Usuario:'
        ws[f'B{row}'] = movimiento.usuario.get_full_name() if movimiento.usuario else 'Sistema'
        
        # Ajustar columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="movimiento_{pk}.xlsx"'
        wb.save(response)
        return response
        
    except MovimientoInventario.DoesNotExist:
        messages.error(request, 'Movimiento no encontrado')
        return redirect('custom_admin:movimientos_inventario')


@ensure_csrf_cookie
@auth_required
def exportar_movimiento_pdf(request, pk):
    """Exporta un movimiento individual a PDF"""
    from apps.inventory_management.models import MovimientoInventario
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from django.http import HttpResponse
    
    try:
        movimiento = MovimientoInventario.objects.select_related(
            'producto_normal__producto',
            'usuario'
        ).get(pk=pk)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="movimiento_{pk}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title = Paragraph('<b>DETALLE DE MOVIMIENTO DE INVENTARIO</b>', styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Datos
        data = [
            ['Fecha:', movimiento.fecha_movimiento.strftime('%d/%m/%Y %H:%M')],
            ['Tipo:', movimiento.get_tipo_movimiento_display()],
            ['Producto:', movimiento.producto_normal.producto.nombre],
            ['Cantidad:', str(movimiento.cantidad)],
            ['Stock Antes:', str(movimiento.stock_antes)],
            ['Stock Después:', str(movimiento.stock_despues)],
            ['Costo Unitario:', f'${movimiento.costo_unitario}'],
            ['Costo Total:', f'${movimiento.costo_total}'],
            ['Usuario:', movimiento.usuario.get_full_name() if movimiento.usuario else 'Sistema'],
        ]
        
        table = Table(data, colWidths=[150, 350])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        return response
        
    except MovimientoInventario.DoesNotExist:
        messages.error(request, 'Movimiento no encontrado')
        return redirect('custom_admin:movimientos_inventario')


@ensure_csrf_cookie
@auth_required
def exportar_movimientos_excel_general(request):
    """Exportar todos los movimientos filtrados a Excel"""
    from apps.inventory_management.models import MovimientoInventario
    from django.http import HttpResponse
    from django.db.models import Q
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime
    from io import BytesIO
    
    # Obtener filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_movimiento = request.GET.get('tipo_movimiento')
    producto_id = request.GET.get('producto')
    usuario_id = request.GET.get('usuario')
    search = request.GET.get('search')
    
    # Filtrar movimientos
    movimientos = MovimientoInventario.objects.select_related(
        'producto_normal__producto',
        'usuario'
    ).all().order_by('-fecha_movimiento')
    
    if fecha_inicio:
        movimientos = movimientos.filter(fecha_movimiento__date__gte=fecha_inicio)
    if fecha_fin:
        movimientos = movimientos.filter(fecha_movimiento__date__lte=fecha_fin)
    if tipo_movimiento:
        movimientos = movimientos.filter(tipo_movimiento=tipo_movimiento)
    if producto_id:
        movimientos = movimientos.filter(producto_normal__producto_id=producto_id)
    if usuario_id:
        movimientos = movimientos.filter(usuario_id=usuario_id)
    if search:
        movimientos = movimientos.filter(
            Q(producto_normal__producto__nombre__icontains=search) |
            Q(observaciones__icontains=search)
        )
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    
    # Estilos
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # TÍTULO
    ws.merge_cells('A1:J1')
    ws['A1'] = f"REPORTE DE MOVIMIENTOS DE INVENTARIO - {datetime.now().strftime('%d/%m/%Y')}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # HEADERS
    headers = ['Fecha', 'Tipo', 'Producto', 'Cantidad', 'Stock Antes', 
               'Stock Después', 'Costo Unit.', 'Costo Total', 'Usuario', 'Observaciones']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # DATOS
    row = 4
    for mov in movimientos:
        ws.cell(row=row, column=1, value=mov.fecha_movimiento.strftime('%d/%m/%Y %H:%M')).border = border
        ws.cell(row=row, column=2, value=mov.get_tipo_movimiento_display()).border = border
        ws.cell(row=row, column=3, value=mov.producto_normal.producto.nombre).border = border
        ws.cell(row=row, column=4, value=float(mov.cantidad)).border = border
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=float(mov.stock_antes)).border = border
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=6, value=float(mov.stock_despues)).border = border
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=7, value=float(mov.costo_unitario)).border = border
        ws.cell(row=row, column=7).number_format = '$#,##0.00'
        ws.cell(row=row, column=8, value=float(mov.costo_total)).border = border
        ws.cell(row=row, column=8).number_format = '$#,##0.00'
        ws.cell(row=row, column=9, value=mov.usuario.get_full_name() if mov.usuario else 'Sistema').border = border
        ws.cell(row=row, column=10, value=mov.observaciones or 'N/A').border = border
        row += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 25
    
    # Preparar respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"movimientos_inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


@ensure_csrf_cookie
@auth_required
def exportar_movimientos_pdf_general(request):
    """Exportar todos los movimientos filtrados a PDF"""
    from apps.inventory_management.models import MovimientoInventario
    from django.http import HttpResponse
    from django.db.models import Q
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    from datetime import datetime
    
    # Obtener filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_movimiento = request.GET.get('tipo_movimiento')
    producto_id = request.GET.get('producto')
    usuario_id = request.GET.get('usuario')
    search = request.GET.get('search')
    
    # Filtrar movimientos
    movimientos = MovimientoInventario.objects.select_related(
        'producto_normal__producto',
        'usuario'
    ).all().order_by('-fecha_movimiento')
    
    if fecha_inicio:
        movimientos = movimientos.filter(fecha_movimiento__date__gte=fecha_inicio)
    if fecha_fin:
        movimientos = movimientos.filter(fecha_movimiento__date__lte=fecha_fin)
    if tipo_movimiento:
        movimientos = movimientos.filter(tipo_movimiento=tipo_movimiento)
    if producto_id:
        movimientos = movimientos.filter(producto_normal__producto_id=producto_id)
    if usuario_id:
        movimientos = movimientos.filter(usuario_id=usuario_id)
    if search:
        movimientos = movimientos.filter(
            Q(producto_normal__producto__nombre__icontains=search) |
            Q(observaciones__icontains=search)
        )
    
    # Limitar a 200 registros para PDF
    movimientos = movimientos[:200]
    
    # Crear buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    title = Paragraph(f"REPORTE DE MOVIMIENTOS DE INVENTARIO - {datetime.now().strftime('%d/%m/%Y')}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 10))
    
    # Tabla de movimientos
    data = [['Fecha', 'Tipo', 'Producto', 'Cant.', 'Stock Antes', 'Stock Después', 'Usuario']]
    
    for mov in movimientos:
        data.append([
            mov.fecha_movimiento.strftime('%d/%m/%Y'),
            mov.get_tipo_movimiento_display()[:15],
            mov.producto_normal.producto.nombre[:25],
            str(mov.cantidad),
            str(mov.stock_antes),
            str(mov.stock_despues),
            (mov.usuario.get_full_name() if mov.usuario else 'Sistema')[:20],
        ])
    
    table = Table(data, colWidths=[1*inch, 1.3*inch, 2*inch, 0.7*inch, 1*inch, 1*inch, 1.5*inch])
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Body
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 8),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
        ('ALIGN', (3, 1), (5, -1), 'CENTER'),
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    filename = f"movimientos_inventario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response


# ========================================
# FINANZAS
# ========================================

@ensure_csrf_cookie
@auth_required
def finanzas_dashboard_view(request):
    """Dashboard general de finanzas con resumen de cuentas por cobrar y pagar"""
    
    from apps.financial_management.models import ReporteCuentasPorCobrar, ReporteCuentasPorPagar
    
    # Resúmenes generales
    resumen_cobrar = ReporteCuentasPorCobrar.resumen_general()
    resumen_pagar = ReporteCuentasPorPagar.resumen_general()
    
    # Antigüedad de saldos
    antiguedad_cobrar = ReporteCuentasPorCobrar.antiguedad_saldos()
    antiguedad_pagar = ReporteCuentasPorPagar.antiguedad_saldos()
    
    # Cuentas más urgentes (próximas a vencer en 7 días)
    hoy = timezone.now().date()
    fecha_limite = hoy + timedelta(days=7)
    
    cuentas_cobrar_urgentes = CuentaPorCobrar.objects.filter(
        estado__in=['PENDIENTE', 'PARCIAL'],
        fecha_vencimiento__lte=fecha_limite,
        fecha_vencimiento__gte=hoy
    ).select_related('cliente').order_by('fecha_vencimiento')[:10]
    
    cuentas_pagar_urgentes = CuentaPorPagar.objects.filter(
        estado__in=['PENDIENTE', 'PARCIAL'],
        fecha_vencimiento__lte=fecha_limite,
        fecha_vencimiento__gte=hoy
    ).select_related('proveedor').order_by('fecha_vencimiento')[:10]
    
    # Cuentas vencidas
    cuentas_cobrar_vencidas = CuentaPorCobrar.objects.filter(
        estado='VENCIDA'
    ).select_related('cliente').order_by('fecha_vencimiento')[:10]
    
    cuentas_pagar_vencidas = CuentaPorPagar.objects.filter(
        estado='VENCIDA'
    ).select_related('proveedor').order_by('fecha_vencimiento')[:10]
    
    context = {
        'resumen_cobrar': resumen_cobrar,
        'resumen_pagar': resumen_pagar,
        'antiguedad_cobrar': antiguedad_cobrar,
        'antiguedad_pagar': antiguedad_pagar,
        'cuentas_cobrar_urgentes': cuentas_cobrar_urgentes,
        'cuentas_pagar_urgentes': cuentas_pagar_urgentes,
        'cuentas_cobrar_vencidas': cuentas_cobrar_vencidas,
        'cuentas_pagar_vencidas': cuentas_pagar_vencidas,
    }
    
    return render(request, 'custom_admin/finanzas/dashboard.html', context)



@ensure_csrf_cookie
@auth_required
def cajas_view(request):
    """Lista de cajas"""
    return render(request, 'custom_admin/finanzas/cajas_list.html')


@ensure_csrf_cookie
@auth_required
def caja_detail_view(request, pk):
    """Detalle de caja"""
    return render(request, 'custom_admin/finanzas/caja_detail.html', {'caja_id': pk})


@ensure_csrf_cookie
@auth_required
def arqueos_view(request):
    """Lista de arqueos"""
    return render(request, 'custom_admin/finanzas/arqueos_list.html')


@ensure_csrf_cookie
@auth_required
def caja_chica_view(request):
    """Lista de cajas chicas"""
    return render(request, 'custom_admin/finanzas/caja_chica_list.html')


# ========================================
# REPORTES
# ========================================

@ensure_csrf_cookie
@auth_required
def reportes_dashboard_view(request):
    """Dashboard de reportes"""
    return render(request, 'custom_admin/reportes/dashboard.html')


@ensure_csrf_cookie
@auth_required
def reporte_ventas_view(request):
    """Reporte de ventas"""
    return render(request, 'custom_admin/reportes/ventas.html')


@ensure_csrf_cookie
@auth_required
def reporte_inventario_view(request):
    """Reporte de inventario"""
    return render(request, 'custom_admin/reportes/inventario.html')


@ensure_csrf_cookie
@auth_required
def reporte_financiero_view(request):
    """Reporte financiero"""
    return render(request, 'custom_admin/reportes/financiero.html')


# ========================================
# ALERTAS
# ========================================

@ensure_csrf_cookie
@auth_required
def alertas_view(request):
    """Dashboard de alertas"""
    return render(request, 'custom_admin/alertas/dashboard.html')


# ========================================
# LOGS Y AUDITORÍA
# ========================================

@ensure_csrf_cookie
@auth_required
def logs_view(request):
    """Lista de logs"""
    return render(request, 'custom_admin/logs/list.html')


@ensure_csrf_cookie
@auth_required
def logs_accesos_view(request):
    """Logs de acceso"""
    return render(request, 'custom_admin/logs/accesos.html')


# ========================================
# SESIONES
# ========================================

@ensure_csrf_cookie
@auth_required
def sesiones_view(request):
    """Sesiones activas"""
    return render(request, 'custom_admin/sesiones/activas.html')


# ========================================
# CONFIGURACIÓN
# ========================================

@ensure_csrf_cookie
@auth_required
def configuracion_view(request):
    """Configuración general del sistema"""
    return render(request, 'custom_admin/configuracion/general.html')


@ensure_csrf_cookie
@auth_required
def config_empresa_view(request):
    """Configuración de empresa"""
    return render(request, 'custom_admin/configuracion/empresa.html')


@ensure_csrf_cookie
@auth_required
def config_facturacion_view(request):
    """Configuración de facturación"""
    return render(request, 'custom_admin/configuracion/facturacion.html')


# ========================================
# BÚSQUEDA GLOBAL
# ========================================

@ensure_csrf_cookie
@auth_required
def busqueda_view(request):
    """Búsqueda global"""
    query = request.GET.get('q', '')
    return render(request, 'custom_admin/busqueda/results.html', {'query': query})


# ========================================
# NOTIFICACIONES
# ========================================

@ensure_csrf_cookie
@auth_required
def notificaciones_view(request):
    """Centro de notificaciones"""
    return render(request, 'custom_admin/notificaciones/list.html')


# ========================================
# PERFIL
# ========================================

@ensure_csrf_cookie
@auth_required
def perfil_view(request):
    """Perfil del usuario actual"""
    return render(request, 'custom_admin/perfil/view.html')


@ensure_csrf_cookie
@auth_required
def perfil_editar_view(request):
    """Editar perfil"""
    return render(request, 'custom_admin/perfil/edit.html')


@ensure_csrf_cookie
@auth_required
def perfil_cambiar_password_view(request):
    """Cambiar contraseña"""
    return render(request, 'custom_admin/perfil/change_password.html')


# ========================================
# APIs MOCK PARA DASHBOARD
# ========================================

@ensure_csrf_cookie
def api_dashboard_stats(request):
    """Estadísticas del dashboard"""
    return JsonResponse({
        'ventas_hoy': 1250.50,
        'num_ventas': 15,
        'productos_total': 250,
        'alertas_criticas': 3,
        'ventas_semana': [
            {'fecha': '2025-10-27', 'total': 1500},
            {'fecha': '2025-10-28', 'total': 1800},
            {'fecha': '2025-10-29', 'total': 2100},
            {'fecha': '2025-10-30', 'total': 1900},
            {'fecha': '2025-10-31', 'total': 2300},
            {'fecha': '2025-11-01', 'total': 2500},
            {'fecha': '2025-11-02', 'total': 1250.50},
        ],
        'top_productos': [
            {'nombre': 'Arroz', 'cantidad': 50},
            {'nombre': 'Frijol', 'cantidad': 30},
            {'nombre': 'Azúcar', 'cantidad': 25},
        ],
        'ultimas_ventas': [
            {
                'numero_venta': 'VNT-2025-00001',
                'cliente': 'Juan Pérez',
                'total': 50.00,
                'estado': 'COMPLETADA'
            },
        ],
        'alertas': [
            {
                'titulo': 'Stock bajo en Arroz',
                'mensaje': 'Solo quedan 5 unidades',
                'prioridad': 'ALTA'
            },
        ]
    })


# ========================================
# UTILIDADES
# ========================================

def health_check_view(request):
    """Health check del panel"""
    return JsonResponse({
        'status': 'healthy',
        'timestamp': timezone.now().isoformat()
    })

@ensure_csrf_cookie
def api_alertas_count(request):
    """API para obtener el conteo de alertas activas"""
    try:
        from apps.stock_alert_system.models import Alerta
        
        # Contar alertas no leídas
        count = Alerta.objects.filter(
            resuelta=False,
            activa=True
        ).count()
        
        return JsonResponse({
            'success': True,
            'count': count
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'count': 0
        })

# API para quintales (agregar al final del archivo)
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from apps.inventory_management.models import Quintal
from apps.sales_management.quintal_service import QuintalSalesService
from decimal import Decimal
import json

def api_quintales_por_producto(request, producto_id):
    """Obtiene quintales disponibles de un producto - VERSIÓN CORREGIDA"""
    try:
        from apps.inventory_management.models import Quintal
        
        quintales = Quintal.objects.filter(
            producto_id=producto_id,
            estado='DISPONIBLE',
            peso_actual__gt=0
        ).select_related('producto', 'unidad_medida', 'proveedor').order_by('fecha_ingreso')  # FIFO
        
        data = []
        for q in quintales:
            data.append({
                'id': str(q.id),
                'codigo_unico': q.codigo_quintal,  # ✅ CORRECTO: es codigo_quintal, NO codigo_unico
                'peso_actual': float(q.peso_actual),
                'unidad_medida': q.unidad_medida.abreviatura if q.unidad_medida else 'lb',
                'costo_por_unidad': float(q.costo_por_unidad),
            })
        
        return JsonResponse({
            'success': True, 
            'quintales': data,
            'total': len(data)
        })
    except Exception as e:
        import traceback
        print(f"❌ Error en api_quintales_por_producto: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=500)

@csrf_exempt
def api_calcular_quintal(request):
    """Calcula conversión entre peso y dinero"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    try:
        data = json.loads(request.body)
        quintal_id = data.get('quintal_id')
        tipo = data.get('tipo')  # 'peso_a_dinero' o 'dinero_a_peso'
        valor = Decimal(str(data.get('valor', 0)))
        
        quintal = Quintal.objects.get(id=quintal_id)
        service = QuintalSalesService()
        
        if tipo == 'peso_a_dinero':
            dinero = service.calcular_dinero_por_peso(quintal, valor)
            return JsonResponse({
                'success': True,
                'peso': float(valor),
                'dinero': float(dinero),
                'unidad': quintal.unidad_medida.abreviatura
            })
        else:  # dinero_a_peso
            peso = service.calcular_peso_por_dinero(quintal, valor)
            # Validar disponibilidad
            if peso > quintal.peso_actual:
                return JsonResponse({
                    'success': False,
                    'error': f'Peso insuficiente. Máximo disponible: {quintal.peso_actual} {quintal.unidad_medida.abreviatura}'
                })
            return JsonResponse({
                'success': True,
                'peso': float(peso),
                'dinero': float(valor),
                'unidad': quintal.unidad_medida.abreviatura
            })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
# ============================================================================
# VISTAS DE FINANZAS - CAJAS
# ============================================================================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from apps.authentication.decorators import role_required_html
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.http import JsonResponse
from django.utils import timezone
from django.db import transaction
from decimal import Decimal
from datetime import datetime, timedelta

from apps.financial_management.models import (
    Caja, MovimientoCaja, ArqueoCaja,
    CajaChica, MovimientoCajaChica
)
from apps.authentication.decorators import role_required_html, role_required


@auth_required
def cajas_list(request):
    """Lista de cajas con filtros y estadísticas"""
    
    # Obtener todas las cajas
    cajas = Caja.objects.all().select_related('usuario_apertura', 'usuario_cierre')
    
    # Filtros
    estado = request.GET.get('estado', '')
    tipo = request.GET.get('tipo', '')
    search = request.GET.get('search', '')
    
    if estado:
        cajas = cajas.filter(estado=estado)
    
    if tipo:
        cajas = cajas.filter(tipo=tipo)
    
    if search:
        cajas = cajas.filter(
            Q(nombre__icontains=search) |
            Q(codigo__icontains=search)
        )
    
    # Estadísticas
    cajas_abiertas = Caja.objects.filter(estado='ABIERTA').count()
    cajas_cerradas = Caja.objects.filter(estado='CERRADA').count()
    total_efectivo = Caja.objects.filter(estado='ABIERTA').aggregate(
        total=Sum('monto_actual')
    )['total'] or 0
    
    # Movimientos de hoy
    hoy = timezone.now().date()
    movimientos_hoy = MovimientoCaja.objects.filter(
        fecha_movimiento__date=hoy
    ).count()
    
    # Paginación
    paginator = Paginator(cajas, 12)  # 12 cajas por página
    page_number = request.GET.get('page')
    cajas_page = paginator.get_page(page_number)
    
    context = {
        'cajas': cajas_page,
        'cajas_abiertas': cajas_abiertas,
        'cajas_cerradas': cajas_cerradas,
        'total_efectivo': total_efectivo,
        'movimientos_hoy': movimientos_hoy,
    }
    
    return render(request, 'custom_admin/finanzas/cajas_list.html', context)


@auth_required
def crear_caja(request):
    """Crear nueva caja"""
    
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            codigo = request.POST.get('codigo')
            tipo = request.POST.get('tipo')
            monto_apertura = Decimal(request.POST.get('monto_apertura', '0'))
            requiere_autorizacion = request.POST.get('requiere_autorizacion_cierre') == 'on'
            
            # Validar que el código no exista
            if Caja.objects.filter(codigo=codigo).exists():
                messages.error(request, f'Ya existe una caja con el código {codigo}')
                return redirect('custom_admin:cajas_list')
            
            # Crear la caja
            caja = Caja.objects.create(
                nombre=nombre,
                codigo=codigo,
                tipo=tipo,
                monto_apertura=monto_apertura,
                requiere_autorizacion_cierre=requiere_autorizacion,
                estado='CERRADA',
                monto_actual=Decimal('0'),
                activa=True
            )
            
            messages.success(request, f'Caja {nombre} creada exitosamente')
            
        except Exception as e:
            messages.error(request, f'Error al crear caja: {str(e)}')
    
    return redirect('custom_admin:cajas_list')


@auth_required
def abrir_caja(request, caja_id):
    """Abrir una caja"""
    
    caja = get_object_or_404(Caja, id=caja_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Validar que la caja esté cerrada
                if caja.estado == 'ABIERTA':
                    messages.warning(request, f'La caja {caja.nombre} ya está abierta')
                    return redirect('custom_admin:cajas_list')
                
                monto_apertura = Decimal(request.POST.get('monto_apertura', '0'))
                observaciones = request.POST.get('observaciones', '')
                
                # Abrir la caja
                caja.estado = 'ABIERTA'
                caja.monto_apertura = monto_apertura
                caja.monto_actual = monto_apertura
                caja.fecha_apertura = timezone.now()
                caja.usuario_apertura = request.user
                caja.fecha_cierre = None
                caja.usuario_cierre = None
                caja.save()
                
                # Registrar movimiento de apertura
                MovimientoCaja.objects.create(
                    caja=caja,
                    tipo_movimiento='APERTURA',
                    monto=monto_apertura,
                    saldo_anterior=Decimal('0'),
                    saldo_nuevo=monto_apertura,
                    usuario=get_authenticated_user(request),
                    observaciones=observaciones or f'Apertura de caja - {caja.nombre}'
                )
                
                messages.success(request, f'Caja {caja.nombre} abierta exitosamente')
                
        except Exception as e:
            messages.error(request, f'Error al abrir caja: {str(e)}')
    
    return redirect('custom_admin:cajas_list')


@auth_required
def cerrar_caja(request, caja_id):
    """Cerrar una caja con arqueo"""
    
    caja = get_object_or_404(Caja, id=caja_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Validar que la caja esté abierta
                if caja.estado == 'CERRADA':
                    messages.warning(request, f'La caja {caja.nombre} ya está cerrada')
                    return redirect('custom_admin:cajas_list')
                
                # Obtener datos del formulario
                billetes_100 = int(request.POST.get('billetes_100', 0))
                billetes_50 = int(request.POST.get('billetes_50', 0))
                billetes_20 = int(request.POST.get('billetes_20', 0))
                billetes_10 = int(request.POST.get('billetes_10', 0))
                billetes_5 = int(request.POST.get('billetes_5', 0))
                billetes_1 = int(request.POST.get('billetes_1', 0))
                monedas = Decimal(request.POST.get('monedas', '0'))
                observaciones = request.POST.get('observaciones', '')
                
                # Calcular monto contado
                monto_contado = (
                    (billetes_100 * 100) +
                    (billetes_50 * 50) +
                    (billetes_20 * 20) +
                    (billetes_10 * 10) +
                    (billetes_5 * 5) +
                    billetes_1 +
                    monedas
                )
                
                # Calcular totales del período
                # Validar que la caja tenga fecha de apertura
                if not caja.fecha_apertura:
                    messages.error(request, 'No se puede cerrar una caja que no ha sido abierta')
                    return redirect('custom_admin:cajas_list')
                
                movimientos = MovimientoCaja.objects.filter(
                    caja=caja,
                    fecha_movimiento__gte=caja.fecha_apertura
                )
                
                total_ventas = movimientos.filter(tipo_movimiento='VENTA').aggregate(
                    total=Sum('monto')
                )['total'] or Decimal('0')
                
                total_ingresos = movimientos.filter(
                    tipo_movimiento__in=['INGRESO', 'AJUSTE_POSITIVO', 'TRANSFERENCIA_ENTRADA']
                ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
                
                total_retiros = movimientos.filter(
                    tipo_movimiento__in=['RETIRO', 'AJUSTE_NEGATIVO', 'TRANSFERENCIA_SALIDA']
                ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
                
                monto_esperado = caja.monto_apertura + total_ventas + total_ingresos - total_retiros
                
                # Crear arqueo
                arqueo = ArqueoCaja.objects.create(
                    caja=caja,
                    fecha_apertura=caja.fecha_apertura,
                    fecha_cierre=timezone.now(),
                    monto_apertura=caja.monto_apertura,
                    total_ventas=total_ventas,
                    total_ingresos=total_ingresos,
                    total_retiros=total_retiros,
                    monto_esperado=monto_esperado,
                    billetes_100=billetes_100,
                    billetes_50=billetes_50,
                    billetes_20=billetes_20,
                    billetes_10=billetes_10,
                    billetes_5=billetes_5,
                    billetes_1=billetes_1,
                    monedas=monedas,
                    monto_contado=monto_contado,
                    observaciones=observaciones,
                    usuario_apertura=caja.usuario_apertura,
                    usuario_cierre=get_authenticated_user(request)
                )
                
                # El método save() del modelo ArqueoCaja calculará automáticamente:
                # - diferencia
                # - estado (CUADRADO, SOBRANTE, FALTANTE)
                # - numero_arqueo
                
                # Registrar movimiento de cierre
                MovimientoCaja.objects.create(
                    caja=caja,
                    tipo_movimiento='CIERRE',
                    monto=monto_contado,
                    saldo_anterior=caja.monto_actual,
                    saldo_nuevo=Decimal('0'),
                    usuario=get_authenticated_user(request),
                    observaciones=f'Cierre de caja - Arqueo: {arqueo.numero_arqueo}'
                )
                
                # Cerrar la caja
                caja.estado = 'CERRADA'
                caja.fecha_cierre = timezone.now()
                caja.usuario_cierre = request.user
                caja.monto_actual = Decimal('0')
                caja.save()
                
                # Mensaje según el estado del arqueo
                if arqueo.estado == 'CUADRADO':
                    messages.success(request, f'Caja cerrada exitosamente. Arqueo cuadrado ✅')
                elif arqueo.estado == 'SOBRANTE':
                    messages.warning(
                        request, 
                        f'Caja cerrada. Sobrante de ${arqueo.diferencia:.2f} 💰'
                    )
                else:
                    messages.error(
                        request, 
                        f'Caja cerrada. Faltante de ${abs(arqueo.diferencia):.2f} ⚠️'
                    )
                
        except Exception as e:
            messages.error(request, f'Error al cerrar caja: {str(e)}')
    
    return redirect('custom_admin:cajas_list')


@auth_required
def movimientos_caja(request, caja_id):
    """Ver movimientos de una caja"""
    
    caja = get_object_or_404(Caja, id=caja_id)
    
    # Obtener movimientos
    movimientos = MovimientoCaja.objects.filter(caja=caja).select_related('usuario')
    
    # Filtro por fecha
    fecha_desde = request.GET.get('desde')
    fecha_hasta = request.GET.get('hasta')
    
    if fecha_desde:
        movimientos = movimientos.filter(fecha_movimiento__date__gte=fecha_desde)
    
    if fecha_hasta:
        movimientos = movimientos.filter(fecha_movimiento__date__lte=fecha_hasta)
    
    # Ordenar por fecha descendente
    movimientos = movimientos.order_by('-fecha_movimiento')[:50]  # Últimos 50
    
    # Si es AJAX, devolver solo el contenido de la tabla
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'custom_admin/finanzas/movimientos_partial.html', {
            'movimientos': movimientos,
            'caja': caja
        })
    
    return render(request, 'custom_admin/finanzas/movimientos.html', {
        'movimientos': movimientos,
        'caja': caja
    })


@auth_required
def crear_movimiento(request, caja_id):
    """Crear un nuevo movimiento en la caja"""
    
    caja = get_object_or_404(Caja, id=caja_id)
    
    if request.method == 'POST':
        try:
            # Validar que la caja esté abierta
            if caja.estado != 'ABIERTA':
                messages.error(request, 'No se pueden registrar movimientos en una caja cerrada')
                return redirect('custom_admin:cajas_list')
            
            tipo_movimiento = request.POST.get('tipo_movimiento')
            monto = Decimal(request.POST.get('monto'))
            observaciones = request.POST.get('observaciones')
            
            # Crear el movimiento (el método save() calculará los saldos automáticamente)
            MovimientoCaja.objects.create(
                caja=caja,
                tipo_movimiento=tipo_movimiento,
                monto=monto,
                usuario=get_authenticated_user(request),
                observaciones=observaciones
            )
            
            messages.success(request, 'Movimiento registrado exitosamente')
            
        except Exception as e:
            messages.error(request, f'Error al registrar movimiento: {str(e)}')
    
    return redirect('custom_admin:cajas_list')
# ============================================================================
# VISTAS DE FINANZAS - ARQUEOS
# ============================================================================

@auth_required
def arqueos_list(request):
    """Lista de arqueos de caja"""
    
    # Obtener todos los arqueos
    arqueos = ArqueoCaja.objects.all().select_related(
        'caja', 'usuario_apertura', 'usuario_cierre'
    )
    
    # Filtros
    estado = request.GET.get('estado', '')
    caja_id = request.GET.get('caja', '')
    fecha_desde = request.GET.get('desde', '')
    fecha_hasta = request.GET.get('hasta', '')
    
    if estado:
        arqueos = arqueos.filter(estado=estado)
    
    if caja_id:
        arqueos = arqueos.filter(caja_id=caja_id)
    
    if fecha_desde:
        arqueos = arqueos.filter(fecha_cierre__date__gte=fecha_desde)
    
    if fecha_hasta:
        arqueos = arqueos.filter(fecha_cierre__date__lte=fecha_hasta)
    
    # Ordenar por fecha de cierre descendente
    arqueos = arqueos.order_by('-fecha_cierre')
    
    # Estadísticas
    total_arqueos = arqueos.count()
    cuadrados = arqueos.filter(estado='CUADRADO').count()
    sobrantes = arqueos.filter(estado='SOBRANTE').count()
    faltantes = arqueos.filter(estado='FALTANTE').count()
    
    # Diferencia total
    diferencia_total = arqueos.aggregate(
        total=Sum('diferencia')
    )['total'] or Decimal('0')
    
    # Paginación
    paginator = Paginator(arqueos, 20)
    page_number = request.GET.get('page')
    arqueos_page = paginator.get_page(page_number)
    
    # Cajas para el filtro
    cajas = Caja.objects.all()
    
    context = {
        'arqueos': arqueos_page,
        'cajas': cajas,
        'total_arqueos': total_arqueos,
        'cuadrados': cuadrados,
        'sobrantes': sobrantes,
        'faltantes': faltantes,
        'diferencia_total': diferencia_total,
    }
    
    return render(request, 'custom_admin/finanzas/arqueos_list.html', context)


@auth_required
def arqueo_detalle(request, arqueo_id):
    """Detalle de un arqueo"""
    
    arqueo = get_object_or_404(ArqueoCaja, id=arqueo_id)
    
    # Obtener movimientos del período
    movimientos = MovimientoCaja.objects.filter(
        caja=arqueo.caja,
        fecha_movimiento__gte=arqueo.fecha_apertura,
        fecha_movimiento__lte=arqueo.fecha_cierre
    ).select_related('usuario').order_by('fecha_movimiento')
    
    context = {
        'arqueo': arqueo,
        'movimientos': movimientos,
    }
    
    return render(request, 'custom_admin/finanzas/arqueo_detalle.html', context)


# ============================================================================
# VISTAS DE FINANZAS - CAJA CHICA
# ============================================================================

@auth_required
def caja_chica_list(request):
    """Lista de cajas chicas"""
    from apps.authentication.models import Usuario
    
    # Obtener todas las cajas chicas
    cajas_chicas = CajaChica.objects.all().select_related('responsable')
    
    # Filtros
    estado = request.GET.get('estado', '')
    search = request.GET.get('search', '')
    
    if estado:
        cajas_chicas = cajas_chicas.filter(estado=estado)
    
    if search:
        cajas_chicas = cajas_chicas.filter(
            Q(nombre__icontains=search) |
            Q(codigo__icontains=search)
        )
    
    # Estadísticas
    total_fondos = cajas_chicas.filter(estado='ACTIVA').aggregate(
        total=Sum('monto_fondo')
    )['total'] or Decimal('0')
    
    total_disponible = cajas_chicas.filter(estado='ACTIVA').aggregate(
        total=Sum('monto_actual')
    )['total'] or Decimal('0')
    
    requieren_reposicion = sum(1 for cc in cajas_chicas if cc.necesita_reposicion())
    
    # Paginación
    paginator = Paginator(cajas_chicas, 12)
    page_number = request.GET.get('page')
    cajas_page = paginator.get_page(page_number)
    
    # Obtener usuarios activos para el formulario de creación
    # ✅ CAMBIO: 'first_name', 'last_name' → 'nombres', 'apellidos'
    usuarios = Usuario.objects.filter(is_active=True).order_by('nombres', 'apellidos')
    
    context = {
        'cajas_chicas': cajas_page,
        'total_fondos': total_fondos,
        'total_disponible': total_disponible,
        'requieren_reposicion': requieren_reposicion,
        'usuarios': usuarios,
    }
    
    return render(request, 'custom_admin/finanzas/caja_chica_list.html', context)
@auth_required
def crear_caja_chica(request):
    """Crear nueva caja chica"""
    
    if request.method == 'POST':
        try:
            from apps.authentication.models import Usuario
            
            nombre = request.POST.get('nombre')
            codigo = request.POST.get('codigo')
            responsable_id = request.POST.get('responsable')
            monto_fondo = Decimal(request.POST.get('monto_fondo'))
            umbral_reposicion = Decimal(request.POST.get('umbral_reposicion'))
            limite_gasto = Decimal(request.POST.get('limite_gasto_individual', '0'))
            
            # Validar código único
            if CajaChica.objects.filter(codigo=codigo).exists():
                messages.error(request, f'Ya existe una caja chica con el código {codigo}')
                return redirect('custom_admin:caja_chica_list')
            
            responsable = Usuario.objects.get(id=responsable_id)
            
            # Crear caja chica
            caja_chica = CajaChica.objects.create(
                nombre=nombre,
                codigo=codigo,
                responsable=responsable,
                monto_fondo=monto_fondo,
                monto_actual=monto_fondo,
                umbral_reposicion=umbral_reposicion,
                limite_gasto_individual=limite_gasto,
                estado='ACTIVA'
            )
            
            # Registrar movimiento inicial
            MovimientoCajaChica.objects.create(
                caja_chica=caja_chica,
                tipo_movimiento='APERTURA',
                monto=monto_fondo,
                saldo_anterior=Decimal('0'),
                saldo_nuevo=monto_fondo,
                descripcion=f'Apertura de caja chica - {nombre}',
                usuario=get_authenticated_user(request)
            )
            
            messages.success(request, f'Caja chica {nombre} creada exitosamente')
            
        except Exception as e:
            messages.error(request, f'Error al crear caja chica: {str(e)}')
    
    return redirect('custom_admin:caja_chica_list')


@auth_required
def registrar_gasto_caja_chica(request, caja_chica_id):
    """Registrar un gasto en caja chica"""
    
    caja_chica = get_object_or_404(CajaChica, id=caja_chica_id)
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            monto = Decimal(request.POST.get('monto'))
            categoria = request.POST.get('categoria_gasto', '')
            descripcion = request.POST.get('descripcion')
            numero_comprobante = request.POST.get('numero_comprobante', '')
            comprobante = request.FILES.get('comprobante_adjunto')
            
            # Validar estado de la caja
            if caja_chica.estado != 'ACTIVA':
                messages.error(request, '❌ La caja chica no está activa')
                return redirect('custom_admin:caja_chica_list')
            
            # Validar monto
            if monto <= 0:
                messages.error(request, '❌ El monto debe ser mayor a cero')
                return redirect('custom_admin:caja_chica_list')
            
            # Validar saldo suficiente
            if monto > caja_chica.monto_actual:
                messages.error(request, f'❌ Saldo insuficiente. Disponible: Q. {caja_chica.monto_actual}')
                return redirect('custom_admin:caja_chica_list')
            
            # Validar límite de gasto individual si existe
            if caja_chica.limite_gasto_individual > 0 and monto > caja_chica.limite_gasto_individual:
                messages.error(
                    request, 
                    f'❌ El monto excede el límite permitido de Q. {caja_chica.limite_gasto_individual}'
                )
                return redirect('custom_admin:caja_chica_list')
            
            # Calcular saldos
            saldo_anterior = caja_chica.monto_actual
            saldo_nuevo = saldo_anterior - monto
            
            # Crear movimiento
            movimiento = MovimientoCajaChica.objects.create(
                caja_chica=caja_chica,
                tipo_movimiento='GASTO',
                categoria_gasto=categoria,
                monto=monto,
                saldo_anterior=saldo_anterior,
                saldo_nuevo=saldo_nuevo,
                descripcion=descripcion,
                numero_comprobante=numero_comprobante,
                comprobante_adjunto=comprobante,
                usuario=get_authenticated_user(request)
            )
            
            # Actualizar saldo de la caja
            caja_chica.monto_actual = saldo_nuevo
            caja_chica.save()
            
            messages.success(
                request, 
                f'✅ Gasto de Q. {monto} registrado exitosamente. Saldo actual: Q. {saldo_nuevo}'
            )
            
            # ✅ CORREGIDO: Verificar si necesita reposición con el SALDO NUEVO
            if caja_chica.umbral_reposicion > 0 and saldo_nuevo < caja_chica.umbral_reposicion:
                messages.warning(
                    request,
                    f'⚠️ La caja chica necesita reposición. Umbral: Q. {caja_chica.umbral_reposicion}'
                )
            
        except Exception as e:
            messages.error(request, f'❌ Error al registrar movimiento: {str(e)}')
            logger.error(f"Error en registrar_gasto_caja_chica: {str(e)}")
    
    return redirect('custom_admin:caja_chica_list')


@auth_required
def reponer_caja_chica(request, caja_chica_id):
    """Reponer fondo de caja chica"""
    
    caja_chica = get_object_or_404(CajaChica, id=caja_chica_id)
    
    if request.method == 'POST':
        try:
            # ✅ Obtener el monto del formulario (no calcularlo automáticamente)
            monto = Decimal(request.POST.get('monto', 0))
            observaciones = request.POST.get('observaciones', '').strip()
            
            # Validar estado de la caja
            if caja_chica.estado != 'ACTIVA':
                messages.error(request, '❌ La caja chica no está activa')
                return redirect('custom_admin:caja_chica_list')
            
            # Validar monto
            if monto <= 0:
                messages.error(request, '❌ El monto debe ser mayor a cero')
                return redirect('custom_admin:caja_chica_list')
            
            # Calcular saldos
            saldo_anterior = caja_chica.monto_actual
            saldo_nuevo = saldo_anterior + monto
            
            # Preparar descripción
            descripcion = f'Reposición de fondo por Q. {monto}'
            if observaciones:
                descripcion += f' - {observaciones}'
            
            # ✅ Registrar movimiento de reposición CON saldos
            movimiento = MovimientoCajaChica.objects.create(
                caja_chica=caja_chica,
                tipo_movimiento='REPOSICION',  # ✅ ESTABLECER DIRECTAMENTE
                monto=monto,
                saldo_anterior=saldo_anterior,
                saldo_nuevo=saldo_nuevo,
                descripcion=descripcion,
                usuario=get_authenticated_user(request)
            )
            
            # Actualizar saldo y fecha de última reposición
            caja_chica.monto_actual = saldo_nuevo
            caja_chica.fecha_ultima_reposicion = timezone.now()
            caja_chica.save()
            
            messages.success(
                request, 
                f'✅ Reposición de Q. {monto} registrada exitosamente. Saldo actual: Q. {saldo_nuevo}'
            )
            
        except Exception as e:
            messages.error(request, f'❌ Error al reponer caja chica: {str(e)}')
            logger.error(f"Error en reponer_caja_chica: {str(e)}")
    
    return redirect('custom_admin:caja_chica_list')

logger = logging.getLogger(__name__)
@ensure_csrf_cookie
@auth_required
@require_http_methods(["POST"])
def producto_imprimir_codigo(request, producto_id):
    """
    Imprime el código de barras de un producto
    Ahora con soporte para múltiples copias
    
    POST /panel/inventario/productos/{producto_id}/imprimir-codigo/
    
    Body (opcional):
    {
        "copias": 3  // Número de copias a imprimir (default: 1)
    }
    """
    try:
        from apps.inventory_management.models import Producto
        from apps.hardware_integration.models import Impresora, TrabajoImpresion
        from apps.hardware_integration.printers.printer_service import PrinterService
        from django.utils import timezone
        import uuid
        
        # Obtener número de copias del request
        copias = 1
        if request.body:
            try:
                body = json.loads(request.body)
                copias = int(body.get('copias', 1))
                if copias < 1:
                    copias = 1
                elif copias > 100:  # Límite de seguridad
                    copias = 100
            except (json.JSONDecodeError, ValueError, TypeError):
                copias = 1
        
        # Obtener el producto
        try:
            producto = Producto.objects.get(id=producto_id)
        except Producto.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Producto no encontrado'
            }, status=404)
        
        # Validar que el producto tenga código de barras
        if not producto.codigo_barras:
            return JsonResponse({
                'success': False,
                'error': 'El producto no tiene código de barras asignado'
            }, status=400)
        
        logger.info(f"🏷️ Solicitando impresión de código de barras")
        logger.info(f"   Producto: {producto.nombre}")
        logger.info(f"   Código: {producto.codigo_barras}")
        logger.info(f"   Copias: {copias}")
        
        # Obtener impresora de etiquetas principal
        impresora = Impresora.objects.filter(
            tipo_impresora='ETIQUETAS',
            es_principal_etiquetas=True,
            estado='ACTIVA'
        ).first()
        
        if not impresora:
            # Buscar cualquier impresora de etiquetas activa
            impresora = Impresora.objects.filter(
                tipo_impresora='ETIQUETAS',
                estado='ACTIVA'
            ).first()
        
        if not impresora:
            return JsonResponse({
                'success': False,
                'error': 'No hay impresoras de etiquetas configuradas. Por favor configure una impresora primero.'
            }, status=404)
        
        logger.info(f"   Impresora seleccionada: {impresora.nombre}")
        
            # Generar comandos para las copias
        comandos_completos = b''
        
        for i in range(copias):
            # Generar etiqueta individual
            comandos = PrinterService.generar_etiqueta_producto(
            nombre=producto.nombre,
            codigo=producto.codigo_barras,
            precio=float(producto.precio_unitario or 0),
            codigo_barras=producto.codigo_barras
        )

        if not comandos:
            return JsonResponse({
                'success': False,
                'error': 'No se pudieron generar los comandos de impresión'
            }, status=500)

        # Convertir a hexadecimal
        comandos_hex = comandos.hex()

        # Crear trabajo de impresión
        trabajo = TrabajoImpresion.objects.create(
            id=uuid.uuid4(),
            tipo='CODIGO_BARRAS',
            impresora=impresora,
            producto=producto,
            datos_impresion=comandos_hex,
            formato='TSPL',
            prioridad=2,
            estado='PENDIENTE',
            copias=copias,  # ✅ El agente imprimirá N copias automáticamente
            creado_por=request.user,
            metadata={
                'descripcion': f'Código de barras: {producto.nombre} ({copias} copia(s))',
                'codigo_barras': producto.codigo_barras,
                'nombre_producto': producto.nombre
            }
        )
        
        logger.info(f" Trabajo de impresión creado: {trabajo.id}")
        logger.info(f"   Estado: {trabajo.estado}")
        logger.info(f"   Impresora: {trabajo.impresora_nombre}")
        logger.info(f"   Copias: {copias}")
        
        return JsonResponse({
            'success': True,
            'mensaje': f'{"Código de barras" if copias == 1 else f"{copias} etiquetas"} enviado(s) a imprimir en {impresora.nombre}',
            'trabajo_id': str(trabajo.id),
            'producto': {
                'id': str(producto.id),
                'nombre': producto.nombre,
                'codigo_barras': producto.codigo_barras
            },
            'impresora': impresora.nombre,
            'copias': copias
        }, status=201)
        
    except Exception as e:
        logger.error(f"❌ Error imprimiendo código de barras: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': f'Error al procesar la impresión: {str(e)}'
        }, status=500)
@auth_required
def caja_chica_movimientos_api(request, caja_chica_id):
    """API para obtener los movimientos de una caja chica"""
    try:
        from django.db.models import Sum
        
        caja = get_object_or_404(CajaChica, id=caja_chica_id)
        
        # Obtener movimientos (últimos 50)
        movimientos = MovimientoCajaChica.objects.filter(
            caja_chica=caja
        ).select_related('usuario').order_by('-fecha_movimiento')[:50]
        
        # Calcular total de gastos
        total_gastos = MovimientoCajaChica.objects.filter(
            caja_chica=caja,
            tipo_movimiento='GASTO'
        ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
        
        # Preparar datos de movimientos
        movimientos_data = []
        for mov in movimientos:
            movimientos_data.append({
                'id': str(mov.id),
                'fecha_movimiento': mov.fecha_movimiento.isoformat(),
                'tipo_movimiento': mov.tipo_movimiento,
                'categoria_gasto': mov.categoria_gasto or '',
                'descripcion': mov.descripcion,
                'monto': float(mov.monto),
                'numero_comprobante': mov.numero_comprobante or '',
                'comprobante_adjunto': mov.comprobante_adjunto.url if mov.comprobante_adjunto else None,
                'saldo_anterior': float(mov.saldo_anterior) if mov.saldo_anterior else 0,
                'saldo_nuevo': float(mov.saldo_nuevo) if mov.saldo_nuevo else 0,
                'usuario': mov.usuario.get_full_name() if mov.usuario else 'Sistema',
            })
        
        # Preparar respuesta
        data = {
            'id': str(caja.id),
            'nombre': caja.nombre,
            'codigo': caja.codigo or '',
            'descripcion': caja.descripcion if hasattr(caja, 'descripcion') else '',
            'monto_fondo': float(caja.monto_fondo),
            'monto_inicial': float(caja.monto_fondo),  # Para compatibilidad
            'monto_actual': float(caja.monto_actual),
            'total_gastos': float(total_gastos),
            'estado': caja.estado,
            'responsable': caja.responsable.get_full_name() if caja.responsable else 'Sin asignar',
            'movimientos': movimientos_data
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        logger.error(f"Error en caja_chica_movimientos_api: {str(e)}")
        return JsonResponse({'error': str(e)}, status=400)
@auth_required
@require_http_methods(["POST"])
def editar_caja_chica(request, caja_chica_id):
    """Editar una caja chica existente"""
    try:
        from apps.authentication.models import Usuario
        
        caja_chica = get_object_or_404(CajaChica, id=caja_chica_id)
        
        # Obtener datos del formulario
        nombre = request.POST.get('nombre', '').strip()
        codigo = request.POST.get('codigo', '').strip()
        monto_fondo = Decimal(request.POST.get('monto_fondo', 0))
        responsable_id = request.POST.get('responsable')
        umbral_reposicion = Decimal(request.POST.get('umbral_reposicion', 0))
        limite_gasto = Decimal(request.POST.get('limite_gasto_individual', 0))
        estado = request.POST.get('estado', 'ACTIVA')
        
        # DEBUG: Imprimir valores recibidos
        print(f"DEBUG - Umbral recibido: {umbral_reposicion}")
        print(f"DEBUG - POST data: {request.POST}")
        
        # Validaciones
        if not nombre:
            messages.error(request, '❌ El nombre es obligatorio.')
            return redirect('custom_admin:caja_chica_list')
        
        if not codigo:
            messages.error(request, '❌ El código es obligatorio.')
            return redirect('custom_admin:caja_chica_list')
        
        # Validar código único (excepto para la misma caja)
        if CajaChica.objects.filter(codigo=codigo).exclude(id=caja_chica_id).exists():
            messages.error(request, f'❌ Ya existe otra caja chica con el código {codigo}')
            return redirect('custom_admin:caja_chica_list')
        
        if monto_fondo <= 0:
            messages.error(request, '❌ El monto del fondo debe ser mayor a cero.')
            return redirect('custom_admin:caja_chica_list')
        
        # Obtener responsable
        try:
            responsable = Usuario.objects.get(id=responsable_id)
        except Usuario.DoesNotExist:
            messages.error(request, '❌ El responsable seleccionado no existe.')
            return redirect('custom_admin:caja_chica_list')
        
        # Actualizar la caja chica
        caja_chica.nombre = nombre
        caja_chica.codigo = codigo
        caja_chica.monto_fondo = monto_fondo
        caja_chica.responsable = responsable
        caja_chica.umbral_reposicion = umbral_reposicion
        caja_chica.limite_gasto_individual = limite_gasto
        caja_chica.estado = estado
        caja_chica.save()
        
        # DEBUG: Verificar que se guardó
        caja_chica.refresh_from_db()
        print(f"DEBUG - Umbral guardado: {caja_chica.umbral_reposicion}")
        
        messages.success(
            request,
            f'✅ Caja chica "{nombre}" actualizada exitosamente. Umbral: Q. {caja_chica.umbral_reposicion}'
        )
        
    except Exception as e:
        messages.error(request, f'❌ Error al editar caja chica: {str(e)}')
        logger.error(f"Error en editar_caja_chica: {str(e)}")
        import traceback
        traceback.print_exc()
    
    return redirect('custom_admin:caja_chica_list')
@require_http_methods(["GET"])
def buscar_venta_api(request):
    numero = request.GET.get('numero', '').strip()
    
    print(f"🔍 Buscando venta con número: '{numero}'")
    
    if not numero:
        return JsonResponse({'success': False, 'error': 'Número de venta requerido'})
    
    try:
        venta = Venta.objects.get(numero_venta__iexact=numero)
        print(f"✅ Venta encontrada: {venta.numero_venta}")
        
        detalles = venta.detalles.all()
        print(f"📦 Productos encontrados: {detalles.count()}")
        
        productos = []
        for detalle in detalles:
            # Debug cada campo
            print(f"Producto nombre: {type(detalle.producto.nombre)} - {detalle.producto.nombre}")
            
            producto_data = {
                'id': str(detalle.id),
                'nombre': str(detalle.producto.nombre),  # Forzar a string
                'cantidad': float(detalle.cantidad_unidades),
                'precio': float(detalle.precio_unitario),
                'subtotal': float(detalle.subtotal)
            }
            print(f"Producto data: {producto_data}")
            productos.append(producto_data)
        
        # Debug venta
        print(f"Venta ID: {type(venta.id)} - {venta.id}")
        print(f"Venta numero: {type(venta.numero_venta)} - {venta.numero_venta}")
        print(f"Cliente: {type(venta.cliente)} - {venta.cliente}")
        print(f"Fecha: {type(venta.fecha_venta)} - {venta.fecha_venta}")
        print(f"Total: {type(venta.total)} - {venta.total}")
        
        venta_data = {
            'success': True,
            'venta': {
                'id': str(venta.id),
                'numero': str(venta.numero_venta),
                'cliente': str(venta.cliente.nombre_completo) if venta.cliente else 'Cliente General',
                'fecha': str(venta.fecha_venta.strftime('%d/%m/%Y %H:%M')),
                'total': float(venta.total),
                'productos': productos
            }
        }
        
        print(f"✅ Venta data construida correctamente")
        return JsonResponse(venta_data)
        
    except Venta.DoesNotExist:
        print(f"❌ Venta no encontrada con número: '{numero}'")
        return JsonResponse({'success': False, 'error': 'Venta no encontrada'})
    except Exception as e:
        import traceback
        print(f"❌ Error completo: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'error': f'Error: {str(e)}'})

@require_http_methods(["POST"])
def procesar_devolucion_api(request):
    import json
    from decimal import Decimal
    
    data = json.loads(request.body)
    
    print(f"📥 Datos recibidos: {data}")
    
    try:
        venta = Venta.objects.get(id=data['venta_id'])
        print(f"✅ Venta encontrada: {venta.numero_venta}")
        
        detalle = DetalleVenta.objects.get(id=data['detalle_venta_id'])
        print(f"✅ Detalle encontrado: {detalle.producto.nombre}")
        
        # Calcular monto de devolución
        cantidad_devuelta = Decimal(str(data['cantidad_devuelta']))
        monto_devolucion = cantidad_devuelta * detalle.precio_unitario
        
        # Generar número de devolución
        año = timezone.now().year
        ultimo = Devolucion.objects.filter(
            numero_devolucion__startswith=f'DEV-{año}-'
        ).order_by('-numero_devolucion').first()
        
        if ultimo and ultimo.numero_devolucion:
            try:
                ultimo_num = int(ultimo.numero_devolucion.split('-')[-1])
                siguiente_num = ultimo_num + 1
            except (ValueError, IndexError):
                siguiente_num = 1
        else:
            siguiente_num = 1
        
        numero_devolucion = f'DEV-{año}-{siguiente_num:05d}'
        
        # Crear la devolución
        devolucion = Devolucion.objects.create(
            numero_devolucion=numero_devolucion,
            venta_original=venta,  # ✅ CORREGIDO
            detalle_venta=detalle,
            cantidad_devuelta=cantidad_devuelta,
            monto_devolucion=monto_devolucion,  # ✅ AGREGADO
            motivo=data['motivo'],
            descripcion=data.get('descripcion', ''),
            usuario_solicita=request.user,
            estado='PENDIENTE'
        )
        
        print(f"✅ Devolución creada: {devolucion.numero_devolucion}")
        
        return JsonResponse({
            'success': True,
            'numero_devolucion': devolucion.numero_devolucion,
            'message': 'Devolución registrada exitosamente'
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error completo: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'error': str(e)})
@require_http_methods(["POST"])
def aprobar_devolucion_api(request, id):
    import json
    data = json.loads(request.body)
    
    try:
        devolucion = Devolucion.objects.get(id=id)
        decision = data.get('decision')
        
        if decision == 'APROBADA':
            devolucion.estado = 'APROBADA'
            devolucion.usuario_aprueba = request.user
            # Aquí agregarías lógica para devolver al inventario
        elif decision == 'RECHAZADA':
            devolucion.estado = 'RECHAZADA'
            devolucion.observaciones_rechazo = data.get('observaciones', '')
        
        devolucion.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Devolución {decision.lower()} exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
# ============================================================================
@require_http_methods(["GET"])
def devolucion_detalle_api(request, devolucion_id):
    """API para obtener detalles de una devolución"""
    try:
        devolucion = Devolucion.objects.get(id=devolucion_id)
        
        data = {
            'success': True,
            'devolucion': {
                'id': str(devolucion.id),
                'numero_devolucion': devolucion.numero_devolucion,
                'venta_original': devolucion.venta_original.numero_venta,
                'cliente': devolucion.venta_original.cliente.nombre_completo if devolucion.venta_original.cliente else 'Cliente General',
                'producto': devolucion.detalle_venta.producto.nombre,
                'cantidad_devuelta': float(devolucion.cantidad_devuelta),
                'monto': float(devolucion.monto_devolucion),
                'motivo': devolucion.motivo,
                'motivo_display': devolucion.get_motivo_display(),
                'descripcion': devolucion.descripcion,
                'estado': devolucion.estado,
                'estado_display': devolucion.get_estado_display(),
                'fecha_devolucion': devolucion.fecha_devolucion.strftime('%d/%m/%Y %H:%M'),
                'usuario_solicita': devolucion.usuario_solicita.get_full_name() if hasattr(devolucion.usuario_solicita, 'get_full_name') else str(devolucion.usuario_solicita),
                'usuario_aprueba': devolucion.usuario_aprueba.get_full_name() if devolucion.usuario_aprueba and hasattr(devolucion.usuario_aprueba, 'get_full_name') else (str(devolucion.usuario_aprueba) if devolucion.usuario_aprueba else None),
                'fecha_procesado': devolucion.fecha_procesado.strftime('%d/%m/%Y %H:%M') if devolucion.fecha_procesado else None
            }
        }
        
        return JsonResponse(data)
        
    except Devolucion.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Devolución no encontrada'})
    except Exception as e:
        import traceback
        print(f"❌ Error en devolucion_detalle_api: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'error': str(e)})


@ensure_csrf_cookie
@auth_required
@require_http_methods(["POST"])
def aprobar_devolucion_api(request, id):
    """API para aprobar o rechazar una devolución"""
    import json
    from django.http import HttpResponse
    from django.db import transaction
    from apps.sales_management.models import Devolucion
    from apps.inventory_management.models import Quintal, MovimientoInventario, MovimientoQuintal, ProductoNormal
    from django.utils import timezone
    from decimal import Decimal
    
    try:
        data = json.loads(request.body)
        print(f"📥 Data recibida: {data}")
        
        devolucion = Devolucion.objects.get(id=id)
        print(f"✅ Devolución encontrada: {devolucion.numero_devolucion}")
        
        if devolucion.estado != 'PENDIENTE':
            return HttpResponse(
                json.dumps({
                    'success': False,
                    'error': 'Solo se pueden procesar devoluciones pendientes'
                }),
                content_type='application/json'
            )
        
        decision = data.get('decision')
        print(f"📋 Decisión: {decision}")
        
        if decision == 'APROBADA':
            with transaction.atomic():
                devolucion.estado = 'APROBADA'
                devolucion.usuario_aprueba = request.user
                devolucion.fecha_procesado = timezone.now()
                devolucion.save()
                print(f"✅ Estado actualizado a APROBADA")
                
                # ✅ REINTEGRAR AL INVENTARIO
                producto = devolucion.detalle_venta.producto
                cantidad_devuelta = devolucion.cantidad_devuelta
                
                print(f"🔍 Producto: {producto.nombre}")
                print(f"🔍 Tipo inventario: {producto.tipo_inventario}")
                print(f"🔍 Cantidad devuelta: {cantidad_devuelta}")
                print(f"🔍 Es quintal?: {producto.es_quintal()}")
                
                if producto.es_quintal():
                    print("📦 Procesando producto QUINTAL...")
                    try:
                        # Crear nuevo quintal con el producto devuelto
                        nuevo_quintal = Quintal.objects.create(
                            producto=producto,
                            peso_inicial=cantidad_devuelta,
                            peso_actual=cantidad_devuelta,
                            unidad_medida=producto.unidad_medida_base,
                            precio_por_unidad=producto.precio_por_unidad_peso,
                            costo_por_unidad=producto.precio_por_unidad_peso * Decimal('0.7'),  # Estimado
                            costo_total=cantidad_devuelta * producto.precio_por_unidad_peso * Decimal('0.7'),
                            fecha_ingreso=timezone.now(),
                            estado='DISPONIBLE',
                            proveedor=producto.proveedor,
                            usuario_registro=request.user,
                            observaciones=f'Devolución {devolucion.numero_devolucion} - Venta {devolucion.venta_original.numero_venta}'
                        )
                        
                        print(f"✅ Quintal creado: ID={nuevo_quintal.codigo_unico}, Peso={nuevo_quintal.peso_actual}")
                        
                        # El signal post_save de Quintal creará automáticamente el MovimientoQuintal
                        
                    except Exception as e:
                        import traceback
                        print(f"❌ Error al crear quintal: {traceback.format_exc()}")
                        raise  # Re-lanzar para que el transaction.atomic haga rollback
                        
                else:
                    print("📦 Procesando producto NORMAL...")
                    try:
                        # Obtener o crear el inventario normal del producto
                        inventario, created = ProductoNormal.objects.get_or_create(
                            producto=producto,
                            defaults={
                                'stock_actual': 0,
                                'stock_minimo': 10,
                                'costo_unitario': producto.precio_unitario or Decimal('0')
                            }
                        )
                        
                        if created:
                            print(f"📝 Inventario creado para {producto.nombre}")
                        
                        # Guardar stock anterior
                        stock_antes = inventario.stock_actual
                        
                        # Reintegrar las unidades devueltas
                        inventario.stock_actual += int(cantidad_devuelta)
                        inventario.save()
                        
                        print(f"✅ Stock de {producto.nombre} actualizado: {stock_antes} → {inventario.stock_actual}")
                        
                        # Registrar movimiento de inventario
                        MovimientoInventario.objects.create(
                            producto_normal=inventario,
                            tipo_movimiento='ENTRADA_AJUSTE',
                            cantidad=int(cantidad_devuelta),
                            stock_antes=stock_antes,
                            stock_despues=inventario.stock_actual,
                            costo_unitario=inventario.costo_unitario,
                            costo_total=int(cantidad_devuelta) * inventario.costo_unitario,
                            usuario=request.user,
                            observaciones=f'Devolución {devolucion.numero_devolucion}'
                        )
                        
                        print(f"✅ Movimiento de inventario registrado")
                        
                    except Exception as e:
                        import traceback
                        print(f"❌ Error al actualizar inventario normal: {traceback.format_exc()}")
                        raise  # Re-lanzar para que el transaction.atomic haga rollback
            
            return HttpResponse(
                json.dumps({
                    'success': True,
                    'message': 'Devolución aprobada exitosamente e inventario reintegrado'
                }),
                content_type='application/json'
            )
            
        elif decision == 'RECHAZADA':
            devolucion.estado = 'RECHAZADA'
            devolucion.usuario_aprueba = request.user
            devolucion.fecha_procesado = timezone.now()
            
            if 'observaciones' in data:
                devolucion.descripcion += f"\n\n--- MOTIVO DE RECHAZO ---\n{data['observaciones']}"
            
            devolucion.save()
            
            return HttpResponse(
                json.dumps({
                    'success': True,
                    'message': 'Devolución rechazada'
                }),
                content_type='application/json'
            )
        
        else:
            return HttpResponse(
                json.dumps({
                    'success': False,
                    'error': 'Decisión no válida'
                }),
                content_type='application/json'
            )
        
    except Devolucion.DoesNotExist:
        return HttpResponse(
            json.dumps({'success': False, 'error': 'Devolución no encontrada'}),
            content_type='application/json'
        )
    except Exception as e:
        import traceback
        print(f"❌ Error en aprobar_devolucion_api: {traceback.format_exc()}")
        return HttpResponse(
            json.dumps({'success': False, 'error': str(e)}),
            content_type='application/json'
        )
# ============================================================================
# TODAS LAS VISTAS API CORREGIDAS CON CAMPOS CORRECTOS
# ============================================================================
# Reemplaza las 5 funciones API en tu views.py con estas versiones corregidas

import json
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib.auth.decorators import login_required
from apps.inventory_management.models import (
    UnidadMedida, Categoria, Marca, Proveedor, Producto
)
from decimal import Decimal


# ============================================
# API: CONTADOR DE PRODUCTOS
# ============================================
@login_required
@require_http_methods(["GET"])
def api_productos_count(request):
    """API para obtener el contador total de productos"""
    try:
        count = Producto.objects.filter(activo=True).count()
        return JsonResponse({
            'success': True,
            'count': count
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================
# API: CREAR UNIDAD DE MEDIDA (MODAL)
# ============================================
@ensure_csrf_cookie
@login_required
@require_http_methods(["POST"])
def api_crear_unidad_medida(request):
    """API para crear una unidad de medida desde el modal - Retorna JSON"""
    try:
        data = json.loads(request.body)
        nombre = data.get('nombre', '').strip()
        abreviatura = data.get('abreviatura', '').strip()
        
        if not nombre or not abreviatura:
            return JsonResponse({
                'success': False,
                'error': 'El nombre y la abreviatura son requeridos'
            }, status=400)
        
        # Verificar si ya existe
        if UnidadMedida.objects.filter(nombre__iexact=nombre).exists():
            return JsonResponse({
                'success': False,
                'error': f'Ya existe una unidad de medida con el nombre "{nombre}"'
            }, status=400)
        
        if UnidadMedida.objects.filter(abreviatura__iexact=abreviatura).exists():
            return JsonResponse({
                'success': False,
                'error': f'Ya existe una unidad de medida con la abreviatura "{abreviatura}"'
            }, status=400)
        
        # Crear la unidad - usando 'activa' en lugar de 'activo'
        unidad = UnidadMedida.objects.create(
            nombre=nombre,
            abreviatura=abreviatura,
            activa=True  # ← CORREGIDO
        )
        
        return JsonResponse({
            'success': True,
            'id': str(unidad.id),
            'nombre': unidad.nombre,
            'abreviatura': unidad.abreviatura,
            'message': 'Unidad de medida creada exitosamente'
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error en api_crear_unidad_medida: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================
# API: CREAR CATEGORÍA (MODAL)
# ============================================
@ensure_csrf_cookie
@login_required
@require_http_methods(["POST"])
def categoria_crear_api(request):
    """API para crear una categoría desde el modal - Retorna JSON - VERSIÓN COMPLETA"""
    try:
        data = json.loads(request.body)
        
        # Obtener todos los campos
        nombre = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        margen_ganancia = data.get('margen_ganancia_sugerido', 30.00)
        descuento_maximo = data.get('descuento_maximo_permitido', 10.00)
        orden = data.get('orden', 0)
        activa = data.get('activa', True)
        
        # Validaciones
        if not nombre:
            return JsonResponse({
                'success': False,
                'error': 'El nombre es requerido'
            }, status=400)
        
        # Verificar si ya existe
        if Categoria.objects.filter(nombre__iexact=nombre).exists():
            return JsonResponse({
                'success': False,
                'error': f'Ya existe una categoría con el nombre "{nombre}"'
            }, status=400)
        
        # Crear la categoría con TODOS los campos
        categoria = Categoria.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            activa=activa,
            margen_ganancia_sugerido=Decimal(str(margen_ganancia)),
            descuento_maximo_permitido=Decimal(str(descuento_maximo)),
            orden=int(orden)
        )
        
        return JsonResponse({
            'success': True,
            'id': str(categoria.id),
            'nombre': categoria.nombre,
            'descripcion': categoria.descripcion,
            'margen_ganancia_sugerido': float(categoria.margen_ganancia_sugerido),
            'descuento_maximo_permitido': float(categoria.descuento_maximo_permitido),
            'orden': categoria.orden,
            'activa': categoria.activa,
            'message': 'Categoría creada exitosamente'
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error en categoria_crear_api: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

# ============================================
# VISTA API MARCA CON FORMDATA (MANEJA LOGO)
# ============================================
# AGREGAR esta nueva vista en views.py

@ensure_csrf_cookie
@login_required
@require_http_methods(["POST"])
def marca_crear_formdata(request):
    """API para crear una marca desde el modal con logo - Recibe FormData"""
    try:
        # Obtener datos del FormData
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        pais_origen = request.POST.get('pais_origen', '').strip()
        fabricante = request.POST.get('fabricante', '').strip()
        sitio_web = request.POST.get('sitio_web', '').strip()
        orden = request.POST.get('orden', '0')
        activa = request.POST.get('activa', '0') == '1'
        destacada = request.POST.get('destacada', '0') == '1'
        logo = request.FILES.get('logo')
        
        # Validaciones
        if not nombre:
            return JsonResponse({
                'success': False,
                'error': 'El nombre es requerido'
            }, status=400)
        
        # Verificar si ya existe
        if Marca.objects.filter(nombre__iexact=nombre).exists():
            return JsonResponse({
                'success': False,
                'error': f'Ya existe una marca con el nombre "{nombre}"'
            }, status=400)
        
        # Preparar datos de la marca
        # Usar string vacío '' en lugar de None para campos que no permiten NULL
        marca_data = {
            'nombre': nombre,
            'descripcion': descripcion,  # String vacío si está vacío
            'pais_origen': pais_origen if pais_origen else '',  # String vacío
            'fabricante': fabricante if fabricante else '',      # String vacío
            'sitio_web': sitio_web if sitio_web else '',        # String vacío (era el problema)
            'activa': activa,
            'destacada': destacada,
            'orden': int(orden)
        }
        
        # Agregar logo solo si existe
        if logo:
            marca_data['logo'] = logo
        
        # Crear la marca
        marca = Marca.objects.create(**marca_data)
        
        return JsonResponse({
            'success': True,
            'id': str(marca.id),
            'nombre': marca.nombre,
            'descripcion': marca.descripcion,
            'pais_origen': marca.pais_origen,
            'fabricante': marca.fabricante,
            'sitio_web': marca.sitio_web,
            'orden': marca.orden,
            'activa': marca.activa,
            'destacada': marca.destacada,
            'logo_url': marca.logo.url if marca.logo else None,
            'message': 'Marca creada exitosamente'
        })
        
    except IntegrityError as e:
        error_msg = str(e)
        if 'not-null constraint' in error_msg.lower():
            return JsonResponse({
                'success': False,
                'error': 'Falta un campo obligatorio. Por favor complete todos los campos requeridos.'
            }, status=400)
        return JsonResponse({
            'success': False,
            'error': f'Error de integridad: {error_msg}'
        }, status=400)
    except Exception as e:
        import traceback
        print(f"❌ Error en marca_crear_formdata: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'error': f'Error al crear la marca: {str(e)}'
        }, status=500)
    
@ensure_csrf_cookie
@login_required
@require_http_methods(["POST"])
def marca_crear_api(request):
    """API para crear una marca desde el modal - Retorna JSON"""
    try:
        data = json.loads(request.body)
        nombre = data.get('nombre', '').strip()
        descripcion = data.get('descripcion', '').strip()
        
        if not nombre:
            return JsonResponse({
                'success': False,
                'error': 'El nombre es requerido'
            }, status=400)
        
        # Verificar si ya existe
        if Marca.objects.filter(nombre__iexact=nombre).exists():
            return JsonResponse({
                'success': False,
                'error': f'Ya existe una marca con el nombre "{nombre}"'
            }, status=400)
        
        # Crear la marca - el campo ya es 'activa'
        marca = Marca.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            activa=True  # ← Este ya estaba correcto
        )
        
        return JsonResponse({
            'success': True,
            'id': str(marca.id),
            'nombre': marca.nombre,
            'descripcion': marca.descripcion,
            'message': 'Marca creada exitosamente'
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error en marca_crear_api: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================
# API: CREAR PROVEEDOR (MODAL)
# ============================================
@ensure_csrf_cookie
@login_required
@require_http_methods(["POST"])
def proveedor_crear_api(request):
    """API para crear un proveedor desde el modal - Retorna JSON"""
    try:
        data = json.loads(request.body)
        nombre_comercial = data.get('nombre_comercial', '').strip()
        contacto = data.get('contacto', '').strip()
        telefono = data.get('telefono', '').strip()
        email = data.get('email', '').strip()
        
        if not nombre_comercial:
            return JsonResponse({
                'success': False,
                'error': 'El nombre comercial es requerido'
            }, status=400)
        
        # Verificar si ya existe
        if Proveedor.objects.filter(nombre_comercial__iexact=nombre_comercial).exists():
            return JsonResponse({
                'success': False,
                'error': f'Ya existe un proveedor con el nombre "{nombre_comercial}"'
            }, status=400)
        
        # Crear el proveedor - el campo es 'activo' (masculino)
        proveedor = Proveedor.objects.create(
            nombre_comercial=nombre_comercial,
            contacto=contacto if contacto else None,
            telefono=telefono if telefono else None,
            email=email if email else None,
            activo=True  # ← Este es 'activo' (masculino)
        )
        
        return JsonResponse({
            'success': True,
            'id': str(proveedor.id),
            'nombre_comercial': proveedor.nombre_comercial,
            'contacto': proveedor.contacto,
            'telefono': proveedor.telefono,
            'email': proveedor.email,
            'message': 'Proveedor creado exitosamente'
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error en proveedor_crear_api: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
# ============================================
# VISTA API PROVEEDOR - SIN CAMPO CONTACTO
# ============================================
# REEMPLAZAR proveedor_crear_api en views.py

@ensure_csrf_cookie
@login_required
@require_http_methods(["POST"])
def proveedor_crear_api(request):
    """API para crear un proveedor desde el modal"""
    try:
        data = json.loads(request.body)
        
        # Obtener campos
        nombre_comercial = data.get('nombre_comercial', '').strip()
        razon_social = data.get('razon_social', '').strip()
        ruc_nit = data.get('ruc_nit', '').strip()
        direccion = data.get('direccion', '').strip()
        telefono = data.get('telefono', '').strip()
        email = data.get('email', '').strip()
        dias_credito = data.get('dias_credito', 0)
        limite_credito = data.get('limite_credito', 0.00)
        activo = data.get('activo', True)
        
        # Validaciones
        if not nombre_comercial:
            return JsonResponse({
                'success': False,
                'error': 'El nombre comercial es requerido'
            }, status=400)
        
        if not ruc_nit:
            return JsonResponse({
                'success': False,
                'error': 'El RUC/NIT es requerido'
            }, status=400)
        
        # Validar longitud de RUC/NIT
        if len(ruc_nit) < 10 or len(ruc_nit) > 20:
            return JsonResponse({
                'success': False,
                'error': 'El RUC/NIT debe tener entre 10 y 20 dígitos'
            }, status=400)
        
        # Verificar duplicados
        if Proveedor.objects.filter(nombre_comercial__iexact=nombre_comercial).exists():
            return JsonResponse({
                'success': False,
                'error': f'Ya existe un proveedor con el nombre "{nombre_comercial}"'
            }, status=400)
        
        if Proveedor.objects.filter(ruc_nit=ruc_nit).exists():
            return JsonResponse({
                'success': False,
                'error': f'Ya existe un proveedor con el RUC/NIT "{ruc_nit}"'
            }, status=400)
        
        # Crear el proveedor SIN el campo 'contacto'
        proveedor = Proveedor.objects.create(
            nombre_comercial=nombre_comercial,
            razon_social=razon_social if razon_social else '',
            ruc_nit=ruc_nit,
            direccion=direccion if direccion else '',
            telefono=telefono if telefono else '',
            email=email if email else '',
            dias_credito=int(dias_credito),
            limite_credito=Decimal(str(limite_credito)),
            activo=activo
        )
        
        return JsonResponse({
            'success': True,
            'id': str(proveedor.id),
            'nombre_comercial': proveedor.nombre_comercial,
            'razon_social': proveedor.razon_social,
            'ruc_nit': proveedor.ruc_nit,
            'direccion': proveedor.direccion,
            'telefono': proveedor.telefono,
            'email': proveedor.email,
            'dias_credito': proveedor.dias_credito,
            'limite_credito': float(proveedor.limite_credito),
            'activo': proveedor.activo,
            'message': 'Proveedor creado exitosamente'
        })
        
    except IntegrityError as e:
        error_msg = str(e).lower()
        if 'unique' in error_msg or 'ruc' in error_msg:
            return JsonResponse({
                'success': False,
                'error': 'Ya existe un proveedor con este RUC/NIT'
            }, status=400)
        return JsonResponse({
            'success': False,
            'error': f'Error de integridad en la base de datos'
        }, status=400)
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌ Error en proveedor_crear_api: {error_trace}")
        
        # Mostrar el error completo para debugging
        return JsonResponse({
            'success': False,
            'error': f'Error al crear el proveedor: {str(e)}'
        }, status=500)

# ============================================================================
# REPORTES Y DASHBOARDS
# ============================================================================

@ensure_csrf_cookie
@login_required
def finanzas_dashboard(request):
    """Dashboard general de finanzas con resumen de cuentas por cobrar y pagar"""
    
    from apps.financial_management.models import ReporteCuentasPorCobrar, ReporteCuentasPorPagar
    
    # Resúmenes generales
    resumen_cobrar = ReporteCuentasPorCobrar.resumen_general()
    resumen_pagar = ReporteCuentasPorPagar.resumen_general()
    
    # Antigüedad de saldos
    antiguedad_cobrar = ReporteCuentasPorCobrar.antiguedad_saldos()
    antiguedad_pagar = ReporteCuentasPorPagar.antiguedad_saldos()
    
    # Cuentas más urgentes (próximas a vencer)
    hoy = timezone.now().date()
    fecha_limite = hoy + timedelta(days=7)
    
    cuentas_cobrar_urgentes = CuentaPorCobrar.objects.filter(
        estado__in=['PENDIENTE', 'PARCIAL'],
        fecha_vencimiento__lte=fecha_limite,
        fecha_vencimiento__gte=hoy
    ).select_related('cliente').order_by('fecha_vencimiento')[:10]
    
    cuentas_pagar_urgentes = CuentaPorPagar.objects.filter(
        estado__in=['PENDIENTE', 'PARCIAL'],
        fecha_vencimiento__lte=fecha_limite,
        fecha_vencimiento__gte=hoy
    ).select_related('proveedor').order_by('fecha_vencimiento')[:10]
    
    context = {
        'resumen_cobrar': resumen_cobrar,
        'resumen_pagar': resumen_pagar,
        'antiguedad_cobrar': antiguedad_cobrar,
        'antiguedad_pagar': antiguedad_pagar,
        'cuentas_cobrar_urgentes': cuentas_cobrar_urgentes,
        'cuentas_pagar_urgentes': cuentas_pagar_urgentes,
    }
    
    return render(request, 'custom_admin/finanzas/dashboard.html', context)
@ensure_csrf_cookie
@auth_required
@require_http_methods(["GET"])
def api_cuentas_por_cobrar_list(request):
    """API para obtener listado de cuentas por cobrar (para uso en JavaScript)"""
    try:
        cuentas = CuentaPorCobrar.objects.select_related(
            'cliente', 'venta'
        ).all().order_by('-fecha_emision')
        
        # Aplicar filtros si existen
        estado = request.GET.get('estado', '')
        if estado:
            cuentas = cuentas.filter(estado=estado)
        
        cliente_id = request.GET.get('cliente_id', '')
        if cliente_id:
            cuentas = cuentas.filter(cliente_id=cliente_id)
        
        # Serializar las cuentas
        cuentas_data = []
        for cuenta in cuentas:
            cuentas_data.append({
                'id': str(cuenta.id),
                'numero_cuenta': cuenta.numero_cuenta,
                'cliente': cuenta.cliente.nombre or cuenta.cliente.razon_social,
                'monto_total': float(cuenta.monto_total),
                'saldo_pendiente': float(cuenta.saldo_pendiente),
                'estado': cuenta.estado,
                'fecha_emision': cuenta.fecha_emision.strftime('%Y-%m-%d'),
                'fecha_vencimiento': cuenta.fecha_vencimiento.strftime('%Y-%m-%d'),
            })
        
        return JsonResponse({
            'success': True,
            'cuentas': cuentas_data
        })
        
    except Exception as e:
        logger.error(f"Error al listar cuentas: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

# ============================================================================
# VIEWS PARA SYSTEM CONFIGURATION
# Agregar estas views al archivo apps/custom_admin/views.py
# ============================================================================
@ensure_csrf_cookie
@auth_required
def configuracion_sistema_view(request):
    """
    Vista principal de configuración del sistema
    Muestra todos los parámetros configurables
    """
    config = ConfiguracionSistema.get_config()
    
    context = {
        'config': config,
        'active_menu': 'system',
        'active_submenu': 'configuracion',
    }
    
    return render(request, 'custom_admin/system/configuracion.html', context)


@ensure_csrf_cookie
@auth_required
@require_http_methods(["POST"])
def configuracion_sistema_actualizar(request):
    """
    API para actualizar la configuración general del sistema
    """
    try:
        config = ConfiguracionSistema.get_config()
        usuario = get_authenticated_user(request)
        
        # Información General
        if 'nombre_empresa' in request.POST:
            config.nombre_empresa = request.POST.get('nombre_empresa', '').strip()
        if 'ruc_empresa' in request.POST:
            config.ruc_empresa = request.POST.get('ruc_empresa', '').strip()
        if 'direccion_empresa' in request.POST:
            config.direccion_empresa = request.POST.get('direccion_empresa', '').strip()
        if 'telefono_empresa' in request.POST:
            config.telefono_empresa = request.POST.get('telefono_empresa', '').strip()
        if 'email_empresa' in request.POST:
            config.email_empresa = request.POST.get('email_empresa', '').strip()
        if 'sitio_web' in request.POST:
            config.sitio_web = request.POST.get('sitio_web', '').strip()
        
        # Logo de la empresa
        if 'logo_empresa' in request.FILES:
            config.logo_empresa = request.FILES['logo_empresa']
        
        # Configuración de Inventario
        if 'prefijo_codigo_quintal' in request.POST:
            config.prefijo_codigo_quintal = request.POST.get('prefijo_codigo_quintal', '').strip()
        if 'prefijo_codigo_producto' in request.POST:
            config.prefijo_codigo_producto = request.POST.get('prefijo_codigo_producto', '').strip()
        if 'longitud_codigo_secuencial' in request.POST:
            config.longitud_codigo_secuencial = int(request.POST.get('longitud_codigo_secuencial', 5))
        if 'umbral_stock_critico_porcentaje' in request.POST:
            config.umbral_stock_critico_porcentaje = Decimal(request.POST.get('umbral_stock_critico_porcentaje', '10.00'))
        if 'umbral_stock_bajo_porcentaje' in request.POST:
            config.umbral_stock_bajo_porcentaje = Decimal(request.POST.get('umbral_stock_bajo_porcentaje', '25.00'))
        if 'stock_minimo_default' in request.POST:
            config.stock_minimo_default = int(request.POST.get('stock_minimo_default', 10))
        if 'dias_alerta_vencimiento' in request.POST:
            config.dias_alerta_vencimiento = int(request.POST.get('dias_alerta_vencimiento', 30))
        if 'peso_minimo_quintal_critico' in request.POST:
            config.peso_minimo_quintal_critico = Decimal(request.POST.get('peso_minimo_quintal_critico', '5.000'))
        
        # Configuración de Ventas
        if 'prefijo_numero_factura' in request.POST:
            config.prefijo_numero_factura = request.POST.get('prefijo_numero_factura', '').strip()
        if 'prefijo_numero_venta' in request.POST:
            config.prefijo_numero_venta = request.POST.get('prefijo_numero_venta', '').strip()
        if 'iva_default' in request.POST:
            config.iva_default = Decimal(request.POST.get('iva_default', '15.00'))
        if 'max_descuento_sin_autorizacion' in request.POST:
            config.max_descuento_sin_autorizacion = Decimal(request.POST.get('max_descuento_sin_autorizacion', '10.00'))
        if 'permitir_ventas_credito' in request.POST:
            config.permitir_ventas_credito = request.POST.get('permitir_ventas_credito') == 'true'
        if 'dias_credito_default' in request.POST:
            config.dias_credito_default = int(request.POST.get('dias_credito_default', 30))
        
        # Configuración Financiera
        if 'moneda' in request.POST:
            config.moneda = request.POST.get('moneda', '').strip()
        if 'simbolo_moneda' in request.POST:
            config.simbolo_moneda = request.POST.get('simbolo_moneda', '').strip()
        if 'decimales_moneda' in request.POST:
            config.decimales_moneda = int(request.POST.get('decimales_moneda', 2))
        if 'monto_inicial_caja' in request.POST:
            config.monto_inicial_caja = Decimal(request.POST.get('monto_inicial_caja', '100.00'))
        if 'monto_fondo_caja_chica' in request.POST:
            config.monto_fondo_caja_chica = Decimal(request.POST.get('monto_fondo_caja_chica', '50.00'))
        if 'alerta_diferencia_caja' in request.POST:
            config.alerta_diferencia_caja = Decimal(request.POST.get('alerta_diferencia_caja', '5.00'))
        
        # Notificaciones
        if 'notificaciones_email_activas' in request.POST:
            config.notificaciones_email_activas = request.POST.get('notificaciones_email_activas') == 'true'
        if 'email_notificaciones' in request.POST:
            config.email_notificaciones = request.POST.get('email_notificaciones', '').strip()
        if 'notificar_stock_bajo' in request.POST:
            config.notificar_stock_bajo = request.POST.get('notificar_stock_bajo') == 'true'
        if 'notificar_vencimientos' in request.POST:
            config.notificar_vencimientos = request.POST.get('notificar_vencimientos') == 'true'
        if 'notificar_cierre_caja' in request.POST:
            config.notificar_cierre_caja = request.POST.get('notificar_cierre_caja') == 'true'
        
        # Backups
        if 'backup_automatico_activo' in request.POST:
            config.backup_automatico_activo = request.POST.get('backup_automatico_activo') == 'true'
        if 'frecuencia_backup' in request.POST:
            config.frecuencia_backup = request.POST.get('frecuencia_backup', 'DIARIO')
        if 'dias_retencion_backup' in request.POST:
            config.dias_retencion_backup = int(request.POST.get('dias_retencion_backup', 30))
        
        # Mantenimiento y Sistema
        if 'modo_mantenimiento' in request.POST:
            config.modo_mantenimiento = request.POST.get('modo_mantenimiento') == 'true'
        if 'mensaje_mantenimiento' in request.POST:
            config.mensaje_mantenimiento = request.POST.get('mensaje_mantenimiento', '').strip()
        if 'timezone' in request.POST:
            config.timezone = request.POST.get('timezone', '').strip()
        if 'idioma_default' in request.POST:
            config.idioma_default = request.POST.get('idioma_default', 'es')
        
        config.actualizado_por = usuario
        config.save()
        
        # Registrar en log
        LogConfiguracion.objects.create(
            tabla='ConfiguracionSistema',
            registro_id='1',
            tipo_cambio='MODIFICACION',
            descripcion='Actualización de configuración general del sistema',
            usuario=usuario,
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Configuración actualizada exitosamente'
        })
        
    except Exception as e:
        logger.error(f"Error al actualizar configuración: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Error al actualizar configuración: {str(e)}'
        }, status=500)


# ============================================================================
# PARÁMETROS DEL SISTEMA
# ============================================================================

@ensure_csrf_cookie
@auth_required
def parametros_sistema_view(request):
    """
    Lista de parámetros del sistema con filtros
    """
    parametros = ParametroSistema.objects.all().order_by('categoria', 'modulo', 'clave')
    
    # Filtros
    search = request.GET.get('search', '')
    categoria = request.GET.get('categoria', '')
    modulo = request.GET.get('modulo', '')
    activo = request.GET.get('activo', '')
    
    if search:
        parametros = parametros.filter(
            Q(nombre_display__icontains=search) |
            Q(clave__icontains=search) |
            Q(descripcion__icontains=search)
        )
    
    if categoria:
        parametros = parametros.filter(categoria=categoria)
    
    if modulo:
        parametros = parametros.filter(modulo=modulo)
    
    if activo:
        parametros = parametros.filter(activo=(activo == 'true'))
    
    # Obtener categorías y módulos únicos para filtros
    categorias = ParametroSistema.objects.values_list('categoria', flat=True).distinct().order_by('categoria')
    modulos = ParametroSistema.objects.values_list('modulo', flat=True).distinct().order_by('modulo')
    
    # Paginación
    paginator = Paginator(parametros, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'parametros': page_obj,
        'page_obj': page_obj,
        'categorias': categorias,
        'modulos': modulos,
        'search': search,
        'categoria_selected': categoria,
        'modulo_selected': modulo,
        'activo_selected': activo,
        'active_menu': 'system',
        'active_submenu': 'parametros',
    }
    
    return render(request, 'custom_admin/system/parametros_list.html', context)


@ensure_csrf_cookie
@auth_required
@require_http_methods(["GET"])
def api_parametro_obtener(request, parametro_id):
    """
    API para obtener los detalles de un parámetro
    """
    try:
        parametro = get_object_or_404(ParametroSistema, pk=parametro_id)
        
        return JsonResponse({
            'success': True,
            'parametro': {
                'id': str(parametro.id),
                'modulo': parametro.modulo,
                'clave': parametro.clave,
                'nombre_display': parametro.nombre_display,
                'descripcion': parametro.descripcion,
                'categoria': parametro.categoria,
                'tipo_dato': parametro.tipo_dato,
                'valor': parametro.valor,
                'valor_default': parametro.valor_default,
                'validacion_regex': parametro.validacion_regex,
                'valor_minimo': parametro.valor_minimo,
                'valor_maximo': parametro.valor_maximo,
                'opciones_validas': parametro.opciones_validas,
                'activo': parametro.activo,
                'editable': parametro.editable,
                'requiere_reinicio': parametro.requiere_reinicio,
            }
        })
        
    except Exception as e:
        logger.error(f"Error al obtener parámetro: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@ensure_csrf_cookie
@auth_required
@require_http_methods(["POST"])
def api_parametro_actualizar(request, parametro_id):
    """
    API para actualizar un parámetro del sistema
    """
    try:
        parametro = get_object_or_404(ParametroSistema, pk=parametro_id)
        usuario = get_authenticated_user(request)
        
        if not parametro.editable:
            return JsonResponse({
                'success': False,
                'error': 'Este parámetro no es editable'
            }, status=400)
        
        data = json.loads(request.body)
        
        valor_anterior = parametro.valor
        nuevo_valor = data.get('valor', '')
        
        # Usar el helper para establecer el valor con el tipo correcto
        parametro.set_valor_typed(nuevo_valor)
        parametro.actualizado_por = usuario
        parametro.save()
        
        # Registrar cambio en log
        LogConfiguracion.objects.create(
            tabla='ParametroSistema',
            registro_id=str(parametro.id),
            tipo_cambio='MODIFICACION',
            campo_modificado='valor',
            valor_anterior=valor_anterior,
            valor_nuevo=parametro.valor,
            usuario=usuario,
            descripcion=f"Cambio de parámetro {parametro.modulo}.{parametro.clave}",
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Parámetro actualizado exitosamente',
            'parametro': {
                'id': str(parametro.id),
                'valor': parametro.valor,
                'requiere_reinicio': parametro.requiere_reinicio,
            }
        })
        
    except Exception as e:
        logger.error(f"Error al actualizar parámetro: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Error al actualizar parámetro: {str(e)}'
        }, status=500)


@ensure_csrf_cookie
@auth_required
@require_http_methods(["POST"])
def api_parametro_crear(request):
    """
    API para crear un nuevo parámetro del sistema
    """
    try:
        usuario = get_authenticated_user(request)
        data = json.loads(request.body)
        
        parametro = ParametroSistema.objects.create(
            modulo=data.get('modulo', '').strip(),
            clave=data.get('clave', '').strip(),
            nombre_display=data.get('nombre_display', '').strip(),
            descripcion=data.get('descripcion', '').strip(),
            categoria=data.get('categoria', 'SYSTEM'),
            tipo_dato=data.get('tipo_dato', 'STRING'),
            valor=data.get('valor', ''),
            valor_default=data.get('valor_default', ''),
            validacion_regex=data.get('validacion_regex', ''),
            valor_minimo=data.get('valor_minimo', ''),
            valor_maximo=data.get('valor_maximo', ''),
            opciones_validas=data.get('opciones_validas', []),
            activo=data.get('activo', True),
            editable=data.get('editable', True),
            requiere_reinicio=data.get('requiere_reinicio', False),
            actualizado_por=usuario
        )
        
        # Registrar en log
        LogConfiguracion.objects.create(
            tabla='ParametroSistema',
            registro_id=str(parametro.id),
            tipo_cambio='CREACION',
            descripcion=f"Creación de parámetro {parametro.modulo}.{parametro.clave}",
            usuario=usuario,
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Parámetro creado exitosamente',
            'parametro': {
                'id': str(parametro.id),
                'modulo': parametro.modulo,
                'clave': parametro.clave,
                'nombre_display': parametro.nombre_display,
            }
        })
        
    except Exception as e:
        logger.error(f"Error al crear parámetro: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Error al crear parámetro: {str(e)}'
        }, status=500)


@ensure_csrf_cookie
@auth_required
@require_http_methods(["DELETE"])
def api_parametro_eliminar(request, parametro_id):
    """
    API para eliminar un parámetro del sistema
    """
    try:
        parametro = get_object_or_404(ParametroSistema, pk=parametro_id)
        usuario = get_authenticated_user(request)
        
        if not parametro.editable:
            return JsonResponse({
                'success': False,
                'error': 'Este parámetro no se puede eliminar'
            }, status=400)
        
        # Registrar en log antes de eliminar
        LogConfiguracion.objects.create(
            tabla='ParametroSistema',
            registro_id=str(parametro.id),
            tipo_cambio='ELIMINACION',
            descripcion=f"Eliminación de parámetro {parametro.modulo}.{parametro.clave}",
            usuario=usuario,
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        parametro.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Parámetro eliminado exitosamente'
        })
        
    except Exception as e:
        logger.error(f"Error al eliminar parámetro: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================================================
# LOGS DE CONFIGURACIÓN
# ============================================================================

@ensure_csrf_cookie
@auth_required
def logs_configuracion_view(request):
    """
    Vista de logs de cambios en la configuración
    """
    logs = LogConfiguracion.objects.select_related('usuario').all().order_by('-fecha_cambio')
    
    # Filtros
    tabla = request.GET.get('tabla', '')
    tipo_cambio = request.GET.get('tipo_cambio', '')
    usuario_id = request.GET.get('usuario_id', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if tabla:
        logs = logs.filter(tabla=tabla)
    
    if tipo_cambio:
        logs = logs.filter(tipo_cambio=tipo_cambio)
    
    if usuario_id:
        logs = logs.filter(usuario_id=usuario_id)
    
    if fecha_desde:
        logs = logs.filter(fecha_cambio__date__gte=fecha_desde)
    
    if fecha_hasta:
        logs = logs.filter(fecha_cambio__date__lte=fecha_hasta)
    
    # Obtener valores únicos para filtros
    from apps.authentication.models import Usuario
    tablas = LogConfiguracion.objects.values_list('tabla', flat=True).distinct().order_by('tabla')
    usuarios = Usuario.objects.filter(
        id__in=LogConfiguracion.objects.values_list('usuario_id', flat=True).distinct()
    ).order_by('username')
    
    # Paginación
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'logs': page_obj,
        'page_obj': page_obj,
        'tablas': tablas,
        'usuarios': usuarios,
        'tabla_selected': tabla,
        'tipo_cambio_selected': tipo_cambio,
        'usuario_selected': usuario_id,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'active_menu': 'system',
        'active_submenu': 'logs',
    }
    
    return render(request, 'custom_admin/system/logs_list.html', context)


@ensure_csrf_cookie
@auth_required
@require_http_methods(["GET"])
def api_log_detalle(request, log_id):
    """
    API para obtener los detalles de un log
    """
    try:
        log = get_object_or_404(
            LogConfiguracion.objects.select_related('usuario'),
            pk=log_id
        )
        
        return JsonResponse({
            'success': True,
            'log': {
                'id': str(log.id),
                'tabla': log.tabla,
                'registro_id': log.registro_id,
                'tipo_cambio': log.tipo_cambio,
                'tipo_cambio_display': log.get_tipo_cambio_display(),
                'campo_modificado': log.campo_modificado,
                'valor_anterior': log.valor_anterior,
                'valor_nuevo': log.valor_nuevo,
                'descripcion': log.descripcion,
                'ip_address': log.ip_address,
                'usuario': log.usuario.get_full_name() if log.usuario else 'Sistema',
                'fecha_cambio': log.fecha_cambio.strftime('%Y-%m-%d %H:%M:%S'),
            }
        })
        
    except Exception as e:
        logger.error(f"Error al obtener log: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================================================
# HEALTH CHECK DEL SISTEMA
# ============================================================================

@ensure_csrf_cookie
@auth_required
def health_check_view(request):
    """
    Vista de health checks del sistema
    """
    checks = HealthCheck.objects.all().order_by('-fecha_check')[:50]
    
    # Último health check
    ultimo_check = checks.first() if checks else None
    
    # Estadísticas de los últimos 30 días
    from django.db.models import Count
    from datetime import timedelta
    
    fecha_inicio = timezone.now() - timedelta(days=30)
    
    estadisticas = HealthCheck.objects.filter(
        fecha_check__gte=fecha_inicio
    ).values('estado_general').annotate(
        total=Count('id')
    )
    
    context = {
        'checks': checks,
        'ultimo_check': ultimo_check,
        'estadisticas': estadisticas,
        'active_menu': 'system',
        'active_submenu': 'health',
    }
    
    return render(request, 'custom_admin/system/health_check.html', context)


@ensure_csrf_cookie
@auth_required
@require_http_methods(["POST"])
def api_ejecutar_health_check(request):
    """
    API para ejecutar un health check manual del sistema
    """
    try:
        import psutil
        from django.db import connection
        from django.core.cache import cache
        import time
        
        errores = []
        advertencias = []
        detalles = {}
        
        # Check Base de Datos
        try:
            start_time = time.time()
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
            tiempo_respuesta_db = int((time.time() - start_time) * 1000)
            base_datos_ok = True
            detalles['base_datos'] = 'Conectado correctamente'
        except Exception as e:
            tiempo_respuesta_db = None
            base_datos_ok = False
            errores.append(f"Error en base de datos: {str(e)}")
        
        # Check Redis/Cache
        try:
            cache.set('health_check_test', 'ok', 10)
            test_value = cache.get('health_check_test')
            redis_ok = test_value == 'ok'
            detalles['redis'] = 'Funcionando correctamente' if redis_ok else 'No responde'
        except Exception as e:
            redis_ok = False
            advertencias.append(f"Redis no disponible: {str(e)}")
        
        # Check Disco
        try:
            disk = psutil.disk_usage('/')
            espacio_libre_gb = disk.free / (1024**3)
            disco_ok = espacio_libre_gb > 5  # Al menos 5GB libres
            detalles['disco'] = f"{espacio_libre_gb:.2f} GB libres de {disk.total / (1024**3):.2f} GB"
            
            if not disco_ok:
                advertencias.append(f"Poco espacio en disco: {espacio_libre_gb:.2f} GB")
        except Exception as e:
            espacio_libre_gb = None
            disco_ok = False
            advertencias.append(f"No se pudo verificar espacio en disco: {str(e)}")
        
        # Check Memoria
        try:
            memory = psutil.virtual_memory()
            uso_memoria = memory.percent
            memoria_ok = uso_memoria < 90
            detalles['memoria'] = f"Uso: {uso_memoria:.2f}%"
            
            if not memoria_ok:
                advertencias.append(f"Uso alto de memoria: {uso_memoria:.2f}%")
        except Exception as e:
            uso_memoria = None
            memoria_ok = False
            advertencias.append(f"No se pudo verificar memoria: {str(e)}")
        
        # Check Celery (opcional)
        try:
            from celery.app.control import Inspect
            from apps.commercebox.celery import app
            
            inspect = Inspect(app=app)
            active_workers = inspect.active()
            celery_ok = active_workers is not None and len(active_workers) > 0
            detalles['celery'] = f"{len(active_workers) if active_workers else 0} workers activos"
        except Exception as e:
            celery_ok = False
            detalles['celery'] = 'No configurado o no disponible'
        
        # Determinar estado general
        if errores:
            estado_general = 'CRITICO'
        elif advertencias:
            estado_general = 'ADVERTENCIA'
        else:
            estado_general = 'SALUDABLE'
        
        # Crear registro de health check
        health_check = HealthCheck.objects.create(
            estado_general=estado_general,
            base_datos_ok=base_datos_ok,
            redis_ok=redis_ok,
            celery_ok=celery_ok,
            disco_ok=disco_ok,
            memoria_ok=memoria_ok,
            espacio_disco_libre_gb=espacio_libre_gb,
            uso_memoria_porcentaje=uso_memoria,
            tiempo_respuesta_db_ms=tiempo_respuesta_db,
            detalles=detalles,
            errores=errores,
            advertencias=advertencias
        )
        
        return JsonResponse({
            'success': True,
            'health_check': {
                'id': str(health_check.id),
                'estado_general': estado_general,
                'estado_general_display': health_check.get_estado_general_display(),
                'base_datos_ok': base_datos_ok,
                'redis_ok': redis_ok,
                'celery_ok': celery_ok,
                'disco_ok': disco_ok,
                'memoria_ok': memoria_ok,
                'detalles': detalles,
                'errores': errores,
                'advertencias': advertencias,
                'fecha_check': health_check.fecha_check.strftime('%Y-%m-%d %H:%M:%S'),
            }
        })
        
    except Exception as e:
        logger.error(f"Error al ejecutar health check: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Error al ejecutar health check: {str(e)}'
        }, status=500)


# ============================================================================
# DASHBOARD DE CONFIGURACIÓN
# ============================================================================

@ensure_csrf_cookie
@auth_required
def system_dashboard_view(request):
    """
    Dashboard principal del módulo de configuración del sistema
    """
    config = ConfiguracionSistema.get_config()
    
    # Estadísticas de parámetros
    total_parametros = ParametroSistema.objects.count()
    parametros_activos = ParametroSistema.objects.filter(activo=True).count()
    parametros_por_categoria = ParametroSistema.objects.values('categoria').annotate(
        total=Count('id')
    ).order_by('-total')
    
    # Últimos logs
    ultimos_logs = LogConfiguracion.objects.select_related('usuario').order_by('-fecha_cambio')[:10]
    
    # Último health check
    ultimo_health_check = HealthCheck.objects.order_by('-fecha_check').first()
    
    # Alertas del sistema
    alertas = []
    
    # Verificar si hay parámetros que requieren reinicio
    parametros_requieren_reinicio = ParametroSistema.objects.filter(
        requiere_reinicio=True,
        activo=True
    ).exists()
    
    if parametros_requieren_reinicio:
        alertas.append({
            'tipo': 'warning',
            'mensaje': 'Hay parámetros que requieren reiniciar el sistema para aplicar cambios'
        })
    
    # Verificar modo mantenimiento
    if config.modo_mantenimiento:
        alertas.append({
            'tipo': 'info',
            'mensaje': 'El sistema está en modo mantenimiento'
        })
    
    # Verificar health check
    if ultimo_health_check:
        if ultimo_health_check.estado_general == 'CRITICO':
            alertas.append({
                'tipo': 'danger',
                'mensaje': 'El último health check detectó problemas críticos'
            })
        elif ultimo_health_check.estado_general == 'ADVERTENCIA':
            alertas.append({
                'tipo': 'warning',
                'mensaje': 'El último health check detectó advertencias'
            })
    
    context = {
        'config': config,
        'total_parametros': total_parametros,
        'parametros_activos': parametros_activos,
        'parametros_por_categoria': parametros_por_categoria,
        'ultimos_logs': ultimos_logs,
        'ultimo_health_check': ultimo_health_check,
        'alertas': alertas,
        'active_menu': 'system',
        'active_submenu': 'dashboard',
    }
    
    return render(request, 'custom_admin/system/dashboard.html', context)
